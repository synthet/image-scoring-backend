import sqlite3
import json
import os
import re
import datetime
import logging
import time
import threading
from collections import deque
from pathlib import Path
import traceback
import queue

from modules import config
from modules.events import event_manager
try:
    from modules import db_postgres
except ImportError:
    db_postgres = None
try:
    from firebird.driver import connect, driver_config, create_database
except ImportError:
    # Fallback/Mock for linting if package missing
    connect = None
    create_database = None 

import shutil

logger = logging.getLogger(__name__)
DEBUG_DB_CONNECTION = os.environ.get("DEBUG_DB_CONNECTION", "").lower() in ("1", "true", "yes")


JOB_TERMINAL_STATES = {"completed", "failed", "canceled", "cancelled", "interrupted"}
JOB_ALLOWED_TRANSITIONS = {
    "pending": {"queued", "running", "canceled", "cancelled", "interrupted"},
    "queued": {"running", "paused", "cancel_requested", "canceled", "cancelled", "interrupted", "restarting"},
    "running": {"paused", "completed", "failed", "cancel_requested", "canceled", "cancelled", "interrupted", "restarting"},
    "paused": {"queued", "running", "cancel_requested", "canceled", "cancelled", "interrupted", "restarting", "completed", "failed"},
    "cancel_requested": {"canceled", "cancelled", "failed", "interrupted"},
    "restarting": {"queued", "running", "failed", "interrupted", "canceled", "cancelled"},
    "interrupted": {"queued", "running", "canceled", "cancelled", "restarting"},
    "completed": set(),
    "failed": set(),
    "canceled": set(),
    "cancelled": set(),
}


_PIPELINE_TELEMETRY_LOCK = threading.Lock()
_PIPELINE_TELEMETRY_SEQ = 0
_PIPELINE_TELEMETRY_EVENTS = deque(maxlen=3000)


class RowWrapper:
    """
    Wraps a Firebird result row to provide both tuple-like (index) 
    and dict-like (column name) access, mimicking sqlite3.Row.
    """
    def __init__(self, cols, values):
        self._cols = cols
        self._values = values
        self._map = dict(zip(cols, values))
    
    def __getitem__(self, key):
        if isinstance(key, int):
            return self._values[key]
        return self._map.get(key.lower())
    
    def get(self, key, default=None):
        """Dict-like .get() for compatibility with code expecting dict access."""
        if isinstance(key, int):
            try:
                return self._values[key]
            except IndexError:
                return default
        return self._map.get(key.lower(), default)
    
    def keys(self):
        return self._map.keys()
        
    def __iter__(self):
        # Yield (key, value) pairs to support dict(row) correctly
        return iter(self._map.items())

    def to_dict(self, include_binary=False, exclude_keys=None):
        """
        Returns a JSON-serializable dictionary.
        Safe for FastAPI/JSON serialization by handling bytes and datetimes.
        """
        import base64
        d = {}
        exclude = exclude_keys if exclude_keys else set()
        for k, v in self._map.items():
            if k in exclude:
                continue
            if isinstance(v, bytes):
                if include_binary:
                    d[k] = base64.b64encode(v).decode('utf-8')
                else:
                    # Skip raw binary data (like embeddings) that crash JSON serialization
                    continue
            elif isinstance(v, (datetime.datetime, datetime.date)):
                d[k] = v.isoformat()
            else:
                d[k] = v
        return d


# NOTE: FirebirdCursorProxy removed — Firebird has been decommissioned (2026-03).


class PostgresCursorProxy:
    """Proxies a PostgreSQL cursor to provide sqlite/Firebird-style compatibility."""
    def __init__(self, pg_cur):
        self._cur = pg_cur

    def execute(self, query, params=None):
        query = self._translate_query(query)
        if params:
            return self._cur.execute(query, params)
        return self._cur.execute(query)

    def executemany(self, query, params):
        query = self._translate_query(query)
        return self._cur.executemany(query, params)

    def fetchone(self):
        row = self._cur.fetchone()
        if row is None:
            return None
        col_names = self._column_names()
        return RowWrapper(col_names, row)

    def fetchall(self):
        rows = self._cur.fetchall()
        if not rows:
            return []
        col_names = self._column_names()
        return [RowWrapper(col_names, r) for r in rows]

    def _column_names(self):
        names = []
        for d in self._cur.description or []:
            if hasattr(d, "name"):
                names.append(str(d.name).lower())
            else:
                names.append(str(d[0]).lower())
        return names

    def _translate_query(self, query: str) -> str:
        return _translate_fb_to_pg(query)

    def __getattr__(self, name):
        return getattr(self._cur, name)


# NOTE: Dual-write infrastructure removed — Firebird has been decommissioned (2026-03).
# The _translate_fb_to_pg() function below is still used by PostgresCursorProxy and
# the PostgresConnector to translate legacy Firebird-dialect SQL to PostgreSQL.

# ---------------------------------------------------------------------------
# Firebird → PostgreSQL SQL translation helpers
# ---------------------------------------------------------------------------

# UPDATE OR INSERT INTO t (cols) VALUES (?) MATCHING (col) [RETURNING col]
# UPDATE OR INSERT INTO t (cols) VALUES (...) MATCHING (col) [RETURNING col]
# Uses a more non-greedy approach for parens to avoid over-matching, 
# though truly nested parens still require a parser.
_FB_UPSERT_RE = re.compile(
    r'UPDATE\s+OR\s+INSERT\s+INTO\s+(\w+)\s*'   # 1: table name
    r'\((.*?)\)\s*'                              # 2: (col_list)
    r'VALUES\s*\((.*?)\)\s*'                    # 3: VALUES (val_list)
    r'MATCHING\s*\((.*?)\)'                     # 4: MATCHING (match_cols)
    r'((?:\s+RETURNING\s+\w+)?)',               # 5: optional RETURNING
    re.IGNORECASE | re.DOTALL,
)

# DATEDIFF(UNIT FROM col1 TO col2)
_FB_DATEDIFF_RE = re.compile(
    r'DATEDIFF\s*\(\s*(SECOND|MINUTE|HOUR|DAY)\s+FROM\s+([\w.]+)\s+TO\s+([\w.]+)\s*\)',
    re.IGNORECASE,
)

# SELECT FIRST n …  (Firebird top-level row limit)
_FB_FIRST_RE = re.compile(
    r'\bSELECT\s+FIRST\s+(\d+)\s+',
    re.IGNORECASE,
)


def _translate_fb_to_pg(query: str) -> str:
    """Translate Firebird-specific SQL to PostgreSQL for dual-write and read routing.

    Transforms:
    1. ``UPDATE OR INSERT INTO t (cols) VALUES (?) MATCHING (col) [RETURNING c]``
       → ``INSERT INTO t (cols) VALUES (%s) ON CONFLICT (col) DO UPDATE SET … [RETURNING c]``
    2. ``SELECT FIRST n …``
       → ``SELECT … LIMIT n``  (top-level queries only; subquery variant deferred to Phase 3)
    3. ``DATEDIFF(SECOND FROM a TO b)``
       → ``EXTRACT(EPOCH FROM (b - a))::INTEGER``
    4. ``?`` → ``%s``  (placeholder style, outside string literals)
    """
    # 1 – Firebird upsert → PostgreSQL ON CONFLICT
    def _upsert_replace(m: re.Match) -> str:
        table     = m.group(1).strip()
        cols_raw  = m.group(2)
        vals_raw  = m.group(3).strip()
        match_raw = m.group(4)
        returning = (m.group(5) or '').strip()

        cols      = [c.strip() for c in cols_raw.split(',')]
        match_set = {c.strip().lower() for c in match_raw.split(',')}
        set_parts = [f"{c} = EXCLUDED.{c}" for c in cols if c.strip().lower() not in match_set]

        pg = (
            f"INSERT INTO {table} ({', '.join(cols)}) VALUES ({vals_raw})"
            f" ON CONFLICT ({match_raw.strip()}) DO UPDATE SET {', '.join(set_parts)}"
        )
        return f"{pg} {returning}".rstrip() if returning else pg

    query = _FB_UPSERT_RE.sub(_upsert_replace, query)

    # 2 – SELECT FIRST n (top-level only)
    m = _FB_FIRST_RE.search(query)
    if m:
        n = m.group(1)
        query = _FB_FIRST_RE.sub('SELECT ', query, count=1)
        query = query.rstrip().rstrip(';') + f' LIMIT {n}'

    # 3 – DATEDIFF
    def _datediff_replace(m2: re.Match) -> str:
        unit = m2.group(1).upper()
        start = m2.group(2)
        end = m2.group(3)
        if unit == 'SECOND':
            return f'EXTRACT(EPOCH FROM ({end} - {start}))::INTEGER'
        elif unit == 'MINUTE':
            return f'(EXTRACT(EPOCH FROM ({end} - {start})) / 60)::INTEGER'
        elif unit == 'HOUR':
            return f'(EXTRACT(EPOCH FROM ({end} - {start})) / 3600)::INTEGER'
        elif unit == 'DAY':
            return f'EXTRACT(DAY FROM ({end} - {start}))::INTEGER'
        return m2.group(0)

    query = _FB_DATEDIFF_RE.sub(_datediff_replace, query)

    # 3b – Firebird "ROWS ?" row-limit → PostgreSQL "LIMIT ?"
    query = re.sub(r'\bROWS\s+\?', 'LIMIT ?', query, flags=re.IGNORECASE)

    # 3b.5 – Firebird "OFFSET ? ROWS FETCH NEXT ? ROWS ONLY" → PostgreSQL "OFFSET ? LIMIT ?"
    #         Preserves parameter order (offset first, then page_size).
    query = re.sub(
        r'\bOFFSET\s+(\S+)\s+ROWS\s+FETCH\s+NEXT\s+(\S+)\s+ROWS\s+ONLY\b',
        r'OFFSET \1 LIMIT \2',
        query,
        flags=re.IGNORECASE,
    )

    # 3c – Firebird "FETCH FIRST n ROWS ONLY" → PostgreSQL "LIMIT n"
    #      Handles both ? param and inlined numeric literals (e.g. FETCH FIRST 100 ROWS ONLY)
    query = re.sub(r'\bFETCH\s+FIRST\s+(\S+)\s+ROWS\s+ONLY\b', r'LIMIT \1', query, flags=re.IGNORECASE)

    # 3d – Firebird RAND() → PostgreSQL RANDOM()
    query = re.sub(r'\bRAND\s*\(\s*\)', 'RANDOM()', query, flags=re.IGNORECASE)

    # 3e – Firebird LIST(expr, sep) → PostgreSQL STRING_AGG(expr, sep)
    query = re.sub(r'\bLIST\s*\(', 'STRING_AGG(', query, flags=re.IGNORECASE)

    # 3f – Function name normalization (Firebird aliases → standard SQL)
    query = query.replace('substr(', 'substring(')
    query = query.replace('length(', 'char_length(')

    # 4 – ? → %s (skip content inside single-quoted string literals)
    parts = query.split("'")
    for i in range(0, len(parts), 2):
        parts[i] = parts[i].replace('?', '%s')
    return "'".join(parts)


def _count_placeholders_firebird_style(sql: str) -> int:
    """Count ``?`` placeholders outside single-quoted string literals."""
    parts = sql.split("'")
    return sum(parts[i].count("?") for i in range(0, len(parts), 2))


def validate_readonly_sql_for_api(query: str) -> str | None:
    """Return an error message if ``query`` is not allowed for /api/db/query; else None."""
    text = (query or "").strip()
    if not text:
        return "Empty query"
    upper = text.upper()
    if not (upper.startswith("SELECT") or upper.startswith("WITH")):
        return "Only read-only SELECT or WITH...SELECT queries are allowed"
    dangerous_patterns = [
        r"\bDROP\b",
        r"\bDELETE\b",
        r"\bINSERT\b",
        r"\bUPDATE\b",
        r"\bALTER\b",
        r"\bCREATE\b",
        r"\bTRUNCATE\b",
        r"\bGRANT\b",
        r"\bREVOKE\b",
        r"--",
        r";--",
    ]
    for pattern in dangerous_patterns:
        if re.search(pattern, upper):
            return f"Query contains forbidden pattern: {pattern}"
    return None


def execute_readonly_sql_for_api(
    sql: str,
    params: list | None = None,
    *,
    max_rows: int = 5000,
) -> list[dict]:
    """Run a validated read-only query for the Electron HTTP DB bridge (legacy ``?`` params).

    Applies ``_translate_fb_to_pg`` so clients can send legacy-dialect SQL.
    """
    err = validate_readonly_sql_for_api(sql)
    if err:
        raise ValueError(err)
    params = list(params) if params is not None else []
    n_mark = _count_placeholders_firebird_style(sql)
    if n_mark != len(params):
        raise ValueError(
            f"Parameter count mismatch: SQL has {n_mark} placeholders, got {len(params)} bound values"
        )
    max_rows = max(1, min(int(max_rows), 50_000))

    if _get_db_engine() == "postgres":
        if not db_postgres:
            raise RuntimeError("database.engine is postgres but db_postgres is unavailable")
        import psycopg2.extras

        pg_sql = _translate_fb_to_pg(sql)
        with db_postgres.PGConnectionManager() as conn:
            with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
                cur.execute(pg_sql, tuple(params) if params else None)
                rows = cur.fetchmany(max_rows)
                return [dict(r) for r in rows]

    with connection() as conn:
        c = conn.cursor()
        try:
            from firebird.driver import tpb, Isolation, TraAccessMode

            ro_tpb = tpb(
                isolation=Isolation.READ_COMMITTED_RECORD_VERSION,
                access_mode=TraAccessMode.READ,
            )
            conn.begin(tpb=ro_tpb)
        except Exception:
            pass
        if params:
            c.execute(sql, tuple(params))
        else:
            c.execute(sql)
        rows = c.fetchmany(max_rows)
        columns = [d[0] for d in c.description] if c.description else []
        return [dict(zip(columns, row)) for row in rows]


def validate_write_sql_for_api(query: str) -> str | None:
    """Return an error message if ``query`` is not allowed for mutating /api/db/query; else None."""
    text = (query or "").strip()
    if not text:
        return "Empty query"
    upper = text.upper()
    if not (
        upper.startswith("INSERT")
        or upper.startswith("UPDATE")
        or upper.startswith("DELETE")
    ):
        return "Write queries must start with INSERT, UPDATE, or DELETE (Firebird UPDATE OR INSERT is allowed)"
    dangerous_patterns = [
        r"\bDROP\b",
        r"\bALTER\b",
        r"\bCREATE\b",
        r"\bTRUNCATE\b",
        r"\bGRANT\b",
        r"\bREVOKE\b",
        r";--",
    ]
    for pattern in dangerous_patterns:
        if re.search(pattern, upper):
            return f"Query contains forbidden pattern: {pattern}"
    return None


def execute_write_sql_for_api(sql: str, params: list | None = None) -> list[dict]:
    """Run INSERT/UPDATE/DELETE for the Electron HTTP DB bridge (legacy ``?`` params).

    Commits on success. Returns result rows when the statement produces a cursor description
    (e.g. INSERT ... RETURNING), else an empty list.
    """
    err = validate_write_sql_for_api(sql)
    if err:
        raise ValueError(err)
    params = list(params) if params is not None else []
    n_mark = _count_placeholders_firebird_style(sql)
    if n_mark != len(params):
        raise ValueError(
            f"Parameter count mismatch: SQL has {n_mark} placeholders, got {len(params)} bound values"
        )

    if _get_db_engine() == "postgres":
        if not db_postgres:
            raise RuntimeError("database.engine is postgres but db_postgres is unavailable")
        import psycopg2.extras

        pg_sql = _translate_fb_to_pg(sql)
        with db_postgres.PGConnectionManager(commit=True) as conn:
            with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
                cur.execute(pg_sql, tuple(params) if params else None)
                if cur.description:
                    rows = cur.fetchall()
                    return [dict(r) for r in rows]
                return []

    conn = get_db()
    try:
        c = conn.cursor()
        if params:
            c.execute(sql, tuple(params))
        else:
            c.execute(sql)
        if c.description:
            rows = c.fetchall()
            columns = [d[0] for d in c.description]
            out = [
                {str(k).lower(): v for k, v in zip(columns, row)}
                for row in rows
            ]
            conn.commit()
            return out
        conn.commit()
        return []
    except Exception:
        try:
            conn.rollback()
        except Exception:
            pass
        raise
    finally:
        try:
            conn.close()
        except Exception:
            pass


def get_dual_write_stats() -> dict:
    """Legacy stub — dual-write has been removed (Firebird decommissioned 2026-03)."""
    return {"queued": 0, "success": 0, "fail": 0, "queue_depth": 0, "enabled": False}

# NOTE: FirebirdConnectionProxy removed — Firebird has been decommissioned (2026-03).


class PostgresConnectionProxy:
    """Proxies a PostgreSQL connection to provide sqlite3-compatible interface."""
    def __init__(self, pg_conn):
        self._conn = pg_conn
        self.row_factory = sqlite3.Row

    def cursor(self):
        return PostgresCursorProxy(self._conn.cursor())

    def commit(self):
        self._conn.commit()

    def rollback(self):
        self._conn.rollback()

    def close(self):
        if self._conn:
            if db_postgres and hasattr(db_postgres, "release_pg_connection"):
                db_postgres.release_pg_connection(self._conn)
            else:
                self._conn.close()

    def __getattr__(self, name):
        return getattr(self._conn, name)


def record_pipeline_event(event_type, message, *, workflow_run=None, stage_run=None,
                          step_run=None, category=None, severity="info",
                          metadata=None, critical=False, noisy=False, source="db"):
    """Append a normalized pipeline telemetry event to an in-memory ring buffer."""
    global _PIPELINE_TELEMETRY_SEQ
    event = {
        "timestamp": datetime.datetime.now().isoformat(timespec="seconds"),
        "event_type": event_type or "log",
        "message": message or "",
        "workflow_run": workflow_run,
        "stage_run": stage_run,
        "step_run": step_run,
        "category": category or "pipeline",
        "severity": severity or "info",
        "metadata": metadata or {},
        "critical": bool(critical),
        "noisy": bool(noisy),
        "source": source,
    }
    with _PIPELINE_TELEMETRY_LOCK:
        _PIPELINE_TELEMETRY_SEQ += 1
        event["seq"] = _PIPELINE_TELEMETRY_SEQ
        _PIPELINE_TELEMETRY_EVENTS.append(event)
    return event["seq"]


def get_pipeline_events(since_seq=0, limit=250):
    """Return pipeline telemetry events with sequence greater than `since_seq`."""
    try:
        since = int(since_seq or 0)
    except (TypeError, ValueError):
        since = 0
    lim = max(1, min(int(limit or 250), 1000))
    with _PIPELINE_TELEMETRY_LOCK:
        rows = [evt.copy() for evt in _PIPELINE_TELEMETRY_EVENTS if evt.get("seq", 0) > since]
        if len(rows) > lim:
            rows = rows[-lim:]
        latest_seq = _PIPELINE_TELEMETRY_SEQ
    return {"events": rows, "latest_seq": latest_seq}


def generate_image_uuid(metadata: dict | None) -> str:
    """
    Generate a UUID for an image.

    Priority:
    1. Existing ImageUniqueID already embedded in EXIF/XMP metadata.
    2. Deterministic UUID: SHA-256 hash of CreateDate + SubSecTimeOriginal +
       Model + LensModel + ShutterCount (stable across machines and path changes).
    3. Fallback: random uuid4().
    """
    import uuid
    import hashlib

    if metadata and isinstance(metadata, dict):
        # 1. Honour existing embedded UUID
        existing = metadata.get("ImageUniqueID") or metadata.get("xmp:ImageUniqueID")
        if existing and isinstance(existing, str) and existing.strip():
            return existing.strip()

        # 2. Deterministic UUID from EXIF identity fields
        create_date = (
            metadata.get("CreateDate")
            or metadata.get("DateTimeOriginal")
            or ""
        )
        sub_sec = metadata.get("SubSecTimeOriginal") or metadata.get("SubSecTime") or ""
        model = metadata.get("Model") or ""
        lens_model = metadata.get("LensModel") or ""
        shutter_count = str(metadata.get("ShutterCount") or "")

        if create_date:  # Only use deterministic when we have at least a date
            fingerprint = f"{create_date}|{sub_sec}|{model}|{lens_model}|{shutter_count}"
            digest = hashlib.sha256(fingerprint.encode("utf-8")).hexdigest()
            # Format as 8-4-4-4-12 UUID
            return f"{digest[0:8]}-{digest[8:12]}-{digest[12:16]}-{digest[16:20]}-{digest[20:32]}"

    # 3. Random fallback
    return str(uuid.uuid4())

# --- Sort validation whitelist (SQL injection prevention) ---
ALLOWED_SORT_COLUMNS = {
    "score", "score_general", "score_aesthetic", "score_technical",
    "score_spaq", "score_ava", "score_koniq", "score_paq2piq", "score_liqe",
    "rating", "file_name", "file_path", "created_at", "updated_at",
    "id", "label", "folder_id",
    "date_time_original", "make", "model", "lens_model", "iso",
}
ALLOWED_SORT_ORDERS = {"asc", "desc"}

def _validate_sort(sort_by: str, order: str) -> tuple:
    """Validate and sanitize ORDER BY parameters to prevent SQL injection."""
    if sort_by not in ALLOWED_SORT_COLUMNS:
        sort_by = "score_general"
    if order.lower() not in ALLOWED_SORT_ORDERS:
        order = "desc"
    return sort_by, order.upper()


def _add_keyword_filter(conditions, params, keyword_filter, table_ref="images"):
    """Append a keyword EXISTS filter using normalized keyword tables."""
    if keyword_filter and keyword_filter.strip():
        conditions.append(
            f"EXISTS (SELECT 1 FROM image_keywords ik "
            f"JOIN keywords_dim kd ON ik.keyword_id = kd.keyword_id "
            f"WHERE ik.image_id = {table_ref}.id "
            f"AND kd.keyword_norm LIKE ?)"
        )
        params.append(f"%{keyword_filter.strip().lower()}%")


DB_CONFIG = config.get_config_section('database')
DB_FILE = DB_CONFIG.get('filename', "scoring_history.fdb")
DB_USER = str(DB_CONFIG.get('user', "sysdba") or "sysdba")
DB_PASS = str(
    os.environ.get("FIREBIRD_PASSWORD")
    or DB_CONFIG.get('password')
    or "masterkey"
)
if DB_PASS == "masterkey":
    logger.warning("Using default Firebird password 'masterkey' — set FIREBIRD_PASSWORD env var for production")

import sys
if "pytest" in sys.modules or os.environ.get("PYTEST_CURRENT_TEST"):
    # Tests must use only scoring_history_test.fdb — never production (e.g. SCORING_HISTORY.FDB).
    DB_FILE = "scoring_history_test.fdb"
    logger.info("Test environment detected: using DB_FILE=%s only", DB_FILE)

_PROJECT_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DB_PATH = os.path.join(_PROJECT_ROOT, DB_FILE)

def _to_win_path(p_str: str) -> str:
    """Convert a WSL /mnt/ path to a Windows drive path."""
    if p_str.startswith("/mnt/"):
        parts = p_str.split('/')
        if len(parts) >= 3 and parts[1] == 'mnt':
            drive = parts[2]
            rest = "\\".join(parts[3:])
            return f"{drive}:\\{rest}"
    return p_str.replace("/", "\\")


def _is_wsl() -> bool:
    # Conservative detection: WSL exports these env vars.
    return os.name != "nt" and bool(os.environ.get("WSL_DISTRO_NAME") or os.environ.get("WSL_INTEROP"))


def _resolve_firebird_client_library() -> str | None:
    """
    Choose a Firebird client library compatible with the current OS.

    - Windows: prefer repo-bundled `Firebird/fbclient.dll`
    - Linux/WSL: prefer repo-extracted `FirebirdLinux/.../libfbclient.so`, else fall back to
      `libfbclient.so` / `libfbclient.so.2` via the dynamic loader.

    Users can override with env var `FIREBIRD_CLIENT_LIBRARY`.
    """
    override = os.environ.get("FIREBIRD_CLIENT_LIBRARY") or os.environ.get("FB_CLIENT_LIBRARY")
    if override:
        return override

    if os.name == "nt":
        win_dll = os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", "Firebird", "fbclient.dll")
        if os.path.exists(win_dll):
            return win_dll
        return None

    # Linux / WSL
    base_root = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))  # project root
    repo_linux_lib = os.path.join(
        base_root,
        "FirebirdLinux",
        "Firebird-5.0.0.1306-0-linux-x64",
        "opt",
        "firebird",
        "lib",
        "libfbclient.so",
    )
    if os.path.exists(repo_linux_lib):
        return repo_linux_lib

    # Let the dynamic loader resolve these if installed system-wide / via LD_LIBRARY_PATH.
    # Note: many distros expose the SONAME as libfbclient.so.2 even for newer versions.
    return "libfbclient.so"


FB_CLIENT_LIBRARY = _resolve_firebird_client_library()

# Flags to prevent log spam
_logged_wsl_info = False
_logged_dsn = False

# Cached WSL host IP so all threads use the same value (avoids JobDispatcher using
# resolv.conf nameserver 10.255.255.254 when ip route fails in a thread).
_cached_wsl_host_ip = None

# Serialize connect() to avoid "Invalid clumplet buffer structure" when multiple
# threads call get_db() concurrently (driver_config + DPB build are not thread-safe).
_DB_CONNECT_LOCK = threading.Lock()
_IP_RESOLVE_LOCK = threading.Lock()
_cached_host_ip = None

# Configure driver if possible
if connect and FB_CLIENT_LIBRARY:
    try:
        # driver_config might be available if imported
        from firebird.driver import driver_config
        # Fix: client_library is a top-level config option in this driver version
        if hasattr(driver_config, 'fb_client_library'):
             driver_config.fb_client_library.value = FB_CLIENT_LIBRARY
    except Exception as e:
        logger.debug("Firebird driver config setup failed: %s", e)


class FirebirdConnectionFailed(RuntimeError):
    """Raised when Firebird connect() fails; message is intended for logs and API surfaces."""


def _humanize_firebird_connect_error(exc: BaseException, ctx: dict) -> str:
    """Turn firebird.driver errors into short, actionable text."""
    raw = str(exc).strip()
    raw_one = " ".join(raw.splitlines())
    low = raw_one.lower()

    dsn = ctx.get("dsn")
    host_ip = ctx.get("host_ip")
    is_docker = bool(ctx.get("is_docker"))
    use_local = bool(ctx.get("use_local_path"))

    if use_local:
        return (
            "Cannot open the local Firebird database file (embedded mode). "
            "Check that the file exists, paths are correct, and no other process holds a lock. "
            f"DSN: {dsn!r}. Details: {raw_one}"
        )

    if "authentication" in low or ("password" in low and "fail" in low):
        return (
            "Firebird rejected the database login. "
            "Check FIREBIRD_USER / FIREBIRD_PASSWORD (or sysdba credentials). "
            f"Details: {raw_one}"
        )

    if is_docker or "host.docker.internal" in raw_one:
        return (
            "Docker could not reach Firebird on the host. "
            "Start Firebird on Windows and ensure TCP port 3050 is reachable from the container. "
            "On Linux Docker, add extra_hosts (e.g. host.docker.internal:host-gateway) or set FIREBIRD_HOST "
            "to a reachable IP. Set FIREBIRD_WIN_DB_PATH to the Windows path of your .fdb "
            "(e.g. D:/Projects/.../SCORING_HISTORY.FDB). "
            f"Resolved host: {host_ip or 'unknown'}. Details: {raw_one}"
        )

    if _is_wsl() and not is_docker:
        return (
            "Could not connect from WSL to Firebird on Windows (TCP). "
            "Ensure the Firebird server is running on the host (port 3050), the database path in the DSN is valid, "
            "and the file is not exclusively locked. For file-only access in WSL you can try FIREBIRD_USE_LOCAL_PATH=1 "
            f"when appropriate. Details: {raw_one}"
        )

    if any(x in low for x in ("network", "connection", "establish", "unable to complete")):
        return (
            "Could not reach the Firebird server over the network. "
            "Confirm the service is listening (e.g. 127.0.0.1:3050), firewalls allow it, and the database path is correct. "
            f"DSN: {dsn!r}. Details: {raw_one}"
        )

    return f"Firebird connection failed. Details: {raw_one}"


def _resolve_firebird_host() -> str:
    """Resolve the host IP for Firebird connection, especially for WSL/Docker."""
    is_docker = os.environ.get("DOCKER_CONTAINER") == "1"
    
    # 0. Check Cache
    global _cached_host_ip
    with _IP_RESOLVE_LOCK:
        host_ip = _cached_host_ip
    
    if host_ip:
        return host_ip

    # 1. Try Env Var
    host_ip = os.environ.get("FIREBIRD_HOST")
    if host_ip:
        if DEBUG_DB_CONNECTION:
            logger.debug("WSL: Using host_ip from FIREBIRD_HOST: %s", host_ip)
    
    # 2. Check for Docker
    elif is_docker:
        host_ip = "host.docker.internal"
        if DEBUG_DB_CONNECTION:
            logger.debug("WSL: Using host_ip for Docker: %s", host_ip)

    # 3. Try Default Gateway (Most reliable for WSL2)
    if not host_ip:
        try:
            import subprocess
            # output: default via 172.22.144.1 dev eth0 ...
            route_out = subprocess.check_output(["ip", "route", "show", "default"], timeout=2).decode().strip()
            if "via" in route_out:
                host_ip = route_out.split("via")[1].split()[0]
                if DEBUG_DB_CONNECTION:
                    logger.debug("WSL: Resolved host_ip via ip route: %s", host_ip)
        except Exception as e:
            logger.debug("Could not resolve default gateway: %s", e)

    # 4. Fallback to Resolv.conf
    if not host_ip:
        try:
            if os.path.exists("/etc/resolv.conf"):
                with open("/etc/resolv.conf", "r") as f:
                    for line in f:
                        if "nameserver" in line:
                            host_ip = line.split()[1]
                            if DEBUG_DB_CONNECTION:
                                logger.debug("WSL: Fallback host_ip via resolv.conf: %s", host_ip)
                            break
        except (OSError, IndexError) as e:
            logger.debug("Could not read resolv.conf: %s", e)
            
    if not host_ip:
        host_ip = "127.0.0.1"
        if DEBUG_DB_CONNECTION:
            logger.debug("WSL: Host resolution failed, defaulting to 127.0.0.1")
    
    # Update Cache
    with _IP_RESOLVE_LOCK:
        _cached_host_ip = host_ip
    
    return host_ip


def _get_firebird_dsn(ctx: dict) -> str:
    """Construct the Firebird DSN based on OS and configuration."""
    use_local = os.environ.get("FIREBIRD_USE_LOCAL_PATH", "").strip() in ("1", "true", "yes")
    ctx["use_local_path"] = use_local
    
    if os.name == 'nt':
        win_path = DB_PATH
        ctx["win_path"] = win_path
        if use_local:
            return win_path
        return f"127.0.0.1/3050:{win_path}"
    
    # Linux/WSL
    host_ip = _resolve_firebird_host()
    ctx["host_ip"] = host_ip
    
    try:
        win_path = _to_win_path(_PROJECT_ROOT) + "\\" + DB_FILE
        env_win_path = os.environ.get("FIREBIRD_WIN_DB_PATH")
        if env_win_path:
            win_path = env_win_path
    except (TypeError, ValueError):
        win_path = DB_FILE # Fallback
        
    ctx["win_path"] = win_path
    
    if use_local:
        return os.path.join(_PROJECT_ROOT, DB_FILE)
    
    return f"inet://{host_ip}/{win_path}"


def _firebird_path_is_production_scoring_file(path: str) -> bool:
    """True if basename is production scoring_history.fdb (not scoring_history_test.fdb)."""
    if not path:
        return False
    bn = os.path.basename(str(path).replace("/", os.sep))
    return bool(re.fullmatch(r"(?i)scoring_history\.fdb\Z", bn))


def get_db():
    import time
    t0 = time.perf_counter()
    fb_connect_ctx = {
        "dsn": None,
        "host_ip": None,
        "is_docker": os.environ.get("DOCKER_CONTAINER") == "1",
        "use_local_path": False,
        "win_path": None,
    }
    try:
        engine = config.get_database_engine()

        if engine == "postgres":
            if not db_postgres:
                raise RuntimeError("database.engine=postgres but modules.db_postgres is unavailable")
            conn = db_postgres.get_pg_connection()
            if DEBUG_DB_CONNECTION:
                logger.debug("get_db using Postgres primary connection")
            return PostgresConnectionProxy(conn)

        if connect is None:
             raise ImportError("firebird-driver not installed")

        dsn = _get_firebird_dsn(fb_connect_ctx)
        fb_connect_ctx["dsn"] = dsn

        import sys as _sys_mod
        if ("pytest" in _sys_mod.modules or os.environ.get("PYTEST_CURRENT_TEST")) and _firebird_path_is_production_scoring_file(
            fb_connect_ctx.get("win_path") or ""
        ):
            raise RuntimeError(
                "pytest refused Firebird connection to production file scoring_history.fdb; "
                "use scoring_history_test.fdb or a temporary .fdb for tests."
            )

        # OS-specific driver setup and server auto-start
        if os.name != 'nt' and not fb_connect_ctx["use_local_path"]:
            # Linux setup
            if driver_config and hasattr(driver_config, "fb_client_library") and FB_CLIENT_LIBRARY:
                driver_config.fb_client_library.value = FB_CLIENT_LIBRARY
            
            # Auto-start check
            if not fb_connect_ctx["is_docker"] and not _is_firebird_running(fb_connect_ctx["host_ip"]):
                win_root = _to_win_path(_PROJECT_ROOT)
                fb_exe_win = os.path.join(win_root, "Firebird", "firebird.exe")
                _launch_firebird_server_wsl(fb_exe_win)
                time.sleep(3)

        # Connect
        if DEBUG_DB_CONNECTION:
            logger.debug("get_db attempting connect to dsn=%s", dsn)
            
        try:
            with _DB_CONNECT_LOCK:
                conn = connect(dsn, user=DB_USER, password=DB_PASS, charset='UTF8')
        except Exception as e:
            friendly = _humanize_firebird_connect_error(e, fb_connect_ctx)
            logger.error("Database connection failed: %s", friendly)
            raise FirebirdConnectionFailed(friendly) from e

        return FirebirdConnectionProxy(conn)

    except (FirebirdConnectionFailed, ImportError, RuntimeError):
        raise
    except Exception as e:
        logger.error("get_db failed: %s", e)
        traceback.print_exc()
        raise


import contextlib

@contextlib.contextmanager
def connection():
    """Context manager for database connections. Auto-closes on exit."""
    conn = get_db()
    try:
        yield conn
    finally:
        conn.close()


def _get_db_engine() -> str:
    """Return the configured DB engine (``firebird``, ``postgres``, or other explicit value)."""
    return config.get_database_engine()


def get_connector():
    """Return the active IConnector for this process.

    Thin shim that lazily imports ``modules.db_connector.get_connector`` to
    avoid circular imports at module load time.  Callers in this module can
    use ``get_connector()`` instead of ``get_db()`` + cursor + commit + close.
    """
    from modules.db_connector import get_connector as _gc
    return _gc()


def get_image_count(rating_filter=None, label_filter=None, keyword_filter=None, min_score_general=0, min_score_aesthetic=0, min_score_technical=0, date_range=None, folder_path=None, stack_id=None):
    query = "SELECT COUNT(*) FROM images"
    params = []
    conditions = []

    if rating_filter:
        placeholders = ','.join(['?'] * len(rating_filter))
        conditions.append(f"rating IN ({placeholders})")
        params.extend(rating_filter)

    if label_filter:
        clean_labels = [l for l in label_filter if l != "None"]
        has_none = "None" in label_filter

        lbl_conds = []
        if clean_labels:
            placeholders = ','.join(['?'] * len(clean_labels))
            lbl_conds.append(f"label IN ({placeholders})")
            params.extend(clean_labels)

        if has_none:
            lbl_conds.append("(label IS NULL OR label = '')")

        if lbl_conds:
            conditions.append(f"({' OR '.join(lbl_conds)})")

    _add_keyword_filter(conditions, params, keyword_filter)

    # Score Filters
    if min_score_general > 0:
        conditions.append("score_general >= ?")
        params.append(min_score_general)

    if min_score_aesthetic > 0:
        conditions.append("score_aesthetic >= ?")
        params.append(min_score_aesthetic)

    if min_score_technical > 0:
        conditions.append("score_technical >= ?")
        params.append(min_score_technical)

    # Date Filter
    if date_range:
        start_date, end_date = date_range
        logger.debug("Date Range: %s to %s", start_date, end_date)
        if start_date:
            conditions.append("CAST(created_at AS DATE) >= CAST(? AS DATE)")
            params.append(start_date)
        if end_date:
            conditions.append("CAST(created_at AS DATE) <= CAST(? AS DATE)")
            params.append(end_date)

    if folder_path:
        folder_id = get_or_create_folder(folder_path)
        conditions.append("folder_id = ?")
        params.append(folder_id)

    if stack_id:
        conditions.append("stack_id = ?")
        params.append(stack_id)

    if conditions:
        query += " WHERE " + " AND ".join(conditions)

    row = get_connector().query_one(query, tuple(params) if params else None)
    return (row.get("count") or row.get("COUNT(*)") or 0) if row else 0

def get_images_paginated(page=1, page_size=None, sort_by="score", order="desc", rating_filter=None, label_filter=None, keyword_filter=None, min_score_general=0, min_score_aesthetic=0, min_score_technical=0, date_range=None, folder_path=None, stack_id=None):
    # Load page_size from config if not provided
    if page_size is None:
        ui_config = config.get_config_section('ui')
        page_size = ui_config.get('gallery_page_size', 50)
    # Ensure integers
    try: page = int(page)
    except (ValueError, TypeError): page = 1
    try: page_size = int(page_size)
    except (ValueError, TypeError): page_size = 50
    sort_by, order = _validate_sort(sort_by, order)

    offset = (page - 1) * page_size
    if offset < 0: offset = 0

    query = "SELECT * FROM images"
    params = []
    conditions = []

    if rating_filter:
        placeholders = ','.join(['?'] * len(rating_filter))
        conditions.append(f"rating IN ({placeholders})")
        params.extend(rating_filter)

    if label_filter:
        clean_labels = [l for l in label_filter if l != "None"]
        has_none = "None" in label_filter

        lbl_conds = []
        if clean_labels:
            placeholders = ','.join(['?'] * len(clean_labels))
            lbl_conds.append(f"label IN ({placeholders})")
            params.extend(clean_labels)

        if has_none:
            lbl_conds.append("(label IS NULL OR label = '')")

        if lbl_conds:
            conditions.append(f"({' OR '.join(lbl_conds)})")

    _add_keyword_filter(conditions, params, keyword_filter)

    # Score Filters
    if min_score_general > 0:
        conditions.append("score_general >= ?")
        params.append(min_score_general)

    if min_score_aesthetic > 0:
        conditions.append("score_aesthetic >= ?")
        params.append(min_score_aesthetic)

    if min_score_technical > 0:
        conditions.append("score_technical >= ?")
        params.append(min_score_technical)

    # Date Filter
    if date_range:
        start_date, end_date = date_range
        if start_date:
            conditions.append("CAST(created_at AS DATE) >= CAST(? AS DATE)")
            params.append(start_date)
        if end_date:
            conditions.append("CAST(created_at AS DATE) <= CAST(? AS DATE)")
            params.append(end_date)


    if folder_path:
        folder_id = get_or_create_folder(folder_path)
        conditions.append("folder_id = ?")
        params.append(folder_id)

    if stack_id:
        conditions.append("stack_id = ?")
        params.append(stack_id)

    if conditions:
        query += " WHERE " + " AND ".join(conditions)

    query += f" ORDER BY {sort_by} {order} OFFSET ? ROWS FETCH NEXT ? ROWS ONLY"
    params.extend([offset, page_size])

    return list(get_connector().query(query, tuple(params)))

def get_images_paginated_with_count(page=1, page_size=None, sort_by="score", order="desc", rating_filter=None, label_filter=None, keyword_filter=None, min_score_general=0, min_score_aesthetic=0, min_score_technical=0, date_range=None, folder_path=None, stack_id=None, use_exif_date=False, make_filter=None, model_filter=None, lens_filter=None, iso_min=None, iso_max=None):
    """
    Get paginated images AND total count using optimized approach.
    Uses same connection for both queries to reduce overhead.
    
    When use_exif_date=True, date_range uses EXIF date_time_original (fallback to created_at).
    make_filter, model_filter, lens_filter: exact or LIKE match.
    iso_min, iso_max: ISO range filter.
    
    Returns:
        tuple: (rows, total_count) where rows is list of image records and total_count is int
    """
    # Load page_size from config if not provided
    if page_size is None:
        ui_config = config.get_config_section('ui')
        page_size = ui_config.get('gallery_page_size', 50)

    # Ensure integers
    try: page = int(page)
    except (ValueError, TypeError): page = 1
    try: page_size = int(page_size)
    except (ValueError, TypeError): page_size = 50
    sort_by, order = _validate_sort(sort_by, order)

    offset = (page - 1) * page_size
    if offset < 0: offset = 0

    # Need EXIF join for: EXIF filters, EXIF sort, or use_exif_date
    exif_sort_cols = {"date_time_original", "make", "model", "lens_model", "iso"}
    need_exif_join = (
        use_exif_date and date_range
        or make_filter or model_filter or lens_filter or iso_min or iso_max
        or sort_by in exif_sort_cols
    )

    if need_exif_join:
        from_clause = " images LEFT JOIN image_exif ON images.id = image_exif.image_id"
        tbl_prefix = "images."
    else:
        from_clause = " images"
        tbl_prefix = ""

    # Build base query components
    params = []
    conditions = []

    # Rating filter
    if rating_filter:
        placeholders = ','.join(['?'] * len(rating_filter))
        conditions.append(f"{tbl_prefix}rating IN ({placeholders})")
        params.extend(rating_filter)

    # Label filter
    if label_filter:
        clean_labels = [l for l in label_filter if l != "None"]
        has_none = "None" in label_filter

        lbl_conds = []
        if clean_labels:
            placeholders = ','.join(['?'] * len(clean_labels))
            lbl_conds.append(f"{tbl_prefix}label IN ({placeholders})")
            params.extend(clean_labels)

        if has_none:
            lbl_conds.append(f"({tbl_prefix}label IS NULL OR {tbl_prefix}label = '')")

        if lbl_conds:
            conditions.append(f"({' OR '.join(lbl_conds)})")

    # Keyword filter
    _add_keyword_filter(conditions, params, keyword_filter)

    # Score Filters
    if min_score_general > 0:
        conditions.append(f"{tbl_prefix}score_general >= ?")
        params.append(min_score_general)

    if min_score_aesthetic > 0:
        conditions.append(f"{tbl_prefix}score_aesthetic >= ?")
        params.append(min_score_aesthetic)

    if min_score_technical > 0:
        conditions.append(f"{tbl_prefix}score_technical >= ?")
        params.append(min_score_technical)

    # Date Filter
    if date_range:
        start_date, end_date = date_range
        date_col = f"COALESCE(image_exif.date_time_original, images.created_at)" if (need_exif_join and use_exif_date) else f"{tbl_prefix}created_at"
        if start_date:
            conditions.append(f"CAST({date_col} AS DATE) >= CAST(? AS DATE)")
            params.append(start_date)
        if end_date:
            conditions.append(f"CAST({date_col} AS DATE) <= CAST(? AS DATE)")
            params.append(end_date)

    # EXIF filters
    if need_exif_join:
        if make_filter and make_filter.strip():
            conditions.append("image_exif.make = ?")
            params.append(make_filter.strip())
        if model_filter and model_filter.strip():
            conditions.append("image_exif.model = ?")
            params.append(model_filter.strip())
        if lens_filter and lens_filter.strip():
            conditions.append("image_exif.lens_model LIKE ?")
            params.append(f"%{lens_filter.strip()}%")
        if iso_min is not None and iso_min > 0:
            conditions.append("image_exif.iso >= ?")
            params.append(iso_min)
        if iso_max is not None and iso_max > 0:
            conditions.append("image_exif.iso <= ?")
            params.append(iso_max)

    # Folder filter
    if folder_path:
        folder_id = get_or_create_folder(folder_path)
        conditions.append(f"{tbl_prefix}folder_id = ?")
        params.append(folder_id)

    # Stack filter
    if stack_id:
        conditions.append(f"{tbl_prefix}stack_id = ?")
        params.append(stack_id)

    # Build WHERE clause
    where_clause = ""
    if conditions:
        where_clause = " WHERE " + " AND ".join(conditions)

    # ORDER BY - EXIF columns need special handling
    if need_exif_join and sort_by in exif_sort_cols:
        nulls = " NULLS LAST" if order == "DESC" else " NULLS FIRST"
        if sort_by == "date_time_original":
            order_by = f"COALESCE(image_exif.date_time_original, images.created_at) {order}{nulls}"
        else:
            order_by = f"image_exif.{sort_by} {order}{nulls}"
    else:
        order_by = f"{tbl_prefix}{sort_by} {order}" if tbl_prefix else f"{sort_by} {order}"

    # OPTIMIZATION: Use Window Function to get count and data in single query
    query = f"SELECT images.*, COUNT(*) OVER() as total_count FROM{from_clause}{where_clause} ORDER BY {order_by} OFFSET ? ROWS FETCH NEXT ? ROWS ONLY"
    params.extend([offset, page_size])

    try:
        rows = list(get_connector().query(query, tuple(params)))
        total_count = 0
        if rows:
            try:
                total_count = int(rows[0]['total_count'])
            except (KeyError, IndexError, ValueError):
                total_count = len(rows)
                logger.warning("Could not retrieve total_count from window function result")
        return rows, total_count
    except Exception as e:
        logger.error("Error in get_images_paginated_with_count: %s", e)
        # Fallback to separate queries if window function fails
        try:
            count_row = get_connector().query_one(f"SELECT COUNT(*) FROM{from_clause}{where_clause}", tuple(params[:-2]))
            total_count = (count_row.get("count") or count_row.get("COUNT(*)") or 0) if count_row else 0
            query_fallback = f"SELECT images.* FROM{from_clause}{where_clause} ORDER BY {order_by} OFFSET ? ROWS FETCH NEXT ? ROWS ONLY"
            rows = list(get_connector().query(query_fallback, tuple(params)))
            return rows, total_count
        except Exception:
            raise e

def get_filtered_paths(rating_filter=None, label_filter=None, keyword_filter=None, min_score_general=0, min_score_aesthetic=0, min_score_technical=0, date_range=None, folder_path=None, stack_id=None):
    """
    Returns a list of file_paths matching the filters (No pagination).
    """
    query = "SELECT file_path FROM images"
    params = []
    conditions = []
    
    if rating_filter:
        placeholders = ','.join(['?'] * len(rating_filter))
        conditions.append(f"rating IN ({placeholders})")
        params.extend(rating_filter)
        
    if label_filter:
        clean_labels = [l for l in label_filter if l != "None"]
        has_none = "None" in label_filter
        
        lbl_conds = []
        if clean_labels:
            placeholders = ','.join(['?'] * len(clean_labels))
            lbl_conds.append(f"label IN ({placeholders})")
            params.extend(clean_labels)
            
        if has_none:
            lbl_conds.append("(label IS NULL OR label = '')")
            
        if lbl_conds:
            conditions.append(f"({' OR '.join(lbl_conds)})")

    _add_keyword_filter(conditions, params, keyword_filter)

    # Score Filters
    if min_score_general > 0:
        conditions.append("score_general >= ?")
        params.append(min_score_general)

    if min_score_aesthetic > 0:
        conditions.append("score_aesthetic >= ?")
        params.append(min_score_aesthetic)

    if min_score_technical > 0:
        conditions.append("score_technical >= ?")
        params.append(min_score_technical)

    # Date Filter
    if date_range:
        start_date, end_date = date_range
        logger.debug("Date Range: %s to %s", start_date, end_date)
        if start_date:
            conditions.append("CAST(created_at AS DATE) >= CAST(? AS DATE)")
            params.append(start_date)
        if end_date:
            conditions.append("CAST(created_at AS DATE) <= CAST(? AS DATE)")
            params.append(end_date)
            
    if folder_path:
        folder_id = get_or_create_folder(folder_path)
        conditions.append("folder_id = ?")
        params.append(folder_id)

    if stack_id:
        conditions.append("stack_id = ?")
        params.append(stack_id)

    if conditions:
        query += " WHERE " + " AND ".join(conditions)

    rows = get_connector().query(query, tuple(params))
    return [row['file_path'] for row in rows]

def image_exists(file_path, current_version=None):
    row = get_connector().query_one(
        "SELECT score, model_version, score_general, thumbnail_path FROM images WHERE file_path = ?",
        (file_path,)
    )

    if row:
        score = row['score']

        # Check if score is valid
        if score is None or score <= 0:
            return False

        # Check thumbnail existence (User Requirement)
        if not row['thumbnail_path']:
            return False

        # Check version if provided
        if current_version:
            if row['model_version'] != current_version:
                return False  # Version mismatch, treat as stale
            # Also strict check for score_general if we are strictly version checking
            sg = row['score_general']
            if sg is None or sg <= 0:
                return False

        return True
    return False

_db_initialized = False


def reset_init_db_state_for_tests():
    """Allow ``init_db()`` to run again after tests switch ``DB_PATH`` to another file."""
    global _db_initialized
    _db_initialized = False


def init_db():
    global _db_initialized
    if _db_initialized:
        return
    if os.environ.get('SKIP_DB_INIT'):
        logger.debug("SKIP_DB_INIT set, skipping DDL.")
        return

    try:
        conn = get_db()
        conn.close()
        # Initialize/Migrate Schema — Firebird uses RDB$ catalog DDL; Postgres has a dedicated path.
        if _get_db_engine() == "postgres":
            if not db_postgres:
                raise RuntimeError("database.engine is postgres but modules.db_postgres is unavailable")
            db_postgres.init_db()
            seed_pipeline_phases()
        else:
            _init_db_impl()
        _db_initialized = True
    except Exception as e:
        if _get_db_engine() == "postgres":
            logging.error("PostgreSQL initialization failed: %s", e)
        else:
            logging.error(
                "Firebird connection failed: %s. Please run migrate_to_firebird.py first.",
                e,
            )
        raise

def _table_exists(cursor, table_name):
    """Check if a table exists in Firebird database."""
    cursor.execute(
        "SELECT 1 FROM RDB$RELATIONS WHERE RDB$RELATION_NAME = ? AND RDB$SYSTEM_FLAG = 0",
        (table_name.upper(),)
    )
    result = cursor.fetchone() is not None
    logger.debug("Table %s: %s", table_name, result)
    return result

def _index_exists(cursor, index_name):
    """Check if an index exists in Firebird database."""
    cursor.execute(
        "SELECT 1 FROM RDB$INDICES WHERE RDB$INDEX_NAME = ?",
        (index_name.upper(),)
    )
    return cursor.fetchone() is not None

def _backup_db():
    """Create a timestamped backup of the database."""
    if "pytest" in sys.modules:
        return
    if not os.path.exists(DB_PATH):
        return

    try:
        timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        backup_path = f"{DB_PATH}.{timestamp}.bak"
        print(f"Creating DB backup: {backup_path}")
        shutil.copy2(DB_PATH, backup_path)
    except Exception as e:
        print(f"Warning: Failed to create DB backup: {e}")

def _constraint_exists(cursor, constraint_name):
    """Check if a constraint exists in Firebird database."""
    cursor.execute(
        "SELECT 1 FROM RDB$RELATION_CONSTRAINTS WHERE RDB$CONSTRAINT_NAME = ?",
        (constraint_name.upper(),)
    )
    return cursor.fetchone() is not None

def _init_db_impl():
    conn = get_db()
    c = conn.cursor()
    
    # Jobs table
    if not _table_exists(c, 'JOBS'):
        c.execute('''CREATE TABLE jobs (
            id INTEGER GENERATED BY DEFAULT AS IDENTITY PRIMARY KEY,
            input_path VARCHAR(4000),
            phase_id INTEGER,
            job_type VARCHAR(50),
            status VARCHAR(50),
            priority SMALLINT DEFAULT 100,
            retry_count INTEGER DEFAULT 0,
            target_scope VARCHAR(255),
            paused_at TIMESTAMP,
            queue_position INTEGER,
            cancel_requested SMALLINT DEFAULT 0,
            queue_payload BLOB SUB_TYPE TEXT,
            created_at TIMESTAMP,
            enqueued_at TIMESTAMP,
            started_at TIMESTAMP,
            finished_at TIMESTAMP,
            completed_at TIMESTAMP,
            log BLOB SUB_TYPE TEXT,
            current_phase VARCHAR(50),
            next_phase_index INTEGER,
            runner_state VARCHAR(50)
        )''')
        try: conn.commit()
        except Exception: pass

    # Job phases table (persisted multi-step pipeline plans)
    if not _table_exists(c, 'JOB_PHASES'):
        c.execute('''CREATE TABLE job_phases (
            id INTEGER GENERATED BY DEFAULT AS IDENTITY PRIMARY KEY,
            job_id INTEGER NOT NULL,
            phase_order INTEGER NOT NULL,
            phase_code VARCHAR(50) NOT NULL,
            state VARCHAR(20) NOT NULL,
            started_at TIMESTAMP,
            completed_at TIMESTAMP,
            error_message BLOB SUB_TYPE TEXT
        )''')
        try: conn.commit()
        except Exception: pass

    # Job steps table — sub-phase telemetry (e.g. individual ML model runs within Scoring stage)
    if not _table_exists(c, 'JOB_STEPS'):
        try:
            c.execute('''CREATE TABLE job_steps (
                id INTEGER GENERATED BY DEFAULT AS IDENTITY PRIMARY KEY,
                job_id INTEGER NOT NULL,
                phase_code VARCHAR(50) NOT NULL,
                step_code VARCHAR(50) NOT NULL,
                step_name VARCHAR(100) NOT NULL,
                status VARCHAR(20) DEFAULT \'pending\',
                started_at TIMESTAMP,
                completed_at TIMESTAMP,
                items_total INTEGER DEFAULT 0,
                items_done INTEGER DEFAULT 0,
                throughput_rps DOUBLE PRECISION,
                error_message BLOB SUB_TYPE TEXT
            )''')
            conn.commit()
        except Exception:
            conn.rollback()

    # jobs: scope columns for multi-path runs (additive migration)
    try:
        if not _column_exists(c, 'JOBS', 'SCOPE_TYPE'):
            c.execute("ALTER TABLE jobs ADD scope_type VARCHAR(30)")
            conn.commit()
    except Exception:
        conn.rollback()

    try:
        if not _column_exists(c, 'JOBS', 'SCOPE_PATHS'):
            c.execute("ALTER TABLE jobs ADD scope_paths BLOB SUB_TYPE TEXT")
            conn.commit()
    except Exception:
        conn.rollback()

    try:
        if not _column_exists(c, 'JOBS', 'PAUSED_AT'):
            c.execute("ALTER TABLE jobs ADD paused_at TIMESTAMP")
            conn.commit()
    except Exception:
        conn.rollback()
    
    # Images table
    if not _table_exists(c, 'IMAGES'):
        c.execute('''CREATE TABLE images (
            id INTEGER GENERATED BY DEFAULT AS IDENTITY PRIMARY KEY,
            job_id INTEGER,
            file_path VARCHAR(4000),
            file_name VARCHAR(255),
            file_type VARCHAR(20),
            score DOUBLE PRECISION,
            keywords BLOB SUB_TYPE TEXT,
            title VARCHAR(500),
            description BLOB SUB_TYPE TEXT,
            metadata BLOB SUB_TYPE TEXT,
            thumbnail_path VARCHAR(4000),
            scores_json BLOB SUB_TYPE TEXT,
            created_at TIMESTAMP
        )''')
        try: conn.commit()
        except Exception: pass
        c = conn.cursor()
    
    # Check for missing columns (Schema Migration)
    c.execute("SELECT RDB$FIELD_NAME FROM RDB$RELATION_FIELDS WHERE RDB$RELATION_NAME = 'IMAGES'")
    columns = [row[0].strip().lower() for row in c.fetchall()]
    
    if "file_type" not in columns:
        try: c.execute("ALTER TABLE images ADD file_type VARCHAR(20)")
        except Exception: pass

    if "thumbnail_path" not in columns:
        try: c.execute("ALTER TABLE images ADD thumbnail_path VARCHAR(4000)")
        except Exception: pass

    if "thumbnail_path_win" not in columns:
        try: c.execute("ALTER TABLE images ADD thumbnail_path_win VARCHAR(4000)")
        except Exception: pass
        
    if "scores_json" not in columns:
        try: c.execute("ALTER TABLE images ADD scores_json BLOB SUB_TYPE TEXT")
        except Exception: pass

    if "image_embedding" not in columns:
        try: c.execute("ALTER TABLE images ADD image_embedding BLOB SUB_TYPE 0")
        except Exception: pass

    # Stacks table
    if not _table_exists(c, 'STACKS'):
        c.execute('''CREATE TABLE stacks (
            id INTEGER GENERATED BY DEFAULT AS IDENTITY PRIMARY KEY,
            name VARCHAR(255),
            best_image_id INTEGER,
            created_at TIMESTAMP
        )''')
        try: conn.commit()
        except Exception: pass
    
    # Folders table
    if not _table_exists(c, 'FOLDERS'):
        c.execute('''CREATE TABLE folders (
            id INTEGER GENERATED BY DEFAULT AS IDENTITY PRIMARY KEY,
            path VARCHAR(4000),
            parent_id INTEGER,
            is_fully_scored INTEGER DEFAULT 0,
            is_keywords_processed INTEGER DEFAULT 0,
            phase_agg_dirty INTEGER DEFAULT 1,
            phase_agg_updated_at TIMESTAMP,
            phase_agg_json BLOB SUB_TYPE TEXT,
            created_at TIMESTAMP
        )''')
        try: conn.commit()
        except Exception: pass
    
    # Check For Folders Columns
    if _table_exists(c, 'FOLDERS'):
        c.execute("SELECT RDB$FIELD_NAME FROM RDB$RELATION_FIELDS WHERE RDB$RELATION_NAME = 'FOLDERS'")
        folder_cols = [row[0].strip().lower() for row in c.fetchall()]

        if "is_fully_scored" not in folder_cols:
             try: c.execute("ALTER TABLE folders ADD is_fully_scored INTEGER DEFAULT 0")
             except Exception: pass

        if "is_keywords_processed" not in folder_cols:
             try: c.execute("ALTER TABLE folders ADD is_keywords_processed INTEGER DEFAULT 0")
             except Exception: pass

        # Phase aggregate cache (phase-aware replacement for legacy booleans)
        if "phase_agg_dirty" not in folder_cols:
             try: c.execute("ALTER TABLE folders ADD phase_agg_dirty INTEGER DEFAULT 1")
             except Exception: pass

        if "phase_agg_updated_at" not in folder_cols:
             try: c.execute("ALTER TABLE folders ADD phase_agg_updated_at TIMESTAMP")
             except Exception: pass

        if "phase_agg_json" not in folder_cols:
             try: c.execute("ALTER TABLE folders ADD phase_agg_json BLOB SUB_TYPE TEXT")
             except Exception: pass
        
        try: conn.commit()
        except Exception: pass

    if not _table_exists(c, 'CLUSTER_PROGRESS'):
        c.execute('''CREATE TABLE cluster_progress (
            folder_path VARCHAR(512) NOT NULL PRIMARY KEY,
            last_run TIMESTAMP
        )''')
        try: conn.commit()
        except Exception: pass

    # Culling Sessions table
    if not _table_exists(c, 'CULLING_SESSIONS'):
        c.execute('''CREATE TABLE culling_sessions (
            id INTEGER GENERATED BY DEFAULT AS IDENTITY PRIMARY KEY,
            folder_path VARCHAR(4000),
            mode VARCHAR(50),
            status VARCHAR(50) DEFAULT 'active',
            total_images INTEGER DEFAULT 0,
            total_groups INTEGER DEFAULT 0,
            reviewed_groups INTEGER DEFAULT 0,
            picked_count INTEGER DEFAULT 0,
            rejected_count INTEGER DEFAULT 0,
            created_at TIMESTAMP,
            completed_at TIMESTAMP
        )''')
        try: conn.commit()
        except Exception: pass

    # Culling Picks table
    if not _table_exists(c, 'CULLING_PICKS'):
        c.execute('''CREATE TABLE culling_picks (
            id INTEGER GENERATED BY DEFAULT AS IDENTITY PRIMARY KEY,
            session_id INTEGER,
            image_id INTEGER,
            group_id INTEGER,
            decision VARCHAR(50),
            auto_suggested SMALLINT DEFAULT 0,
            is_best_in_group SMALLINT DEFAULT 0,
            created_at TIMESTAMP
        )''')
        try: conn.commit()
        except Exception: pass
    
    # Index for fast lookup
    if not _index_exists(c, 'IDX_CULLING_PICKS_SESSION'):
        try: c.execute("CREATE INDEX idx_culling_picks_session ON culling_picks(session_id)")
        except Exception: pass
    if not _index_exists(c, 'IDX_CULLING_PICKS_IMAGE'):
        try: c.execute("CREATE INDEX idx_culling_picks_image ON culling_picks(image_id)")
        except Exception: pass
    
    # IMAGE_EXIF — cached EXIF metadata (one row per image)
    if not _table_exists(c, 'IMAGE_EXIF'):
        c.execute('''CREATE TABLE image_exif (
            image_id INTEGER NOT NULL PRIMARY KEY,
            make VARCHAR(100),
            model VARCHAR(200),
            lens_model VARCHAR(255),
            focal_length VARCHAR(50),
            focal_length_35mm SMALLINT,
            date_time_original TIMESTAMP,
            create_date TIMESTAMP,
            exposure_time VARCHAR(30),
            f_number VARCHAR(20),
            iso INTEGER,
            exposure_compensation VARCHAR(20),
            image_width INTEGER,
            image_height INTEGER,
            orientation SMALLINT,
            flash SMALLINT,
            image_unique_id VARCHAR(64),
            shutter_count INTEGER,
            sub_sec_time_original VARCHAR(10),
            extracted_at TIMESTAMP
        )''')
        try: conn.commit()
        except Exception: pass
        c = conn.cursor()
    
    if _table_exists(c, 'IMAGE_EXIF') and not _constraint_exists(c, 'FK_IMAGE_EXIF_IMAGES'):
        try:
            c.execute("ALTER TABLE image_exif ADD CONSTRAINT fk_image_exif_images FOREIGN KEY (image_id) REFERENCES images(id) ON DELETE CASCADE")
            conn.commit()
        except Exception: pass
        c = conn.cursor()
    
    if _table_exists(c, 'IMAGE_EXIF'):
        for idx in ('IDX_IMAGE_EXIF_DATE', 'IDX_IMAGE_EXIF_MAKE', 'IDX_IMAGE_EXIF_MODEL', 'IDX_IMAGE_EXIF_LENS', 'IDX_IMAGE_EXIF_ISO'):
            col = {'IDX_IMAGE_EXIF_DATE': 'date_time_original', 'IDX_IMAGE_EXIF_MAKE': 'make', 'IDX_IMAGE_EXIF_MODEL': 'model', 'IDX_IMAGE_EXIF_LENS': 'lens_model', 'IDX_IMAGE_EXIF_ISO': 'iso'}[idx]
            if not _index_exists(c, idx):
                try: c.execute(f"CREATE INDEX {idx.lower()} ON image_exif({col})")
                except Exception: pass
        # Migration: iso SMALLINT overflows for high ISO (e.g. 51200). Change to INTEGER.
        try:
            c.execute("ALTER TABLE image_exif ALTER COLUMN iso TYPE INTEGER")
            conn.commit()
        except Exception:
            pass  # Column may already be INTEGER or Firebird < 4
        c = conn.cursor()
    
    # IMAGE_XMP — cached XMP sidecar metadata (one row per image)
    if not _table_exists(c, 'IMAGE_XMP'):
        c.execute('''CREATE TABLE image_xmp (
            image_id INTEGER NOT NULL PRIMARY KEY,
            rating SMALLINT,
            label VARCHAR(50),
            pick_status SMALLINT,
            burst_uuid VARCHAR(64),
            stack_id VARCHAR(64),
            keywords BLOB SUB_TYPE TEXT,
            title VARCHAR(500),
            description BLOB SUB_TYPE TEXT,
            create_date TIMESTAMP,
            modify_date TIMESTAMP,
            extracted_at TIMESTAMP
        )''')
        try: conn.commit()
        except Exception: pass
        c = conn.cursor()
    
    if _table_exists(c, 'IMAGE_XMP') and not _constraint_exists(c, 'FK_IMAGE_XMP_IMAGES'):
        try:
            c.execute("ALTER TABLE image_xmp ADD CONSTRAINT fk_image_xmp_images FOREIGN KEY (image_id) REFERENCES images(id) ON DELETE CASCADE")
            conn.commit()
        except Exception: pass
        c = conn.cursor()
    
    if _table_exists(c, 'IMAGE_XMP'):
        for idx in ('IDX_IMAGE_XMP_BURST', 'IDX_IMAGE_XMP_PICK'):
            col = {'IDX_IMAGE_XMP_BURST': 'burst_uuid', 'IDX_IMAGE_XMP_PICK': 'pick_status'}[idx]
            if not _index_exists(c, idx):
                try: c.execute(f"CREATE INDEX {idx.lower()} ON image_xmp({col})")
                except Exception: pass
    
    try: conn.commit()
    except Exception: pass
    
    # Additional migrations (Scores etc)
    if not _index_exists(c, 'IDX_STACK_ID') and "stack_id" in columns:
        try: c.execute("CREATE INDEX idx_stack_id ON images(stack_id)")
        except Exception: pass
    if not _index_exists(c, 'IDX_FOLDER_ID') and "folder_id" in columns:
        try: c.execute("CREATE INDEX idx_folder_id ON images(folder_id)")
        except Exception: pass

    
    c = conn.cursor()
    logger.debug("_init_db_impl started, checking for backup...")
    try:
        # Check if migration needed (e.g. missing columns or old tables)
        migration_needed = False
        
        # Check if RESOLVED_PATHS exists
        if _table_exists(c, 'RESOLVED_PATHS'):
             logger.debug("RESOLVED_PATHS exists, migration needed.")
             migration_needed = True
        else:
             logger.debug("RESOLVED_PATHS not found.")
             
        # Check if FILE_PATHS needs columns
        logger.debug("Checking FILE_PATHS columns...")
        c.execute("SELECT RDB$FIELD_NAME FROM RDB$RELATION_FIELDS WHERE RDB$RELATION_NAME = 'FILE_PATHS'")
        fp_cols = [row[0].strip().lower() for row in c.fetchall()]
        logger.debug("FILE_PATHS columns: %s", fp_cols)
        
        if "path_type" not in fp_cols or "is_verified" not in fp_cols:
             migration_needed = True

        if migration_needed:
             print("Database migration required. Starting backup...")
             # Close conn to release lock for backup
             conn.close()
             logger.debug("Closed conn for backup.")
             _backup_db()
             # Re-open
             logger.debug("Re-opening DB after backup.")
             conn = get_db()
             c = conn.cursor()

        # --- 1. Refactor FILE_PATHS & Merge RESOLVED_PATHS ---
        
        # Add columns to FILE_PATHS if missing
        # Re-fetch columns in case we reopened? No, we have fp_cols but we need to check again if we want to be safe or just use logic.
        # But wait, if we closed conn, we need to re-fetch?
        # Yes, if we use fp_cols variable it is fine.
        
        logger.debug("Starting Schema updates...")
        if not _table_exists(c, "FILE_PATHS"):
            c.execute(
                """CREATE TABLE file_paths (
                    id INTEGER GENERATED BY DEFAULT AS IDENTITY PRIMARY KEY,
                    image_id INTEGER NOT NULL,
                    path VARCHAR(4000),
                    last_seen TIMESTAMP,
                    path_type VARCHAR(10) DEFAULT 'WSL',
                    is_verified SMALLINT DEFAULT 0,
                    verification_date TIMESTAMP
                )"""
            )
            try:
                conn.commit()
            except Exception:
                conn.rollback()
            c = conn.cursor()

        c.execute("SELECT RDB$FIELD_NAME FROM RDB$RELATION_FIELDS WHERE RDB$RELATION_NAME = 'FILE_PATHS'")
        fp_cols = [row[0].strip().lower() for row in c.fetchall()]
        
        if "path_type" not in fp_cols:
             print("Migrating FILE_PATHS: Adding path_type...")
             c.execute("ALTER TABLE file_paths ADD path_type VARCHAR(10) DEFAULT 'WSL'")
             conn.commit()
             c = conn.cursor() # Refresh cursor
             
             print("Migrating FILE_PATHS: Updating path_type default...")
             c.execute("UPDATE file_paths SET path_type = 'WSL'") 
             conn.commit()
             c = conn.cursor()
        
        if "is_verified" not in fp_cols:
             print("Migrating FILE_PATHS: Adding is_verified...")
             c.execute("ALTER TABLE file_paths ADD is_verified SMALLINT DEFAULT 0")
             conn.commit()
             c = conn.cursor()
        if "verification_date" not in fp_cols:
             print("Migrating FILE_PATHS: Adding verification_date...")
             c.execute("ALTER TABLE file_paths ADD verification_date TIMESTAMP")
             conn.commit()
             c = conn.cursor()

        # Migrate RESOLVED_PATHS -> FILE_PATHS
        if _table_exists(c, 'RESOLVED_PATHS'):
             print("Migrating RESOLVED_PATHS -> FILE_PATHS...")
             # Move data
             c.execute("""
                INSERT INTO file_paths (image_id, path, path_type, is_verified, verification_date, last_seen)
                SELECT image_id, windows_path, 'WIN', is_verified, verification_date, last_checked 
                FROM resolved_paths
             """)
             # Drop old table
             c.execute("DROP TABLE resolved_paths")

        # --- 2. Enforce Integirty (Foreign Keys) ---

        # FK_FILE_PATHS_IMAGES
        if not _constraint_exists(c, 'FK_FILE_PATHS_IMAGES'):
             print("Applying FK_FILE_PATHS_IMAGES...")
             c.execute("DELETE FROM file_paths WHERE image_id NOT IN (SELECT id FROM images)")
             conn.commit()
             c.execute("ALTER TABLE file_paths ADD CONSTRAINT fk_file_paths_images FOREIGN KEY (image_id) REFERENCES images(id) ON DELETE CASCADE")

        # FK_IMAGES_FOLDERS
        if not _constraint_exists(c, 'FK_IMAGES_FOLDERS') and _table_exists(c, 'FOLDERS'):
             print("Applying FK_IMAGES_FOLDERS...")
             # Set invalid folder_ids to NULL
             c.execute("UPDATE images SET folder_id = NULL WHERE folder_id NOT IN (SELECT id FROM folders)")
             c.execute("ALTER TABLE images ADD CONSTRAINT fk_images_folders FOREIGN KEY (folder_id) REFERENCES folders(id) ON DELETE SET NULL")

        # FK_FOLDERS_PARENT
        if not _constraint_exists(c, 'FK_FOLDERS_PARENT') and _table_exists(c, 'FOLDERS'):
             print("Applying FK_FOLDERS_PARENT...")
             c.execute("UPDATE folders SET parent_id = NULL WHERE parent_id NOT IN (SELECT id FROM folders)")
             # Prevent self-reference issues? Parent must exist.
             c.execute("ALTER TABLE folders ADD CONSTRAINT fk_folders_parent FOREIGN KEY (parent_id) REFERENCES folders(id) ON DELETE CASCADE")

        # FK_CULLING_PICKS
        if _table_exists(c, 'CULLING_PICKS'):
             if not _constraint_exists(c, 'FK_CULLING_PICKS_IMAGES'):
                 print("Applying FK_CULLING_PICKS_IMAGES...")
                 c.execute("DELETE FROM culling_picks WHERE image_id NOT IN (SELECT id FROM images)")
                 c.execute("ALTER TABLE culling_picks ADD CONSTRAINT fk_culling_picks_images FOREIGN KEY (image_id) REFERENCES images(id) ON DELETE CASCADE")
             
             if not _constraint_exists(c, 'FK_CULLING_PICKS_SESSIONS') and _table_exists(c, 'CULLING_SESSIONS'):
                 print("Applying FK_CULLING_PICKS_SESSIONS...")
                 c.execute("DELETE FROM culling_picks WHERE session_id NOT IN (SELECT id FROM culling_sessions)")
                 c.execute("ALTER TABLE culling_picks ADD CONSTRAINT fk_culling_picks_sessions FOREIGN KEY (session_id) REFERENCES culling_sessions(id) ON DELETE CASCADE")

        # Index on path_type
        if _table_exists(c, 'FILE_PATHS') and not _index_exists(c, 'IDX_FILE_PATHS_IMG_TYPE'):
             c.execute("CREATE INDEX idx_file_paths_img_type ON file_paths(image_id, path_type)")

        # Selection feature: cull_decision, cull_policy_version on IMAGES
        c.execute("SELECT RDB$FIELD_NAME FROM RDB$RELATION_FIELDS WHERE RDB$RELATION_NAME = 'IMAGES'")
        img_cols = [row[0].strip().lower() for row in c.fetchall()]
        if "cull_decision" not in img_cols:
            try:
                c.execute("ALTER TABLE images ADD cull_decision VARCHAR(20)")
                conn.commit()
            except Exception as m:
                logger.debug("Adding cull_decision column: %s", m)
                try:
                    conn.rollback()
                except Exception:
                    pass
        if "cull_policy_version" not in img_cols:
            try:
                c.execute("ALTER TABLE images ADD cull_policy_version VARCHAR(50)")
                conn.commit()
            except Exception as m:
                logger.debug("Adding cull_policy_version column: %s", m)
                try:
                    conn.rollback()
                except Exception:
                    pass
        if "image_uuid" not in img_cols:
            try:
                c.execute("ALTER TABLE images ADD image_uuid VARCHAR(36)")
                conn.commit()
            except Exception as m:
                logger.debug("Adding image_uuid column: %s", m)
                try:
                    conn.rollback()
                except Exception:
                    pass
        # Unique index on image_uuid to prevent duplicates (allows multiple NULLs)
        if "image_uuid" in img_cols and not _index_exists(c, 'UQ_IMAGES_IMAGE_UUID'):
            try:
                c.execute("CREATE UNIQUE INDEX uq_images_image_uuid ON images(image_uuid)")
                conn.commit()
            except Exception as e:
                logger.debug("Could not create unique index on image_uuid (may have existing duplicates): %s", e)
                try:
                    conn.rollback()
                except Exception:
                    pass

    except Exception as e:
        logger.error("Migration error: %s", e)
        import traceback
        traceback.print_exc()
        try:
            conn.rollback()
        except Exception:
            pass

    # --- Pipeline Phases Tables ---
    try:
        c = conn.cursor()

        # PIPELINE_PHASES — phase registry
        if not _table_exists(c, 'PIPELINE_PHASES'):
            c.execute('''CREATE TABLE pipeline_phases (
                id          INTEGER GENERATED BY DEFAULT AS IDENTITY PRIMARY KEY,
                code        VARCHAR(50) NOT NULL,
                name        VARCHAR(100) NOT NULL,
                description BLOB SUB_TYPE TEXT,
                sort_order  INTEGER DEFAULT 0 NOT NULL,
                enabled     SMALLINT DEFAULT 1 NOT NULL,
                optional    SMALLINT DEFAULT 0 NOT NULL,
                default_skip SMALLINT DEFAULT 0 NOT NULL
            )''')
            conn.commit()
            c = conn.cursor()
            # Unique index on code
            if not _index_exists(c, 'UQ_PIPELINE_PHASES_CODE'):
                c.execute("CREATE UNIQUE INDEX uq_pipeline_phases_code ON pipeline_phases(code)")
                conn.commit()
                c = conn.cursor()

        # PIPELINE_PHASES — add optional/default_skip columns if missing
        if _table_exists(c, 'PIPELINE_PHASES'):
            c.execute("SELECT RDB$FIELD_NAME FROM RDB$RELATION_FIELDS WHERE RDB$RELATION_NAME = 'PIPELINE_PHASES'")
            pp_cols = [row[0].strip().lower() for row in c.fetchall()]
            if 'optional' not in pp_cols:
                try:
                    c.execute("ALTER TABLE pipeline_phases ADD optional SMALLINT DEFAULT 0 NOT NULL")
                    conn.commit()
                    c = conn.cursor()
                except Exception as m:
                    logger.debug("Adding pipeline_phases.optional: %s", m)
                    try: conn.rollback()
                    except Exception: pass
                    c = conn.cursor()
            if 'default_skip' not in pp_cols:
                try:
                    c.execute("ALTER TABLE pipeline_phases ADD default_skip SMALLINT DEFAULT 0 NOT NULL")
                    conn.commit()
                    c = conn.cursor()
                except Exception as m:
                    logger.debug("Adding pipeline_phases.default_skip: %s", m)
                    try: conn.rollback()
                    except Exception: pass
                    c = conn.cursor()

        # IMAGE_PHASE_STATUS — per-image per-phase tracking
        if not _table_exists(c, 'IMAGE_PHASE_STATUS'):
            c.execute('''CREATE TABLE image_phase_status (
                id               INTEGER GENERATED BY DEFAULT AS IDENTITY PRIMARY KEY,
                image_id         INTEGER NOT NULL,
                phase_id         INTEGER NOT NULL,
                status           VARCHAR(20) DEFAULT 'not_started' NOT NULL,
                executor_version VARCHAR(50),
                app_version      VARCHAR(50),
                job_id           INTEGER,
                attempt_count    SMALLINT DEFAULT 0 NOT NULL,
                last_error       BLOB SUB_TYPE TEXT,
                started_at       TIMESTAMP,
                finished_at      TIMESTAMP,
                updated_at       TIMESTAMP,
                skip_reason      BLOB SUB_TYPE TEXT,
                skipped_by       VARCHAR(255)
            )''')
            conn.commit()
            c = conn.cursor()

            # Unique constraint on (image_id, phase_id)
            if not _index_exists(c, 'UQ_IMAGE_PHASE'):
                c.execute("CREATE UNIQUE INDEX uq_image_phase ON image_phase_status(image_id, phase_id)")
                conn.commit()
                c = conn.cursor()

            # Performance indexes
            if not _index_exists(c, 'IDX_IPS_IMAGE_ID'):
                c.execute("CREATE INDEX idx_ips_image_id ON image_phase_status(image_id)")
                conn.commit()
                c = conn.cursor()
            if not _index_exists(c, 'IDX_IPS_PHASE_ID'):
                c.execute("CREATE INDEX idx_ips_phase_id ON image_phase_status(phase_id)")
                conn.commit()
                c = conn.cursor()
            if not _index_exists(c, 'IDX_IPS_STATUS'):
                c.execute("CREATE INDEX idx_ips_status ON image_phase_status(status)")
                conn.commit()
                c = conn.cursor()

        # IMAGE_PHASE_STATUS — add skip metadata columns if missing
        if _table_exists(c, 'IMAGE_PHASE_STATUS'):
            c.execute("SELECT RDB$FIELD_NAME FROM RDB$RELATION_FIELDS WHERE RDB$RELATION_NAME = 'IMAGE_PHASE_STATUS'")
            ips_cols = [row[0].strip().lower() for row in c.fetchall()]
            if 'skip_reason' not in ips_cols:
                try:
                    c.execute("ALTER TABLE image_phase_status ADD skip_reason BLOB SUB_TYPE TEXT")
                    conn.commit()
                    c = conn.cursor()
                except Exception as m:
                    logger.debug("Adding image_phase_status.skip_reason: %s", m)
                    try: conn.rollback()
                    except Exception: pass
                    c = conn.cursor()
            if 'skipped_by' not in ips_cols:
                try:
                    c.execute("ALTER TABLE image_phase_status ADD skipped_by VARCHAR(255)")
                    conn.commit()
                    c = conn.cursor()
                except Exception as m:
                    logger.debug("Adding image_phase_status.skipped_by: %s", m)
                    try: conn.rollback()
                    except Exception: pass
                    c = conn.cursor()

        # Foreign keys for IMAGE_PHASE_STATUS (safe add)
        if _table_exists(c, 'IMAGE_PHASE_STATUS'):
            if not _constraint_exists(c, 'FK_IPS_IMAGES'):
                try:
                    c.execute("DELETE FROM image_phase_status WHERE image_id NOT IN (SELECT id FROM images)")
                    conn.commit()
                    c = conn.cursor()
                    c.execute("ALTER TABLE image_phase_status ADD CONSTRAINT fk_ips_images FOREIGN KEY (image_id) REFERENCES images(id) ON DELETE CASCADE")
                    conn.commit()
                    c = conn.cursor()
                except Exception as fk_err:
                    logger.debug("FK_IPS_IMAGES: %s", fk_err)
                    try: conn.rollback()
                    except Exception: pass
                    c = conn.cursor()

            if not _constraint_exists(c, 'FK_IPS_PHASES'):
                try:
                    c.execute("ALTER TABLE image_phase_status ADD CONSTRAINT fk_ips_phases FOREIGN KEY (phase_id) REFERENCES pipeline_phases(id)")
                    conn.commit()
                    c = conn.cursor()
                except Exception as fk_err:
                    logger.debug("FK_IPS_PHASES: %s", fk_err)
                    try: conn.rollback()
                    except Exception: pass
                    c = conn.cursor()

        # JOBS — add phase_id column if missing
        if _table_exists(c, 'JOBS'):
            c.execute("SELECT RDB$FIELD_NAME FROM RDB$RELATION_FIELDS WHERE RDB$RELATION_NAME = 'JOBS'")
            jobs_cols = [row[0].strip().lower() for row in c.fetchall()]
            if "phase_id" not in jobs_cols:
                try:
                    c.execute("ALTER TABLE jobs ADD phase_id INTEGER")
                    conn.commit()
                    c = conn.cursor()
                except Exception as m:
                    logger.debug("Adding jobs.phase_id: %s", m)
                    try: conn.rollback()
                    except Exception: pass
                    c = conn.cursor()
            if "job_type" not in jobs_cols:
                try:
                    c.execute("ALTER TABLE jobs ADD job_type VARCHAR(50)")
                    conn.commit()
                    c = conn.cursor()
                except Exception as m:
                    logger.debug("Adding jobs.job_type: %s", m)
                    try: conn.rollback()
                    except Exception: pass
                    c = conn.cursor()
            if "current_phase" not in jobs_cols:
                try:
                    c.execute("ALTER TABLE jobs ADD current_phase VARCHAR(50)")
                    conn.commit()
                    c = conn.cursor()
                except Exception as m:
                    logger.debug("Adding jobs.current_phase: %s", m)
                    try: conn.rollback()
                    except Exception: pass
                    c = conn.cursor()
            if "next_phase_index" not in jobs_cols:
                try:
                    c.execute("ALTER TABLE jobs ADD next_phase_index INTEGER")
                    conn.commit()
                    c = conn.cursor()
                except Exception as m:
                    logger.debug("Adding jobs.next_phase_index: %s", m)
                    try: conn.rollback()
                    except Exception: pass
                    c = conn.cursor()
            if "runner_state" not in jobs_cols:
                try:
                    c.execute("ALTER TABLE jobs ADD runner_state VARCHAR(50)")
                    conn.commit()
                    c = conn.cursor()
                except Exception as m:
                    logger.debug("Adding jobs.runner_state: %s", m)
                    try: conn.rollback()
                    except Exception: pass
                    c = conn.cursor()
            if "queue_position" not in jobs_cols:
                try:
                    c.execute("ALTER TABLE jobs ADD queue_position INTEGER")
                    conn.commit()
                    c = conn.cursor()
                except Exception as m:
                    logger.debug("Adding jobs.queue_position: %s", m)
                    try: conn.rollback()
                    except Exception: pass
                    c = conn.cursor()
            if "enqueued_at" not in jobs_cols:
                try:
                    c.execute("ALTER TABLE jobs ADD enqueued_at TIMESTAMP")
                    conn.commit()
                    c = conn.cursor()
                except Exception as m:
                    logger.debug("Adding jobs.enqueued_at: %s", m)
                    try: conn.rollback()
                    except Exception: pass
                    c = conn.cursor()
            if "started_at" not in jobs_cols:
                try:
                    c.execute("ALTER TABLE jobs ADD started_at TIMESTAMP")
                    conn.commit()
                    c = conn.cursor()
                except Exception as m:
                    logger.debug("Adding jobs.started_at: %s", m)
                    try: conn.rollback()
                    except Exception: pass
                    c = conn.cursor()
            if "finished_at" not in jobs_cols:
                try:
                    c.execute("ALTER TABLE jobs ADD finished_at TIMESTAMP")
                    conn.commit()
                    c = conn.cursor()
                except Exception as m:
                    logger.debug("Adding jobs.finished_at: %s", m)
                    try: conn.rollback()
                    except Exception: pass
                    c = conn.cursor()
            if "cancel_requested" not in jobs_cols:
                try:
                    c.execute("ALTER TABLE jobs ADD cancel_requested SMALLINT DEFAULT 0")
                    conn.commit()
                    c = conn.cursor()
                except Exception as m:
                    logger.debug("Adding jobs.cancel_requested: %s", m)
                    try: conn.rollback()
                    except Exception: pass
                    c = conn.cursor()
            if "queue_payload" not in jobs_cols:
                try:
                    c.execute("ALTER TABLE jobs ADD queue_payload BLOB SUB_TYPE TEXT")
                    conn.commit()
                    c = conn.cursor()
                except Exception as m:
                    logger.debug("Adding jobs.queue_payload: %s", m)
                    try: conn.rollback()
                    except Exception: pass
                    c = conn.cursor()
            if "priority" not in jobs_cols:
                try:
                    c.execute("ALTER TABLE jobs ADD priority SMALLINT DEFAULT 100")
                    conn.commit()
                    c = conn.cursor()
                except Exception as m:
                    logger.debug("Adding jobs.priority: %s", m)
                    try: conn.rollback()
                    except Exception: pass
                    c = conn.cursor()
            if "retry_count" not in jobs_cols:
                try:
                    c.execute("ALTER TABLE jobs ADD retry_count INTEGER DEFAULT 0")
                    conn.commit()
                    c = conn.cursor()
                except Exception as m:
                    logger.debug("Adding jobs.retry_count: %s", m)
                    try: conn.rollback()
                    except Exception: pass
                    c = conn.cursor()
            if "target_scope" not in jobs_cols:
                try:
                    c.execute("ALTER TABLE jobs ADD target_scope VARCHAR(255)")
                    conn.commit()
                    c = conn.cursor()
                except Exception as m:
                    logger.debug("Adding jobs.target_scope: %s", m)
                    try: conn.rollback()
                    except Exception: pass
                    c = conn.cursor()
            if "paused_at" not in jobs_cols:
                try:
                    c.execute("ALTER TABLE jobs ADD paused_at TIMESTAMP")
                    conn.commit()
                    c = conn.cursor()
                except Exception as m:
                    logger.debug("Adding jobs.paused_at: %s", m)
                    try: conn.rollback()
                    except Exception: pass
                    c = conn.cursor()
            # FK for phase_id
            if not _constraint_exists(c, 'FK_JOBS_PHASES'):
                try:
                    c.execute("ALTER TABLE jobs ADD CONSTRAINT fk_jobs_phases FOREIGN KEY (phase_id) REFERENCES pipeline_phases(id)")
                    conn.commit()
                    c = conn.cursor()
                except Exception as fk_err:
                    logger.debug("FK_JOBS_PHASES: %s", fk_err)
                    try: conn.rollback()
                    except Exception: pass
                    c = conn.cursor()

            # Index on jobs.phase_id
            if not _index_exists(c, 'IDX_JOBS_PHASE_ID'):
                try:
                    c.execute("CREATE INDEX idx_jobs_phase_id ON jobs(phase_id)")
                    conn.commit()
                    c = conn.cursor()
                except Exception: pass
            if not _index_exists(c, 'IDX_JOBS_STATUS'):
                try:
                    c.execute("CREATE INDEX idx_jobs_status ON jobs(status)")
                    conn.commit()
                    c = conn.cursor()
                except Exception: pass
            if not _index_exists(c, 'IDX_JOBS_QUEUE_POSITION'):
                try:
                    c.execute("CREATE INDEX idx_jobs_queue_position ON jobs(queue_position)")
                    conn.commit()
                    c = conn.cursor()
                except Exception: pass
            if not _index_exists(c, 'IDX_JOBS_PRIORITY_STATUS'):
                try:
                    c.execute("CREATE INDEX idx_jobs_priority_status ON jobs(status, priority)")
                    conn.commit()
                    c = conn.cursor()
                except Exception:
                    pass
            if not _index_exists(c, 'IDX_JOBS_PAUSED_AT'):
                try:
                    c.execute("CREATE INDEX idx_jobs_paused_at ON jobs(paused_at)")
                    conn.commit()
                    c = conn.cursor()
                except Exception:
                    pass

        # JOB_PHASES migration / constraints / indexes
        if _table_exists(c, 'JOB_PHASES'):
            if not _constraint_exists(c, 'FK_JOB_PHASES_JOB'):
                try:
                    c.execute("DELETE FROM job_phases WHERE job_id NOT IN (SELECT id FROM jobs)")
                    conn.commit()
                    c = conn.cursor()
                    c.execute("ALTER TABLE job_phases ADD CONSTRAINT fk_job_phases_job FOREIGN KEY (job_id) REFERENCES jobs(id) ON DELETE CASCADE")
                    conn.commit()
                    c = conn.cursor()
                except Exception as fk_err:
                    logger.debug("FK_JOB_PHASES_JOB: %s", fk_err)
                    try: conn.rollback()
                    except Exception: pass
                    c = conn.cursor()

            if not _index_exists(c, 'IDX_JOB_PHASES_JOB_ID'):
                try:
                    c.execute("CREATE INDEX idx_job_phases_job_id ON job_phases(job_id)")
                    conn.commit()
                    c = conn.cursor()
                except Exception:
                    pass

            if not _index_exists(c, 'IDX_JOB_PHASES_ORDER'):
                try:
                    c.execute("CREATE INDEX idx_job_phases_order ON job_phases(job_id, phase_order)")
                    conn.commit()
                    c = conn.cursor()
                except Exception:
                    pass

        conn.commit()
    except Exception as e:
        logger.error("Pipeline phases migration error: %s", e)
        try: conn.rollback()
        except Exception: pass

    # --- Phase 1: Integrity + Index Hardening ---
    # Ref: docs/plans/database/DB_SCHEMA_REFACTOR_PLAN.md
    # All DDL is idempotent (check-then-act). Safe to run on every startup.
    try:
        c = conn.cursor()
        print("[Phase 1] Starting integrity + index hardening...")

        # 1.1: Repair orphan STACKS.BEST_IMAGE_ID rows
        print("  [1.1] Repairing orphan STACKS.BEST_IMAGE_ID...")
        try:
            c.execute("""
                UPDATE stacks SET best_image_id = NULL
                WHERE best_image_id IS NOT NULL
                  AND best_image_id NOT IN (SELECT id FROM images)
            """)
            conn.commit()
            c = conn.cursor()
        except Exception as e:
            logger.warning("Phase1 1.1 orphan repair: %s", e)
            try: conn.rollback()
            except Exception: pass
            c = conn.cursor()

        # 1.2: Unique index on IMAGES.FILE_PATH (highest-impact for upsert perf)
        print("  [1.2] Adding UQ_IMAGES_FILE_PATH (with temp index optimization)...")
        if not _index_exists(c, 'UQ_IMAGES_FILE_PATH'):
            try:
                # Optimization: create temporary non-unique index for fast de-dup grouping
                if not _index_exists(c, 'IDX_TMP_FILE_PATH'):
                    c.execute("CREATE INDEX idx_tmp_file_path ON images(file_path)")
                    conn.commit()
                    c = conn.cursor()

                # Safety: delete exact duplicate file_path rows (keep highest id)
                c.execute("""
                    DELETE FROM images i1
                    WHERE i1.file_path IS NOT NULL
                      AND EXISTS (
                        SELECT 1 FROM images i2
                        WHERE i2.file_path = i1.file_path
                        AND i2.id > i1.id
                    )
                """)
                conn.commit()
                c = conn.cursor()

                # Drop temporary index before creating the final UNIQUE one
                if _index_exists(c, 'IDX_TMP_FILE_PATH'):
                    c.execute("DROP INDEX idx_tmp_file_path")
                    conn.commit()
                    c = conn.cursor()

                c.execute("CREATE UNIQUE INDEX uq_images_file_path ON images(file_path)")
                conn.commit()
                c = conn.cursor()
            except Exception as e:
                logger.warning("Phase1 1.2 UQ_IMAGES_FILE_PATH: %s", e)
                try: conn.rollback()
                except Exception: pass
                # Emergency cleanup of temp index on failure
                try:
                    c = conn.cursor()
                    if _index_exists(c, 'IDX_TMP_FILE_PATH'):
                        c.execute("DROP INDEX idx_tmp_file_path")
                        conn.commit()
                except Exception: pass
                c = conn.cursor()

        # 1.3: Composite indexes for query hot paths
        print("  [1.3] Adding composite indexes (folder_score, stack_score)...")
        if not _index_exists(c, 'IDX_IMAGES_FOLDER_SCORE'):
            try:
                c.execute("CREATE INDEX idx_images_folder_score ON images(folder_id, score_general)")
                conn.commit()
                c = conn.cursor()
            except Exception as e:
                logger.warning("Phase1 1.3 IDX_IMAGES_FOLDER_SCORE: %s", e)
                try: conn.rollback()
                except Exception: pass
                c = conn.cursor()

        if not _index_exists(c, 'IDX_IMAGES_STACK_SCORE'):
            try:
                c.execute("CREATE INDEX idx_images_stack_score ON images(stack_id, score_general)")
                conn.commit()
                c = conn.cursor()
            except Exception as e:
                logger.warning("Phase1 1.3 IDX_IMAGES_STACK_SCORE: %s", e)
                try: conn.rollback()
                except Exception: pass
                c = conn.cursor()

        # 1.4a: Drop redundant single-column indexes (superseded by composites)
        print("  [1.4a] Dropping redundant single-column indexes...")
        for old_idx in ('IDX_IMAGES_FOLDER_ID', 'IDX_IMAGES_STACK_ID'):
            if _index_exists(c, old_idx):
                try:
                    c.execute(f"DROP INDEX {old_idx}")
                    conn.commit()
                    c = conn.cursor()
                except Exception as e:
                    logger.warning("Phase1 1.4a DROP INDEX %s: %s", old_idx, e)
                    try: conn.rollback()
                    except Exception: pass
                    c = conn.cursor()

        # 1.4b: Drop legacy auto-named FK artifacts on CULLING_PICKS
        # (INTEG_13, INTEG_14 — duplicates of FK_CULLING_PICKS_IMAGES/SESSIONS)
        print("  [1.4b] Dropping legacy CULLING_PICKS FK artifacts...")
        try:
            c.execute("""
                SELECT rdb$constraint_name FROM rdb$relation_constraints
                WHERE rdb$relation_name = 'CULLING_PICKS'
                  AND rdb$constraint_type = 'FOREIGN KEY'
                  AND rdb$constraint_name NOT STARTING WITH 'FK_'
            """)
            orphan_constraints = [row[0].strip() for row in c.fetchall()]
            for cn in orphan_constraints:
                try:
                    c.execute(f'ALTER TABLE culling_picks DROP CONSTRAINT "{cn}"')
                    conn.commit()
                    c = conn.cursor()
                except Exception as e:
                    logger.warning("Phase1 1.4b DROP CONSTRAINT %s: %s", cn, e)
                    try: conn.rollback()
                    except Exception: pass
                    c = conn.cursor()
        except Exception as e:
            logger.warning("Phase1 1.4b constraint query: %s", e)

        # 1.5a: FK_STACKS_BEST_IMAGE (after orphan repair in 1.1)
        print("  [1.5a] Adding FK_STACKS_BEST_IMAGE...")
        if not _constraint_exists(c, 'FK_STACKS_BEST_IMAGE'):
            try:
                c.execute("""
                    ALTER TABLE stacks ADD CONSTRAINT fk_stacks_best_image
                    FOREIGN KEY (best_image_id) REFERENCES images(id) ON DELETE SET NULL
                """)
                conn.commit()
                c = conn.cursor()
            except Exception as e:
                logger.warning("Phase1 1.5a FK_STACKS_BEST_IMAGE: %s", e)
                try: conn.rollback()
                except Exception: pass
                c = conn.cursor()

        # 1.5b: FK_IMAGES_JOB
        print("  [1.5b] Adding FK_IMAGES_JOB...")
        if not _constraint_exists(c, 'FK_IMAGES_JOB'):
            try:
                c.execute("UPDATE images SET job_id = NULL WHERE job_id IS NOT NULL AND job_id NOT IN (SELECT id FROM jobs)")
                conn.commit()
                c = conn.cursor()
                c.execute("""
                    ALTER TABLE images ADD CONSTRAINT fk_images_job
                    FOREIGN KEY (job_id) REFERENCES jobs(id) ON DELETE SET NULL
                """)
                conn.commit()
                c = conn.cursor()
            except Exception as e:
                logger.warning("Phase1 1.5b FK_IMAGES_JOB: %s", e)
                try: conn.rollback()
                except Exception: pass
                c = conn.cursor()

        # 1.5c: FK_IMAGES_STACK
        print("  [1.5c] Adding FK_IMAGES_STACK...")
        if not _constraint_exists(c, 'FK_IMAGES_STACK'):
            try:
                c.execute("UPDATE images SET stack_id = NULL WHERE stack_id IS NOT NULL AND stack_id NOT IN (SELECT id FROM stacks)")
                conn.commit()
                c = conn.cursor()
                c.execute("""
                    ALTER TABLE images ADD CONSTRAINT fk_images_stack
                    FOREIGN KEY (stack_id) REFERENCES stacks(id) ON DELETE SET NULL
                """)
                conn.commit()
                c = conn.cursor()
            except Exception as e:
                logger.warning("Phase1 1.5c FK_IMAGES_STACK: %s", e)
                try: conn.rollback()
                except Exception: pass
                c = conn.cursor()

        # 1.5d: FK_IPS_JOB (IMAGE_PHASE_STATUS.JOB_ID)
        print("  [1.5d] Adding FK_IPS_JOB...")
        if _table_exists(c, 'IMAGE_PHASE_STATUS') and not _constraint_exists(c, 'FK_IPS_JOB'):
            try:
                c.execute("UPDATE image_phase_status SET job_id = NULL WHERE job_id IS NOT NULL AND job_id NOT IN (SELECT id FROM jobs)")
                conn.commit()
                c = conn.cursor()
                c.execute("""
                    ALTER TABLE image_phase_status ADD CONSTRAINT fk_ips_job
                    FOREIGN KEY (job_id) REFERENCES jobs(id) ON DELETE SET NULL
                """)
                conn.commit()
                c = conn.cursor()
            except Exception as e:
                logger.warning("Phase1 1.5d FK_IPS_JOB: %s", e)
                try: conn.rollback()
                except Exception: pass
                c = conn.cursor()

        # 1.5e: STACK_CACHE FKs
        print("  [1.5e] Adding STACK_CACHE FK constraints...")
        if _table_exists(c, 'STACK_CACHE'):
            if not _constraint_exists(c, 'FK_STACK_CACHE_STACK'):
                try:
                    c.execute("DELETE FROM stack_cache WHERE stack_id NOT IN (SELECT id FROM stacks)")
                    conn.commit()
                    c = conn.cursor()
                    c.execute("""
                        ALTER TABLE stack_cache ADD CONSTRAINT fk_stack_cache_stack
                        FOREIGN KEY (stack_id) REFERENCES stacks(id) ON DELETE CASCADE
                    """)
                    conn.commit()
                    c = conn.cursor()
                except Exception as e:
                    logger.warning("Phase1 1.5e FK_STACK_CACHE_STACK: %s", e)
                    try: conn.rollback()
                    except Exception: pass
                    c = conn.cursor()

            if not _constraint_exists(c, 'FK_STACK_CACHE_REP_IMAGE'):
                try:
                    c.execute("UPDATE stack_cache SET rep_image_id = NULL WHERE rep_image_id IS NOT NULL AND rep_image_id NOT IN (SELECT id FROM images)")
                    conn.commit()
                    c = conn.cursor()
                    c.execute("""
                        ALTER TABLE stack_cache ADD CONSTRAINT fk_stack_cache_rep_image
                        FOREIGN KEY (rep_image_id) REFERENCES images(id) ON DELETE SET NULL
                    """)
                    conn.commit()
                    c = conn.cursor()
                except Exception as e:
                    logger.warning("Phase1 1.5e FK_STACK_CACHE_REP_IMAGE: %s", e)
                    try: conn.rollback()
                    except Exception: pass
                    c = conn.cursor()

            if not _constraint_exists(c, 'FK_STACK_CACHE_FOLDER'):
                try:
                    c.execute("UPDATE stack_cache SET folder_id = NULL WHERE folder_id IS NOT NULL AND folder_id NOT IN (SELECT id FROM folders)")
                    conn.commit()
                    c = conn.cursor()
                    c.execute("""
                        ALTER TABLE stack_cache ADD CONSTRAINT fk_stack_cache_folder
                        FOREIGN KEY (folder_id) REFERENCES folders(id) ON DELETE SET NULL
                    """)
                    conn.commit()
                    c = conn.cursor()
                except Exception as e:
                    logger.warning("Phase1 1.5e FK_STACK_CACHE_FOLDER: %s", e)
                    try: conn.rollback()
                    except Exception: pass
                    c = conn.cursor()

        try:
            n_folders = None
            n_stack_cache = None
            if _table_exists(c, "FOLDERS"):
                c.execute("SELECT COUNT(*) FROM folders")
                n_folders = c.fetchone()[0]
            if _table_exists(c, "STACK_CACHE"):
                c.execute("SELECT COUNT(*) FROM stack_cache")
                n_stack_cache = c.fetchone()[0]
            logger.info(
                "Phase1 after 1.5e row counts (for pause diagnosis): folders=%s stack_cache=%s",
                n_folders,
                n_stack_cache,
            )
        except Exception:
            pass

        # 1.5f: UQ_FOLDERS_PATH
        print("  [1.5f] Adding UQ_FOLDERS_PATH...")
        if not _index_exists(c, 'UQ_FOLDERS_PATH'):
            try:
                # De-dup safety net
                c.execute("""
                    DELETE FROM folders WHERE id NOT IN (
                        SELECT MAX(id) FROM folders GROUP BY path
                    )
                    AND path IN (
                        SELECT path FROM folders GROUP BY path HAVING COUNT(*) > 1
                    )
                """)
                conn.commit()
                c = conn.cursor()
                c.execute("CREATE UNIQUE INDEX uq_folders_path ON folders(path)")
                conn.commit()
                c = conn.cursor()
            except Exception as e:
                logger.warning("Phase1 1.5f UQ_FOLDERS_PATH: %s", e)
                try: conn.rollback()
                except Exception: pass
                c = conn.cursor()

        # 1.7: Recompute statistics for Firebird query planner
        print("  [1.7] Recomputing index statistics...")
        for idx_name in ('UQ_IMAGES_FILE_PATH', 'UQ_IMAGES_IMAGE_UUID',
                         'IDX_IMAGES_FOLDER_SCORE', 'IDX_IMAGES_STACK_SCORE',
                         'IDX_FOLDER_ID', 'IDX_STACK_ID', 'UQ_FOLDERS_PATH'):
            if _index_exists(c, idx_name):
                try:
                    c.execute(f"SET STATISTICS INDEX {idx_name}")
                    conn.commit()
                    c = conn.cursor()
                except Exception:
                    pass

        # 1.8: CHECK constraints for enum validation
        print("  [1.8] Adding CHECK constraints...")
        if not _constraint_exists(c, 'CHK_IMAGES_LABEL'):
            try:
                c.execute("""
                    ALTER TABLE images ADD CONSTRAINT chk_images_label
                    CHECK (label IS NULL OR label IN ('Red','Yellow','Green','Blue','Purple','None',''))
                """)
                conn.commit()
                c = conn.cursor()
            except Exception as e:
                logger.warning("Phase1 1.8 CHK_IMAGES_LABEL: %s", e)
                try: conn.rollback()
                except Exception: pass
                c = conn.cursor()

        # CHK_IMAGES_CULL_DECISION: allow 'neutral','maybe' (used by selection/culling)
        if _constraint_exists(c, 'CHK_IMAGES_CULL_DECISION'):
            try:
                c.execute("ALTER TABLE images DROP CONSTRAINT chk_images_cull_decision")
                conn.commit()
                c = conn.cursor()
            except Exception as e:
                logger.warning("Phase 1.8a DROP CHK_IMAGES_CULL_DECISION: %s", e)
                try: conn.rollback()
                except Exception: pass
                c = conn.cursor()
        if not _constraint_exists(c, 'CHK_IMAGES_CULL_DECISION'):
            try:
                c.execute("""
                    ALTER TABLE images ADD CONSTRAINT chk_images_cull_decision
                    CHECK (cull_decision IS NULL OR cull_decision IN ('pick','reject','skip','neutral','maybe',''))
                """)
                conn.commit()
                c = conn.cursor()
            except Exception as e:
                logger.warning("Phase1 1.8 CHK_IMAGES_CULL_DECISION: %s", e)
                try: conn.rollback()
                except Exception: pass
                c = conn.cursor()

        if _table_exists(c, 'IMAGE_PHASE_STATUS') and not _constraint_exists(c, 'CHK_IPS_STATUS'):
            try:
                c.execute("""
                    ALTER TABLE image_phase_status ADD CONSTRAINT chk_ips_status
                    CHECK (status IN ('not_started','pending','running','done','failed','skipped'))
                """)
                conn.commit()
                c = conn.cursor()
            except Exception as e:
                logger.warning("Phase1 1.8 CHK_IPS_STATUS: %s", e)
                try: conn.rollback()
                except Exception: pass
                c = conn.cursor()

        print("[Phase 1] OK - Complete (integrity + index hardening).")
        logger.info("Phase 1 migration complete (integrity + index hardening).")
    except Exception as e:
        logger.error("Phase 1 migration error: %s", e)
        try: conn.rollback()
        except Exception: pass

    # --- Phase 2: Keyword Normalization + IMAGE_XMP Backfill ---
    try:
        c = conn.cursor()
        print("[Phase 2] Starting Keyword Normalization + IMAGE_XMP Backfill...")

        # 2.1a: KEYWORDS_DIM table
        if not _table_exists(c, 'KEYWORDS_DIM'):
            print("  [2.1a] Creating KEYWORDS_DIM table...")
            try:
                c.execute('''CREATE TABLE keywords_dim (
                    keyword_id      INTEGER GENERATED BY DEFAULT AS IDENTITY PRIMARY KEY,
                    keyword_norm    VARCHAR(200) NOT NULL,
                    keyword_display VARCHAR(200),
                    created_at      TIMESTAMP DEFAULT CURRENT_TIMESTAMP
                )''')
                conn.commit()
                c = conn.cursor()

                c.execute("CREATE UNIQUE INDEX uq_keywords_dim_norm ON keywords_dim(keyword_norm)")
                conn.commit()
                c = conn.cursor()
            except Exception as e:
                logger.warning("KEYWORDS_DIM table creation: %s", e)
                try: conn.rollback()
                except Exception: pass
                c = conn.cursor()

        # 2.1b: IMAGE_KEYWORDS junction table
        if not _table_exists(c, 'IMAGE_KEYWORDS'):
            print("  [2.1b] Creating IMAGE_KEYWORDS table...")
            try:
                c.execute('''CREATE TABLE image_keywords (
                    image_id    INTEGER NOT NULL,
                    keyword_id  INTEGER NOT NULL,
                    source      VARCHAR(20) DEFAULT 'auto',
                    confidence  DOUBLE PRECISION,
                    created_at  TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    PRIMARY KEY (image_id, keyword_id)
                )''')
                conn.commit()
                c = conn.cursor()

                c.execute("""
                    ALTER TABLE image_keywords ADD CONSTRAINT fk_imgkw_image
                    FOREIGN KEY (image_id) REFERENCES images(id) ON DELETE CASCADE
                """)
                conn.commit()
                c = conn.cursor()

                c.execute("""
                    ALTER TABLE image_keywords ADD CONSTRAINT fk_imgkw_keyword
                    FOREIGN KEY (keyword_id) REFERENCES keywords_dim(keyword_id) ON DELETE CASCADE
                """)
                conn.commit()
                c = conn.cursor()

                c.execute("CREATE INDEX idx_imgkw_image_id ON image_keywords(image_id)")
                conn.commit()
                c = conn.cursor()

                c.execute("CREATE INDEX idx_imgkw_keyword_id ON image_keywords(keyword_id)")
                conn.commit()
                c = conn.cursor()
            except Exception as e:
                logger.warning("IMAGE_KEYWORDS table creation: %s", e)
                try: conn.rollback()
                except Exception: pass
                c = conn.cursor()

        # Call backfill after tables are created
        if _table_exists(c, 'KEYWORDS_DIM') and _table_exists(c, 'IMAGE_KEYWORDS'):
            conn.close() 
            _backfill_keywords()
            conn = get_db()
            c = conn.cursor()

        # Call IMAGE_XMP backfill as per 2.6
        if _table_exists(c, 'IMAGE_XMP'):
            conn.close()
            _backfill_image_xmp()
            conn = get_db()
            c = conn.cursor()

    except Exception as e:
        logger.error("Phase 2 table creation error: %s", e)
        try: conn.rollback()
        except Exception: pass

    # Seed phases
    try:
        conn.close()
    except Exception:
        pass
    seed_pipeline_phases()

    # Re-open for final commit check
    try:
        conn = get_db()
        conn.commit()
    except Exception as e:
        logger.warning("Final commit failed: %s", e)
    finally:
        try:
            conn.close()
        except Exception:
            pass

def get_image_by_hash(image_hash):
    row = get_connector().query_one(
        "SELECT * FROM images WHERE image_hash = ?", (image_hash,)
    )
    if not row:
        return None
    data = dict(row)
    data['file_paths'] = get_all_paths(data['id'])
    return data


def update_image_field(image_id: int, field_name: str, value) -> bool:
    """
    Update a single field for an image by ID.
    
    Args:
        image_id: ID of the image to update
        field_name: Name of the column to update (must be a valid column)
        value: New value for the field
        
    Returns:
        True if successful
    """
    # Whitelist valid field names to prevent SQL injection
    valid_fields = {
        'burst_uuid', 'rating', 'label', 'score_general', 'score_aesthetic',
        'score_technical', 'keywords', 'title', 'description', 'stack_id',
        'thumbnail_path', 'thumbnail_path_win', 'metadata', 'image_hash',
        'cull_decision', 'cull_policy_version', 'image_uuid'
    }
    
    if field_name not in valid_fields:
        logging.warning(f"Invalid field name for update: {field_name}")
        return False
    
    try:
        # field_name is safe: validated against whitelist above
        get_connector().execute(
            f"UPDATE images SET {field_name} = ? WHERE id = ?", (value, image_id)
        )

        # Dual-write: sync normalized keyword tables
        if field_name == 'keywords':
            try:
                _sync_image_keywords(image_id, value)
            except Exception as e:
                logging.warning(f"Keyword sync failed for image {image_id}: {e}")

        # Broadcast image update
        try:
            from modules.events import event_manager
            event_manager.broadcast_threadsafe("image_updated", {
                "image_id": image_id,
                "field": field_name,
                "value": value
            })
        except Exception: pass

        return True
    except Exception as e:
        logging.error(f"Failed to update {field_name} for image {image_id}: {e}")
        return False


def update_image_path(image_hash, new_path):
    from pathlib import Path
    new_name = Path(new_path).name
    try:
        def _tx(tx):
            tx.execute(
                "UPDATE images SET file_path = ?, file_name = ? WHERE image_hash = ?",
                (new_path, new_name, image_hash),
            )
            row = tx.query_one("SELECT id FROM images WHERE image_hash = ?", (image_hash,))
            if row:
                tx.execute(
                    "UPDATE OR INSERT INTO file_paths (image_id, path, last_seen) VALUES (?, ?, ?) MATCHING (image_id, path)",
                    (row["id"], new_path, datetime.datetime.now()),
                )
        get_connector().run_transaction(_tx)
        # Post-update folder fix
        try:
            update_image_folder_id(image_hash=image_hash)
        except Exception: pass
        return True
    except Exception as e:
        logging.error(f"Failed to update path for hash {image_hash}: {e}")
        return False


def update_image_uuid(image_id: int, image_uuid: str) -> bool:
    """Update the IMAGE_UUID for a specific image."""
    return update_image_field(image_id, "image_uuid", image_uuid)


def update_image_folder_id(image_hash=None, image_id=None):
    """
    Helper to update folder_id for a single image.
    """
    try:
        if image_hash:
            row = get_connector().query_one(
                "SELECT id, file_path, folder_id FROM images WHERE image_hash = ?", (image_hash,)
            )
        elif image_id:
            row = get_connector().query_one(
                "SELECT id, file_path, folder_id FROM images WHERE id = ?", (image_id,)
            )
        else:
            return
        if row and row["file_path"]:
            img_id = row["id"]
            old_folder_id = row.get("folder_id")
            dirname = os.path.normpath(os.path.dirname(row["file_path"]))
            fid = get_or_create_folder(dirname)
            get_connector().execute("UPDATE images SET folder_id = ? WHERE id = ?", (fid, img_id))
            if fid:
                invalidate_folder_phase_aggregates(folder_id=fid)
            if old_folder_id and old_folder_id != fid:
                invalidate_folder_phase_aggregates(folder_id=old_folder_id)
    except Exception as e:
        logger.warning("Error updating folder_id: %s", e)

def register_image_path(image_id, path):
    """
    Registers a path for a given image ID (default type='WSL').
    """
    path_type = 'WSL'
    try:
        get_connector().execute(
            "UPDATE OR INSERT INTO file_paths (image_id, path, path_type, last_seen) VALUES (?, ?, ?, ?) MATCHING (image_id, path)",
            (image_id, path, path_type, datetime.datetime.now()),
        )
    except Exception as e:
        logging.error(f"Failed to register path {path} for image {image_id}: {e}")



def get_all_paths(image_id):
    rows = get_connector().query("SELECT path FROM file_paths WHERE image_id = ?", (image_id,))
    return [r["path"] for r in rows]


# --- Resolved Paths (Windows Native Viewer Support) ---

def _convert_to_windows_path(path):
    r"""
    Convert any path format to Windows format.
    Handles WSL paths (/mnt/d/...) -> (D:\...).
    """
    if not path:
        return None
    
    # Already Windows format?
    if len(path) >= 2 and path[1] == ':':
        # Normalize slashes
        return path.replace('/', '\\')
    
    # WSL format?
    path_normalized = path.replace('\\', '/')
    if path_normalized.startswith('/mnt/'):
        parts = path_normalized.split('/')
        if len(parts) > 2 and len(parts[2]) == 1:
            drive = parts[2].upper()
            rest = '\\'.join(parts[3:])
            return f"{drive}:\\{rest}"
    
    # Unknown format, return as-is with backslashes
    return path.replace('/', '\\')


def resolve_windows_path(image_id, wsl_path, verify=True):
    """
    Resolves a WSL/Unix path to Windows format and stores in file_paths (type='WIN').
    """
    import platform
    
    windows_path = _convert_to_windows_path(wsl_path)
    if not windows_path:
        return None
    
    # Verify file exists (if on Windows)
    is_verified = 0
    verification_date = None
    now = datetime.datetime.now()
    
    if verify and platform.system() == 'Windows':
        import os
        if os.path.exists(windows_path):
            is_verified = 1
            verification_date = now

    if image_id is None:
        return windows_path

    try:
        row = get_connector().query_one(
            "SELECT id FROM file_paths WHERE image_id = ? AND path_type = 'WIN'", (image_id,))
        if row:
            get_connector().execute(
                "UPDATE file_paths SET path = ?, is_verified = ?, verification_date = ?, last_seen = ? WHERE id = ?",
                (windows_path, is_verified, verification_date, now, row["id"]))
        else:
            get_connector().execute(
                "INSERT INTO file_paths (image_id, path, path_type, is_verified, verification_date, last_seen) VALUES (?, ?, 'WIN', ?, ?, ?)",
                (image_id, windows_path, is_verified, verification_date, now))
        return windows_path
    except Exception as e:
        logging.error(f"Failed to resolve path for image {image_id}: {e}")
        return None


def get_resolved_path(image_id, verified_only=True):
    """
    Returns the Windows path for an image from file_paths (type='WIN').
    """
    query = "SELECT path FROM file_paths WHERE image_id = ? AND path_type = 'WIN'"
    if verified_only:
        query += " AND is_verified = 1"
    row = get_connector().query_one(query, (image_id,))
    return row["path"] if row else None


def verify_resolved_path(image_id):
    """
    Verifies that a resolved path still exists on disk.
    """
    import platform
    import os

    if platform.system() != 'Windows':
        return False

    row = get_connector().query_one(
        "SELECT id, path FROM file_paths WHERE image_id = ? AND path_type = 'WIN'", (image_id,))
    if not row:
        return False

    rp_id = row["id"]
    windows_path = row["path"]
    now = datetime.datetime.now()
    exists = os.path.exists(windows_path)

    get_connector().execute(
        "UPDATE file_paths SET is_verified = ?, verification_date = ?, last_seen = ? WHERE id = ?",
        (1 if exists else 0, now if exists else None, now, rp_id))
    return exists

def get_resolved_paths_batch(image_ids):
    """
    Get resolved Windows paths for a batch of image IDs.
    Returns a dictionary mapping image_id -> windows_path.
    Only returns verified paths that are correctly formatted.
    """
    if not image_ids:
        return {}

    placeholders = ','.join(['?'] * len(image_ids))
    query = f"SELECT image_id, path FROM file_paths WHERE image_id IN ({placeholders}) AND path_type = 'WIN' AND is_verified = 1"
    rows = get_connector().query(query, tuple(image_ids))

    result = {}
    for row in rows:
        path = row["path"]
        # Ensure path is truly Windows format (backslashes)
        # Mixed separators can cause "file not found" in some Windows APIs
        if path and '/' not in path:
            result[row["image_id"]] = path
    return result

def get_folder_by_id(folder_id):
    row = get_connector().query_one("SELECT path FROM folders WHERE id = ?", (folder_id,))
    return row["path"] if row else None


def find_image_id_by_path(file_path):
    """Returns image id if exists by file_path, else None."""
    row = get_connector().query_one("SELECT id FROM images WHERE file_path = ?", (file_path,))
    return row["id"] if row else None


def find_image_id_by_uuid(image_uuid):
    """Returns image id if exists by image_uuid, else None."""
    if not image_uuid or not isinstance(image_uuid, str) or not image_uuid.strip():
        return None
    row = get_connector().query_one("SELECT id FROM images WHERE image_uuid = ?", (image_uuid.strip(),))
    return row["id"] if row else None


def _get_folder_ancestor_ids(folder_id):
    """Return folder_id plus all parents up to root."""
    if not folder_id:
        return []

    seen = set()
    ids = []
    current = folder_id
    while current and current not in seen:
        seen.add(current)
        ids.append(current)
        row = get_connector().query_one("SELECT parent_id FROM folders WHERE id = ?", (current,))
        current = row["parent_id"] if row else None
    return ids


def invalidate_folder_phase_aggregates(folder_id=None, folder_path=None):
    """
    Mark phase aggregate cache dirty for the target folder and all its parents.
    """
    if not folder_id and folder_path:
        folder_id = get_or_create_folder(folder_path)

    ancestor_ids = _get_folder_ancestor_ids(folder_id)
    if not ancestor_ids:
        return

    placeholders = ",".join(["?"] * len(ancestor_ids))
    get_connector().execute(
        f"UPDATE folders SET phase_agg_dirty = 1 WHERE id IN ({placeholders})",
        tuple(ancestor_ids)
    )


def register_image_for_import(file_path, file_name, file_type, folder_id, image_uuid=None):
    """
    Insert a minimal image record for import (no scoring).
    Returns (image_id, was_new): image_id on success, None on failure; was_new True if inserted, False if already existed.
    On duplicate (UQ_IMAGES_FILE_PATH), returns the existing image_id with was_new=False.
    """
    try:
        rows = get_connector().execute_returning(
            """INSERT INTO images (file_path, file_name, file_type, folder_id, image_uuid, created_at)
               VALUES (?, ?, ?, ?, ?, CURRENT_TIMESTAMP) RETURNING id""",
            (file_path, file_name, file_type, folder_id, image_uuid or None)
        )
        image_id = rows[0]["id"] if rows else None
        if image_id:
            register_image_path(image_id, file_path)
            try:
                resolve_windows_path(image_id, file_path, verify=False)
            except Exception:
                pass
        return (image_id, True)
    except Exception as e:
        err_str = str(e)
        if "UQ_IMAGES_FILE_PATH" in err_str or "duplicate value" in err_str.lower():
            existing_id = find_image_id_by_path(file_path)
            if existing_id:
                return (existing_id, False)
        logger.warning("register_image_for_import failed for %s: %s", file_path, e)
        return (None, False)


def get_or_create_folder(folder_path, _depth=0):
    """
    Gets folder ID from cache/DB, creating it if it doesn't exist.
    Recursively creates parent folders to establish hierarchy.
    """
    # Normalize path
    if not folder_path or folder_path == '.':
        folder_path = os.getcwd()

    folder_path = os.path.normpath(folder_path)
    
    # Auto-convert Windows paths to WSL if we are on Windows but DB has WSL paths
    # This is critical because scoring runs in WSL (saving /mnt/d/...) 
    # but UI runs in Windows (sending D:\...)
    # This MUST happen before os.path.abspath to prevent Linux from prepending cwd to Windows paths
    try:
        from modules import utils
        if ":" in folder_path or "\\" in folder_path:
             wsl_path = utils.convert_path_to_wsl(folder_path)
             if wsl_path != folder_path:
                 logging.debug(f"Converted {folder_path} to {wsl_path}")
                 folder_path = wsl_path
    except ImportError:
        pass

    import posixpath
    
    # Check if this is a WSL path (starts with /mnt/)
    # On Windows, os.path.abspath will mangle it (D:\mnt\...)
    normalized_unix = folder_path.replace('\\', '/')
    is_wsl_path = normalized_unix.startswith('/mnt/') or normalized_unix == '/mnt' or normalized_unix == '/'
    
    if not is_wsl_path:
        folder_path = os.path.abspath(folder_path)
        folder_path = os.path.normpath(folder_path)
    else:
        folder_path = folder_path.replace('\\', '/')
        folder_path = posixpath.normpath(folder_path)

    # Sanity check for recursive loops (e.g. /mnt/d/mnt/d/...)
    if "/mnt/d/mnt/d" in folder_path:
        logging.error(f"Refusing to create recursive folder path: {folder_path}")
        return None
        
    # Check depth
    if folder_path.count('/') > 15:
         logging.warning(f"Deep folder path detected: {folder_path} (Depth: {folder_path.count('/')})")
    
    # Base case for recursion / root check
    # On Windows, os.path.dirname("D:\\") is "D:\\". 
    # Stop if parent is same as current or empty.
    
    import posixpath
    if folder_path.startswith('/mnt/'):
        parent_path = posixpath.dirname(folder_path)
    else:
        parent_path = os.path.dirname(folder_path)
    
    if not parent_path or parent_path == folder_path:
        # It's a root or top level?
        # Just create/get it with no parent.
        parent_id = None
    else:
        # Recursive call to get/create parent first
        # We assume this won't be too deep (max 10-20 levels typically)
        parent_id = get_or_create_folder(parent_path, _depth=_depth+1)

    try:
        row = get_connector().query_one("SELECT id, parent_id FROM folders WHERE path = ?", (folder_path,))
        if row:
            # Check if parent_id needs update (if it was created without parent before)
            curr_pid = row["parent_id"]
            if parent_id and curr_pid != parent_id:
                get_connector().execute("UPDATE folders SET parent_id = ? WHERE id = ?", (parent_id, row["id"]))
            return row["id"]

        rows = get_connector().execute_returning(
            "INSERT INTO folders (path, parent_id, created_at) VALUES (?, ?, ?) RETURNING id",
            (folder_path, parent_id, datetime.datetime.now()))
        created_id = rows[0]["id"] if rows else None
        if created_id:
            invalidate_folder_phase_aggregates(folder_id=created_id)
        return created_id
    except Exception as e:
        # Race condition or error — retry select
        try:
            row = get_connector().query_one("SELECT id FROM folders WHERE path = ?", (folder_path,))
            if row:
                return row["id"]
        except Exception:
            pass
        logging.error(f"Error getting/creating folder {folder_path}: {e}")
        return None

def rebuild_folder_cache():
    """
    Scans all images, populates folders table with full hierarchy, and updates images.folder_id.
    """
    invalidate_folder_images_cache()
    print("Rebuilding folder cache with hierarchy...")

    # 1. Get all unique folder paths from images
    rows = get_connector().query("SELECT DISTINCT file_path FROM images")

    unique_dirs = set()
    for row in rows:
        fp = row["file_path"]
        if fp:
            unique_dirs.add(os.path.dirname(fp))

    sorted_dirs = sorted(list(unique_dirs))

    # 2. Iterate and create folders (recursive logic in get_or_create_folder handles hierarchy)
    print(f"Found {len(sorted_dirs)} unique image directories. Processing hierarchy...")

    folder_map = {}  # path -> id
    for d in sorted_dirs:
        fid = get_or_create_folder(d)
        if fid:
            folder_map[d] = fid

    # 3. Update images folder_id
    print("Updating image folder_ids...")
    img_rows = get_connector().query("SELECT id, file_path FROM images")

    batch = []
    count = 0
    for row in img_rows:
        img_id = row["id"]
        path = row["file_path"]
        if not path:
            continue
        d = os.path.normpath(os.path.dirname(path))
        fid = folder_map.get(d)
        if fid:
            batch.append((fid, img_id))

        if len(batch) >= 1000:
            get_connector().execute_many("UPDATE images SET folder_id = ? WHERE id = ?", batch)
            count += len(batch)
            batch = []

    if batch:
        get_connector().execute_many("UPDATE images SET folder_id = ? WHERE id = ?", batch)
        count += len(batch)

    msg = f"Folder cache rebuild complete. Processed {len(sorted_dirs)} folders, updated {count} images."
    print(msg)
    return msg



def set_folder_scored(folder_path, is_scored=True):
    folder_id = get_or_create_folder(folder_path)
    if not folder_id: return
    get_connector().execute(
        "UPDATE folders SET is_fully_scored = ? WHERE id = ?",
        (1 if is_scored else 0, folder_id),
    )
    
    # Broadcast folder update
    try:
        from modules.events import event_manager
        event_manager.broadcast_threadsafe("folder_updated", {
            "path": folder_path,
            "is_fully_scored": is_scored
        })
    except Exception: pass

def is_folder_scored(folder_path):
    folder_id = get_or_create_folder(folder_path)
    if not folder_id: return False
    row = get_connector().query_one("SELECT is_fully_scored FROM folders WHERE id = ?", (folder_id,))
    return bool(row and row["is_fully_scored"])

def check_and_update_folder_status(folder_path):
    """
    Verifies if all images in a folder have valid scores in the DB.
    If so, sets is_fully_scored = 1.
    """
    # 1. List files in folder
    path = Path(folder_path)
    if not path.exists() or not path.is_dir():
        return False
        
    extensions = {'.jpg', '.jpeg', '.png', '.nef', '.nrw', '.dng', '.cr2', '.arw'}
    
    try:
        files = [p for p in path.iterdir() if p.is_file() and p.suffix.lower() in extensions]
    except (PermissionError, OSError) as e:
        print(f"Error accessing folder {folder_path}: {e}")
        return False

    if not files:
        # Empty folder is "fully scored" -> mark it done.
        set_folder_scored(folder_path, True)
        return True

    # 2. Check DB for these files using folder_id
    folder_id = get_or_create_folder(folder_path)
    if not folder_id: return False

    rows = list(get_connector().query("SELECT file_name, score_general FROM images WHERE folder_id = ?", (folder_id,)))

    # helper set of scored filenames
    scored_files = {row['file_name'] for row in rows if row['score_general'] and row['score_general'] > 0}
            
    # 3. Compare
    all_scored = True
    for f in files:
        if f.name not in scored_files:
            all_scored = False
            break
            
    # Update status
    set_folder_scored(folder_path, all_scored)
        
    return all_scored


def get_all_folders():
    """
    Returns a sorted list of all unique folder paths from the folders table.
    Does NOT auto-rebuild to avoid blocking the UI - use rebuild_folder_cache() explicitly.
    """
    rows = get_connector().query("SELECT path FROM folders ORDER BY path")
    return [r["path"] for r in rows]


def backfill_folder_phase_aggregates(limit=None):
    """
    Maintenance helper to recalculate folder phase aggregate caches.
    Marks all folders dirty, then recomputes deepest folders first.
    """
    get_connector().execute("UPDATE folders SET phase_agg_dirty = 1")
    rows = list(get_connector().query("SELECT path FROM folders ORDER BY CHAR_LENGTH(path) DESC"))

    paths = [r["path"] for r in rows if r and r["path"]]
    if isinstance(limit, int) and limit > 0:
        paths = paths[:limit]

    recomputed = 0
    for path in paths:
        get_folder_phase_summary(path)
        recomputed += 1

    return {"recomputed": recomputed, "total": len(paths)}


def delete_folder_cache_entry(folder_path: str, delete_descendants: bool = True) -> dict:
    """
    Delete a folder record from the `folders` table (folder tree cache).

    This is intended for removing stale/incorrect folder cache entries that appear
    in the Folder Tree UI.

    Behavior:
    - Deletes the matching folder row(s) (supports Windows or WSL path input)
    - Optionally deletes all descendant folders (via parent_id traversal)
    - Clears `images.folder_id` for images referencing deleted folders
    - Attempts to delete matching `cluster_progress` rows (best-effort)

    Returns:
        dict with keys: success (bool), message (str), deleted_folders (int)
    """
    if not folder_path or not str(folder_path).strip():
        return {"success": False, "message": "No folder path provided.", "deleted_folders": 0}

    # Prepare candidate path representations for lookup (DB may store WSL paths)
    raw = str(folder_path).strip()
    candidates: list[str] = []
    try:
        from modules import utils
        # If it looks like a Windows path, convert to WSL (DB convention)
        if ":" in raw or "\\" in raw:
            wsl = utils.convert_path_to_wsl(raw)
            if wsl and wsl != raw:
                candidates.append(wsl)
            candidates.append(os.path.normpath(raw))
        else:
            candidates.append(raw)
    except Exception:
        candidates.append(raw)

    # De-dup candidates while keeping order
    seen = set()
    candidates = [p for p in candidates if not (p in seen or seen.add(p))]

    def _tx(tx):
        # Find starting folder IDs
        start_ids: list[int] = []
        start_paths: list[str] = []
        for cand in candidates:
            try:
                row = tx.query_one("SELECT id, path FROM folders WHERE path = ?", (cand,))
                if row:
                    start_ids.append(int(row["id"]))
                    start_paths.append(str(row["path"]))
            except Exception:
                continue

        if not start_ids:
            return {"success": False, "message": f"Folder not found in cache: {raw}", "deleted_folders": 0}

        # Traverse descendants by parent_id to get full delete set
        ids_to_delete: list[int] = []
        paths_to_delete: list[str] = []

        queue = list(dict.fromkeys(start_ids))
        while queue:
            batch_ids = queue[:200]
            queue = queue[200:]

            # Add current batch
            for _id in batch_ids:
                if _id not in ids_to_delete:
                    ids_to_delete.append(_id)

            # Collect their paths
            placeholders = ",".join(["?"] * len(batch_ids))
            for r in tx.query(f"SELECT id, path FROM folders WHERE id IN ({placeholders})", tuple(batch_ids)):
                try:
                    _p = str(r["path"])
                    if _p not in paths_to_delete:
                        paths_to_delete.append(_p)
                except Exception:
                    pass

            if not delete_descendants:
                continue

            # Find children
            child_rows = tx.query(f"SELECT id FROM folders WHERE parent_id IN ({placeholders})", tuple(batch_ids))
            for r in child_rows:
                try:
                    cid = int(r["id"])
                    if cid not in ids_to_delete and cid not in queue:
                        queue.append(cid)
                except Exception:
                    continue

        if not ids_to_delete:
            return {"success": False, "message": f"Nothing to delete for: {raw}", "deleted_folders": 0}

        # Clear image folder_id references first (avoid dangling references)
        placeholders = ",".join(["?"] * len(ids_to_delete))
        tx.execute(f"UPDATE images SET folder_id = NULL WHERE folder_id IN ({placeholders})", tuple(ids_to_delete))

        # Best-effort cleanup for cluster_progress rows
        try:
            if paths_to_delete:
                cp_ph = ",".join(["?"] * len(paths_to_delete))
                tx.execute(f"DELETE FROM cluster_progress WHERE folder_path IN ({cp_ph})", tuple(paths_to_delete))
        except Exception:
            pass

        # Delete folders (children first to be safe if FK constraints are added later)
        for fid in reversed(ids_to_delete):
            tx.execute("DELETE FROM folders WHERE id = ?", (fid,))

        return {"ids_deleted": ids_to_delete, "paths_deleted": paths_to_delete}

    try:
        result = get_connector().run_transaction(_tx)
        if isinstance(result, dict) and "success" in result:
            return result  # Early exit (not found)

        ids_deleted = result["ids_deleted"]
        paths_deleted = result["paths_deleted"]

        # Broadcast folder deletions (outside transaction)
        try:
            from modules.events import event_manager
            for path in paths_deleted:
                event_manager.broadcast_threadsafe("folder_deleted", {"path": path})
        except Exception: pass

        return {
            "success": True,
            "message": f"Deleted {len(ids_deleted)} folder cache record(s).",
            "deleted_folders": len(ids_deleted),
        }
    except Exception as e:
        logging.error(f"delete_folder_cache_entry failed for {folder_path}: {e}")
        return {"success": False, "message": f"Error deleting folder cache entry: {e}", "deleted_folders": 0}

_folder_images_cache = {}
_FOLDER_CACHE_TTL = 30  # seconds


def invalidate_folder_images_cache(folder_path=None):
    """Clear cached get_images_by_folder results.

    Args:
        folder_path: Specific folder to invalidate, or None to clear all.
    """
    if folder_path:
        _folder_images_cache.pop(os.path.normpath(folder_path), None)
    else:
        _folder_images_cache.clear()


def get_images_by_folder(folder_path):
    """
    Returns all images located immediately in the specified folder using folder_id.
    Results are cached for up to _FOLDER_CACHE_TTL seconds to avoid redundant
    DB round-trips (e.g. folder tree selection followed by "Open in..." navigation).
    """
    folder_path = os.path.normpath(folder_path)

    now = time.time()
    cached = _folder_images_cache.get(folder_path)
    if cached is not None:
        cached_time, cached_rows = cached
        if now - cached_time < _FOLDER_CACHE_TTL:
            return cached_rows
        del _folder_images_cache[folder_path]

    folder_id = get_or_create_folder(folder_path)

    if not folder_id:
        return []

    result = list(get_connector().query(
        "SELECT * FROM images WHERE folder_id = ? ORDER BY file_name", (folder_id,)
    ))
    _folder_images_cache[folder_path] = (now, result)
    return result


def get_images_with_keyword(folder_path=None, keyword="birds", resolved_image_ids=None):
    """Return image rows that have a specific keyword (via normalized keyword tables).

    Scope priority: resolved_image_ids > folder_path > all DB images.
    Used by BirdSpeciesRunner to restrict processing to images tagged with 'birds'.
    """
    conditions = []
    params = []

    # Always filter by keyword using the normalized junction tables
    conditions.append(
        "EXISTS (SELECT 1 FROM image_keywords ik "
        "JOIN keywords_dim kd ON ik.keyword_id = kd.keyword_id "
        "WHERE ik.image_id = images.id AND kd.keyword_norm LIKE ?)"
    )
    params.append(f"%{keyword.strip().lower()}%")

    # When resolved_image_ids is large, skip the IN clause (Firebird ~900 param limit)
    # and post-filter in Python instead.
    resolved_ids_set = None
    if resolved_image_ids is not None:
        if not resolved_image_ids:
            return []
        if len(resolved_image_ids) <= 900:
            placeholders = ",".join("?" * len(resolved_image_ids))
            conditions.append(f"images.id IN ({placeholders})")
            params.extend(int(i) for i in resolved_image_ids)
        else:
            resolved_ids_set = set(int(i) for i in resolved_image_ids)
    elif folder_path:
        norm_path = os.path.normpath(folder_path)
        folder_id = get_or_create_folder(norm_path)
        if not folder_id:
            return []
        conditions.append("images.folder_id = ?")
        params.append(folder_id)

    where = "WHERE " + " AND ".join(conditions)
    rows = get_connector().query(f"SELECT * FROM images {where} ORDER BY file_name", tuple(params))
    result = [dict(row) for row in rows]
    if resolved_ids_set is not None:
        result = [r for r in result if r["id"] in resolved_ids_set]
    return result


def create_job(input_path, phase_code=None, job_type=None, status="pending", current_phase=None,
               next_phase_index=None, runner_state=None, queue_payload=None):
    """
    Create a new job record.

    Args:
        input_path: Path being processed.
        phase_code: Optional phase code (e.g. 'scoring') — resolves to phase_id FK.
        job_type:   Optional legacy job type string (deprecated, use phase_code).
        status:     Initial status (default: pending).
        current_phase: Current orchestrator phase code.
        next_phase_index: Next phase index in orchestrator order.
        runner_state: High-level runner/orchestrator state.
        queue_payload: Optional queue metadata payload persisted as JSON.
    """
    phase_id = None
    if phase_code:
        phase_id = get_phase_id(phase_code)
        if job_type is None:
            job_type = phase_code  # backfill legacy column

    now = datetime.datetime.now()
    payload_json = json.dumps(queue_payload) if queue_payload is not None else None
    rows = get_connector().execute_returning(
        """INSERT INTO jobs (input_path, phase_id, job_type, status, created_at, current_phase, next_phase_index, runner_state, enqueued_at, queue_payload, cancel_requested)
           VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 0) RETURNING id""",
        (input_path, phase_id, job_type, status, now, current_phase, next_phase_index, runner_state, now, payload_json)
    )
    job_id = rows[0]['id'] if rows else None

    record_pipeline_event(
        "state-change",
        f"Job #{job_id} created ({status})",
        workflow_run=job_id,
        stage_run=phase_code or job_type or "pipeline",
        step_run="job:create",
        category="job",
        metadata={"status": status, "input_path": input_path, "job_type": job_type, "phase_code": phase_code},
        source="db.create_job",
    )
    return job_id


def get_job(job_id):
    row = get_connector().query_one("SELECT * FROM jobs WHERE id = ?", (job_id,))
    return dict(row) if row else None


def set_job_execution_cursor(job_id, current_phase=None, next_phase_index=None, runner_state=None):
    """Persist pipeline execution cursor fields on a job row."""
    get_connector().execute(
        "UPDATE jobs SET current_phase = ?, next_phase_index = ?, runner_state = ? WHERE id = ?",
        (current_phase, next_phase_index, runner_state, job_id),
    )

    record_pipeline_event(
        "state-change",
        f"Job #{job_id} cursor updated",
        workflow_run=job_id,
        stage_run=current_phase or "pipeline",
        step_run="job:cursor",
        category="phase-transition",
        metadata={"current_phase": current_phase, "next_phase_index": next_phase_index, "runner_state": runner_state},
        source="db.set_job_execution_cursor",
    )


def update_job_status(job_id, status, log=None, current_phase=None, next_phase_index=None, runner_state=None):
    def _tx(tx):
        row = tx.query_one("SELECT status, current_phase, next_phase_index, runner_state FROM jobs WHERE id = ?", (job_id,))
        if not row:
            raise ValueError(f"Job not found: {job_id}")

        old_status = (row["status"] or "pending").strip().lower()
        new_status = (status or "").strip().lower()

        allowed_next = JOB_ALLOWED_TRANSITIONS.get(old_status)
        if allowed_next is not None and old_status != new_status and new_status not in allowed_next:
            raise ValueError(f"Invalid job status transition: {old_status} -> {new_status} (job_id={job_id})")

        final_log = log
        # Keep existing cursor values unless caller explicitly overrides
        final_phase = current_phase if current_phase is not None else row["current_phase"]
        final_next_idx = next_phase_index if next_phase_index is not None else row["next_phase_index"]
        final_runner_state = runner_state if runner_state is not None else row["runner_state"]

        now = datetime.datetime.now()
        if new_status == "running":
            tx.execute(
                "UPDATE jobs SET status = ?, started_at = COALESCE(started_at, ?), log = ?, current_phase = ?, next_phase_index = ?, runner_state = ? WHERE id = ?",
                (new_status, now, final_log, final_phase, final_next_idx, final_runner_state, job_id),
            )
        elif new_status in JOB_TERMINAL_STATES:
            tx.execute(
                "UPDATE jobs SET status = ?, finished_at = ?, completed_at = ?, log = ?, current_phase = ?, next_phase_index = ?, runner_state = ? WHERE id = ?",
                (new_status, now, now, final_log, final_phase, final_next_idx, final_runner_state, job_id),
            )
        else:
            tx.execute(
                "UPDATE jobs SET status = ?, log = ?, current_phase = ?, next_phase_index = ?, runner_state = ? WHERE id = ?",
                (new_status, final_log, final_phase, final_next_idx, final_runner_state, job_id),
            )

        # Keep job_phases state in sync for phase-bound jobs
        try:
            count_row = tx.query_one("SELECT COUNT(*) AS cnt FROM job_phases WHERE job_id = ?", (job_id,))
            n_phases = int(count_row["cnt"]) if count_row else 0

            job_row = tx.query_one("SELECT phase_id, job_type FROM jobs WHERE id = ?", (job_id,))
            phase_code = None
            if job_row:
                if job_row["phase_id"]:
                    p_row = tx.query_one("SELECT code FROM pipeline_phases WHERE id = ?", (job_row["phase_id"],))
                    if p_row:
                        phase_code = p_row["code"]
                if not phase_code and job_row["job_type"] != "pipeline":
                    phase_code = job_row["job_type"]
                if not phase_code and job_row["job_type"] == "pipeline":
                    phase_code = get_next_running_job_phase(job_id, tx=tx)

            phase_state_map = {
                "queued": "queued",
                "running": "running",
                "paused": "paused",
                "cancel_requested": "cancel_requested",
                "restarting": "restarting",
                "completed": "completed",
                "failed": "failed",
                "canceled": "canceled",
                "cancelled": "canceled",
                "interrupted": "interrupted",
            }
            phase_state = phase_state_map.get(new_status, "running")

            if n_phases > 1:
                multi = _resolve_multi_phase_job_phases_sync_code(job_id, new_status, tx=tx)
                if multi == "__bulk_completed__":
                    tx.execute(
                        "UPDATE job_phases SET state = 'completed', completed_at = ?, error_message = NULL WHERE job_id = ?",
                        (now, job_id),
                    )
                elif multi:
                    set_job_phase_state(
                        job_id,
                        multi,
                        phase_state,
                        error_message=log if new_status in {"failed", "interrupted"} else None,
                        tx=tx,
                    )
            elif phase_code:
                set_job_phase_state(
                    job_id,
                    phase_code,
                    phase_state,
                    error_message=log if new_status in {"failed", "interrupted"} else None,
                    tx=tx,
                )
        except Exception as e:
            logger.debug("update_job_status: failed to sync job_phases for job %s: %s", job_id, e)

        return old_status, new_status, final_phase, final_next_idx, final_runner_state

    old_status, new_status, final_phase, final_next_idx, final_runner_state = get_connector().run_transaction(_tx)


    event_type = "state-change"
    severity = "info"
    if new_status == "failed":
        event_type = "error"
        severity = "error"
    elif new_status in ("completed", "canceled"):
        event_type = "recovery"
        severity = "warning" if new_status == "canceled" else "info"

    record_pipeline_event(
        event_type,
        f"Job #{job_id} status: {old_status} → {new_status}",
        workflow_run=job_id,
        stage_run=final_phase or "pipeline",
        step_run="job:status",
        category="job",
        severity=severity,
        metadata={
            "old_status": old_status,
            "status": new_status,
            "current_phase": final_phase,
            "next_phase_index": final_next_idx,
            "runner_state": final_runner_state,
        },
        critical=new_status in ("failed", "interrupted"),
        source="db.update_job_status",
    )

    # Broadcast job status update
    try:
        from modules.events import event_manager
        event_manager.broadcast_threadsafe(f"job_{new_status}", {
            "job_id": job_id,
            "status": new_status,
            "current_phase": final_phase,
            "next_phase_index": final_next_idx,
            "runner_state": final_runner_state,
        })
    except Exception:
        pass


def enqueue_job(input_path, phase_code, job_type=None, queue_payload=None):
    """Create a queued job with a stable internal sort key and dense display position."""
    phase_id = get_phase_id(phase_code) if phase_code else None
    if job_type is None:
        job_type = phase_code

    now = datetime.datetime.now()
    payload_json = json.dumps(queue_payload) if queue_payload is not None else None
    priority = int((queue_payload or {}).get("priority", 100)) if isinstance(queue_payload, dict) else 100
    priority = max(1, min(priority, 999))
    target_scope = None
    if isinstance(queue_payload, dict):
        target_scope = queue_payload.get("target_scope") or queue_payload.get("scope")
    if not target_scope:
        target_scope = input_path

    def _tx(tx):
        rows = tx.execute_returning(
            """
            INSERT INTO jobs (
                input_path, phase_id, job_type, status, queue_position,
                created_at, enqueued_at, queue_payload, cancel_requested,
                priority, target_scope, retry_count
            ) VALUES (?, ?, ?, 'queued', NULL, ?, ?, ?, 0, ?, ?, 0) RETURNING id
            """,
            (input_path, phase_id, job_type, now, now, payload_json, priority, target_scope),
        )
        job_id = rows[0]['id'] if rows else None
        if not job_id:
            raise RuntimeError("Failed to insert job row")

        # Persist a stable queue ordering key using the DB identity.
        tx.execute("UPDATE jobs SET queue_position = ? WHERE id = ?", (job_id, job_id))

        # Return dense user-facing queue position (1..N), not the internal sort key.
        count_rows = tx.query(
            """
            SELECT COUNT(*) AS cnt FROM jobs
            WHERE status = 'queued' AND COALESCE(queue_position, id) <= ?
            """,
            (job_id,),
        )
        display_position = int((count_rows[0].get('cnt') or 0) if count_rows else 0)
        return job_id, display_position

    try:
        return get_connector().run_transaction(_tx)
    except RuntimeError:
        return None, 0


def requeue_job(job_id):
    """Reset an existing job row to queued status (in-place resume).

    Resets started_at, finished_at, completed_at and bumps enqueued_at.
    Updates queue_position so it sorts after any already-queued jobs.
    Returns (job_id, display_position).
    """
    now = datetime.datetime.now()

    def _tx(tx):
        row = tx.query_one("SELECT status FROM jobs WHERE id = ?", (job_id,))
        if not row:
            raise ValueError(f"Job {job_id} not found")
        old_status = (row["status"] or "").strip().lower()
        allowed = JOB_ALLOWED_TRANSITIONS.get(old_status, set())
        if "queued" not in allowed:
            raise ValueError(f"Cannot requeue job from status '{old_status}'")

        tx.execute(
            """
            UPDATE jobs
            SET status = 'queued',
                started_at = NULL,
                finished_at = NULL,
                completed_at = NULL,
                enqueued_at = ?,
                queue_position = ?,
                runner_state = NULL,
                current_phase = NULL,
                cancel_requested = 0
            WHERE id = ?
            """,
            (now, job_id, job_id),
        )

        pos_row = tx.query_one(
            "SELECT COUNT(*) FROM jobs WHERE status = 'queued' AND COALESCE(queue_position, id) <= ?",
            (job_id,),
        )
        display_position = int((pos_row.get("count") or pos_row.get("COUNT(*)") or 0) if pos_row else 0)
        return job_id, display_position

    try:
        return get_connector().run_transaction(_tx)
    except RuntimeError:
        return None, 0


def update_job_payload(job_id, queue_payload):
    """Update the queue_payload column on an existing job."""
    get_connector().execute(
        "UPDATE jobs SET queue_payload = ? WHERE id = ?",
        (queue_payload, job_id),
    )


def get_job_by_id(job_id):
    """Return a single job by id as a dictionary."""
    return get_job(job_id)


def dequeue_next_job():
    """Atomically take the oldest queued job and mark it running."""
    def _tx(tx):
        row = tx.query_one(
            """
            SELECT id FROM jobs
            WHERE status = 'queued' AND COALESCE(cancel_requested, 0) = 0
            ORDER BY COALESCE(priority, 100) DESC, COALESCE(queue_position, id) ASC, enqueued_at ASC, id ASC
            FETCH FIRST 1 ROWS ONLY
            """
        )
        if not row:
            return None
        job_id = int(row["id"])
        now = datetime.datetime.now()
        rowcount = tx.execute(
            """
            UPDATE jobs
            SET status = 'running', started_at = ?, queue_position = NULL
            WHERE id = ? AND status = 'queued' AND COALESCE(cancel_requested, 0) = 0
            """,
            (now, job_id),
        )
        if rowcount == 0:
            raise RuntimeError("dequeue race condition")
        return job_id

    try:
        job_id = get_connector().run_transaction(_tx)
    except RuntimeError:
        return None
    if job_id is None:
        return None
    return get_job_by_id(job_id)



def get_queued_jobs(limit=200, include_related=False):
    try:
        limit = int(limit)
    except (ValueError, TypeError):
        limit = 200
    if limit <= 0:
        return []
    limit = min(limit, 1000)

    rows = [dict(r) for r in get_connector().query(
        """
        SELECT
            j.*,
            p.name AS phase_name,
            ph.selected_phases,
            ph.dependency_blockers
        FROM jobs j
        LEFT JOIN pipeline_phases p ON p.id = j.phase_id
        LEFT JOIN (
            SELECT
                jp.job_id,
                LIST(jp.phase_code, ', ') AS selected_phases,
                LIST(CASE WHEN jp.state IN ('blocked', 'waiting', 'pending_dependency') THEN jp.phase_code ELSE NULL END, ', ') AS dependency_blockers
            FROM job_phases jp
            GROUP BY jp.job_id
        ) ph ON ph.job_id = j.id
        WHERE j.status IN ('queued', 'paused', 'failed')
          AND (? = 1 OR j.status = 'queued')
        ORDER BY
            CASE j.status WHEN 'queued' THEN 0 WHEN 'paused' THEN 1 ELSE 2 END,
            COALESCE(j.priority, 100) DESC,
            COALESCE(j.queue_position, j.id) ASC,
            j.enqueued_at ASC,
            j.id ASC
        FETCH FIRST ? ROWS ONLY
        """,
        (1 if include_related else 0, limit),
    )]

    avg_seconds = 120
    try:
        avg_row = get_connector().query_one(
            """
            SELECT AVG(DATEDIFF(SECOND FROM started_at TO completed_at)) AS avg_sec
            FROM jobs
            WHERE status = 'completed' AND started_at IS NOT NULL AND completed_at IS NOT NULL
            """
        )
        if avg_row and avg_row.get("avg_sec"):
            avg_seconds = max(15, int(avg_row["avg_sec"]))
    except Exception:
        pass

    queue_idx = 0
    now = datetime.datetime.now()
    for row in rows:
        if row.get("status") in ("queued", "paused"):
            queue_idx += 1
            row["queue_position"] = queue_idx
            eta = now + datetime.timedelta(seconds=(queue_idx - 1) * avg_seconds)
            row["estimated_start"] = eta.isoformat(sep=" ", timespec="seconds")
        else:
            row["queue_position"] = "-"
            row["estimated_start"] = "-"
        row["target_scope"] = row.get("target_scope") or row.get("input_path") or "-"
        row["selected_phases"] = row.get("selected_phases") or row.get("phase_name") or row.get("job_type") or "-"
        row["dependency_blockers"] = row.get("dependency_blockers") or "None"
        row["retry_count"] = int(row.get("retry_count") or 0)
        row["priority"] = int(row.get("priority") or 100)
    return rows




def bump_job_priority(job_id, delta=10):
    """Increase/decrease job priority for queued/paused jobs."""
    try:
        d = int(delta)
    except Exception:
        d = 10

    def _tx(tx):
        rowcount = tx.execute(
            """
            UPDATE jobs
            SET priority = CASE
                WHEN COALESCE(priority, 100) + ? < 1 THEN 1
                WHEN COALESCE(priority, 100) + ? > 999 THEN 999
                ELSE COALESCE(priority, 100) + ?
            END
            WHERE id = ? AND status IN ('queued', 'paused')
            """,
            (d, d, d, job_id),
        )
        if rowcount > 0:
            row = tx.query_one("SELECT priority FROM jobs WHERE id = ?", (job_id,))
            new_priority = int(row["priority"]) if row and row["priority"] is not None else 100
        else:
            new_priority = None
        return {"success": rowcount > 0, "priority": new_priority}

    return get_connector().run_transaction(_tx)


def set_job_priority(job_id, priority):
    """Update job priority for queued/paused jobs."""
    try:
        p = max(1, min(int(priority), 999))
    except Exception:
        p = 100
    rowcount = get_connector().execute(
        "UPDATE jobs SET priority = ? WHERE id = ? AND status IN ('queued', 'paused')",
        (p, job_id),
    )
    return {"success": rowcount > 0, "priority": p}


def pause_queue_job(job_id):
    """Pause a queued job so it is temporarily skipped by dequeue."""
    now = datetime.datetime.now()
    rowcount = get_connector().execute(
        "UPDATE jobs SET status = 'paused', paused_at = ? WHERE id = ? AND status = 'queued'",
        (now, job_id),
    )
    return {"success": rowcount > 0}


def restart_failed_job(job_id):
    """Move failed job back to queued and increment retry_count."""
    now = datetime.datetime.now()
    rowcount = get_connector().execute(
        """
        UPDATE jobs
        SET status = 'queued', cancel_requested = 0, enqueued_at = ?, queue_position = id,
            retry_count = COALESCE(retry_count, 0) + 1, paused_at = NULL,
            started_at = NULL, finished_at = NULL, completed_at = NULL
        WHERE id = ? AND status = 'failed'
        """,
        (now, job_id),
    )
    return {"success": rowcount > 0}


def request_cancel_job(job_id):
    now = datetime.datetime.now()

    def _tx(tx):
        row = tx.query_one("SELECT status FROM jobs WHERE id = ?", (job_id,))
        if not row:
            return {"success": False, "reason": "not_found"}
        status = (row["status"] or "").strip().lower()
        if status in ("completed", "failed", "cancelled"):
            return {"success": False, "reason": "already_finished", "status": status}
        if status == "running":
            return {"success": False, "reason": "running_not_supported", "status": status}
        if status not in ("queued", "paused"):
            return {"success": False, "reason": "not_cancellable_state", "status": status}
        rowcount = tx.execute(
            """
            UPDATE jobs
            SET status = 'cancelled', cancel_requested = 1, queue_position = NULL,
                finished_at = ?, completed_at = ?
            WHERE id = ? AND status IN ('queued', 'paused')
            """,
            (now, now, job_id),
        )
        if rowcount == 0:
            latest = tx.query_one("SELECT status FROM jobs WHERE id = ?", (job_id,))
            latest_status = (latest["status"] or "").strip().lower() if latest else "not_found"
            if latest_status == "running":
                return {"success": False, "reason": "running_not_supported", "status": latest_status}
            if latest_status in ("completed", "failed", "cancelled"):
                return {"success": False, "reason": "already_finished", "status": latest_status}
            if latest_status == "not_found":
                return {"success": False, "reason": "not_found"}
            return {"success": False, "reason": "cancel_failed", "status": latest_status}
        return {"success": True, "reason": "cancelled", "status": status}

    return get_connector().run_transaction(_tx)


def create_job_phases(job_id, phase_codes, first_phase_state=None):
    """Persist ordered phase plan for a job.

    Args:
        first_phase_state: If ``'queued'``, first phase is queued (job still in queue). If ``None``,
            first phase is ``running`` with ``started_at`` set (immediate pipeline start).
    """
    if not phase_codes:
        return []

    now = datetime.datetime.now()
    rows = []
    for idx, phase_code in enumerate(phase_codes):
        if idx > 0:
            state = "pending"
            started_at = None
        elif first_phase_state == "queued":
            state = "queued"
            started_at = None
        else:
            state = "running"
            started_at = now
        rows.append((job_id, idx, phase_code, state, started_at))

    def _tx(tx):
        tx.execute("DELETE FROM job_phases WHERE job_id = ?", (job_id,))
        for row in rows:
            tx.execute(
                "INSERT INTO job_phases (job_id, phase_order, phase_code, state, started_at) VALUES (?, ?, ?, ?, ?)",
                row,
            )
    get_connector().run_transaction(_tx)

    return [
        {"phase_order": r[1], "phase_code": r[2], "state": r[3], "started_at": r[4], "completed_at": None, "error_message": None}
        for r in rows
    ]


def resume_job_phases(job_id):
    """Reset incomplete phases for resume. Completed/skipped stay; others → pending.

    The first incomplete phase is set to 'queued' so the dispatcher picks it up.
    Returns the updated phase list.
    """
    rows = get_connector().query(
        "SELECT phase_order, phase_code, state FROM job_phases WHERE job_id = ? ORDER BY phase_order",
        (job_id,),
    )
    if not rows:
        return []

    keep_states = {"completed", "skipped"}
    first_incomplete_set = False
    updates = []
    for r in rows:
        state = (r["state"] or "").strip().lower()
        if state in keep_states:
            continue
        new_state = "queued" if not first_incomplete_set else "pending"
        first_incomplete_set = True
        updates.append((new_state, job_id, r["phase_code"]))

    if updates:
        def _tx(tx):
            for params in updates:
                tx.execute(
                    "UPDATE job_phases SET state = ?, started_at = NULL, completed_at = NULL, error_message = NULL "
                    "WHERE job_id = ? AND phase_code = ?",
                    params,
                )
        get_connector().run_transaction(_tx)

    return [dict(r) for r in get_connector().query(
        "SELECT phase_order, phase_code, state, started_at, completed_at, error_message "
        "FROM job_phases WHERE job_id = ? ORDER BY phase_order",
        (job_id,),
    )]


def set_job_phase_state(job_id, phase_code, state, error_message=None, tx=None):
    """Update state metadata for one phase of a job and auto-advance next pending phase."""
    allowed = {
        "pending": {"queued", "running", "skipped", "canceled", "failed"},
        "queued": {"running", "paused", "cancel_requested", "canceled"},
        "running": {
            "paused",
            "completed",
            "failed",
            "interrupted",
            "cancel_requested",
            "restarting",
            "canceled",
        },
        "paused": {"running", "restarting", "cancel_requested", "canceled"},
        "cancel_requested": {"canceled", "failed"},
        "restarting": {"queued", "running", "failed"},
        "completed": set(),
        "failed": {"skipped", "pending"},
        "interrupted": {"running", "failed", "skipped", "pending", "queued"},
        "skipped": set(),
        "canceled": set(),
    }
    now = datetime.datetime.now()

    def _tx(tx):
        row = tx.query_one(
            "SELECT id, state FROM job_phases WHERE job_id = ? AND phase_code = ?",
            (job_id, phase_code),
        )
        if not row:
            return None

        phase_id = row["id"]
        old_state = str(row["state"] or "pending").strip().lower()
        new_state = str(state or "").strip().lower()
        if old_state != new_state and new_state not in allowed.get(old_state, set()):
            if not (old_state == "failed" and new_state in ("skipped", "pending")):
                msg = f"Invalid job phase transition: {old_state} -> {new_state} (job_id={job_id}, phase={phase_code})"
                logger.warning(msg)
                raise ValueError(msg)

        fields = ["state = ?"]
        params = [state]
        if state == "running":
            fields.append("started_at = COALESCE(started_at, ?)")
            params.append(now)
            fields.append("error_message = NULL")
        if state in {"completed", "failed", "skipped", "interrupted"}:
            fields.append("completed_at = ?")
            params.append(now)
        if error_message is not None:
            fields.append("error_message = ?")
            params.append(error_message)

        params.append(phase_id)
        tx.execute(f"UPDATE job_phases SET {', '.join(fields)} WHERE id = ?", params)

        if state in {"completed", "skipped"}:
            next_row = tx.query_one(
                "SELECT id FROM job_phases WHERE job_id = ? AND phase_order > "
                "(SELECT phase_order FROM job_phases WHERE id = ?) AND state = 'pending' "
                "ORDER BY phase_order FETCH FIRST 1 ROWS ONLY",
                (job_id, phase_id),
            )
            if next_row:
                tx.execute(
                    "UPDATE job_phases SET state = 'running', started_at = COALESCE(started_at, ?), error_message = NULL WHERE id = ?",
                    (now, next_row["id"]),
                )
        return True

    if tx:
        return _tx(tx)
    result = get_connector().run_transaction(_tx)
    if result is None:
        return None
    return get_job_phases(job_id, tx=tx)


def get_job_phases(job_id, tx=None):
    """Get ordered phase plan/status rows for a job."""
    conn = tx if tx else get_connector()
    return [dict(r) for r in conn.query(
        "SELECT phase_order, phase_code, state, started_at, completed_at, error_message "
        "FROM job_phases WHERE job_id = ? ORDER BY phase_order",
        (job_id,),
    )]


def _resolve_multi_phase_job_phases_sync_code(job_id, new_status, tx=None):
    """For jobs with multiple job_phases rows, pick which phase row mirrors ``jobs.status``.

    Single-phase jobs return None so callers keep using ``job_type`` as the phase code.

    Returns:
        None — use legacy single-phase sync (job_type / pipeline_phases).
        "__bulk_completed__" — mark every phase row completed (job finished successfully).
        str — a ``phase_code`` to pass to ``set_job_phase_state``.
    """
    phases = get_job_phases(job_id, tx=tx)
    if not phases or len(phases) <= 1:
        return None
    st = (new_status or "").strip().lower()
    if st == "running":
        for p in phases:
            row_state = (p.get("state") or "").strip().lower()
            if row_state in ("queued", "pending"):
                return p.get("phase_code")
        for p in phases:
            if (p.get("state") or "").strip().lower() == "running":
                return p.get("phase_code")
        return phases[0].get("phase_code")
    if st == "completed":
        # Safety: only bulk-complete if every phase has actually been started.
        # If any phase was never started (started_at is None and state is not
        # completed/skipped), return it so set_job_phase_state handles one
        # phase at a time with proper auto-advance.
        unstarted = [
            p for p in phases
            if p.get("started_at") is None
            and (p.get("state") or "").strip().lower() not in ("completed", "skipped", "canceled")
        ]
        if unstarted:
            # Find the currently running phase to mark completed
            for p in phases:
                if (p.get("state") or "").strip().lower() == "running":
                    return p.get("phase_code")
            # Fallback: last phase that isn't completed
            for p in reversed(phases):
                if (p.get("state") or "").strip().lower() not in ("completed", "skipped"):
                    return p.get("phase_code")
        return "__bulk_completed__"
    if st in ("failed", "interrupted", "canceled", "cancelled"):
        for p in phases:
            if (p.get("state") or "").strip().lower() == "running":
                return p.get("phase_code")
        return phases[-1].get("phase_code")
    # queued / paused / restarting — do not map whole-job status onto one stage row
    return None


def get_job_steps(job_id, phase_code):
    """Return step-level telemetry rows for a job+phase from job_steps table."""
    try:
        rows = get_connector().query(
            "SELECT id, step_code, step_name, status, started_at, completed_at, "
            "items_total, items_done, throughput_rps, error_message "
            "FROM job_steps WHERE job_id = ? AND phase_code = ? ORDER BY id",
            (job_id, phase_code),
        )
        return [
            {
                "id": r["id"],
                "step_code": r["step_code"],
                "step_name": r["step_name"],
                "status": r["status"],
                "started_at": str(r["started_at"]) if r["started_at"] else None,
                "completed_at": str(r["completed_at"]) if r["completed_at"] else None,
                "items_total": r["items_total"] or 0,
                "items_done": r["items_done"] or 0,
                "throughput_rps": r["throughput_rps"],
                "error_message": r["error_message"],
            }
            for r in rows
        ]
    except Exception:
        return []


def upsert_job_step(job_id, phase_code, step_code, step_name, status="pending",
                    items_total=0, items_done=0, throughput_rps=None, error_message=None):
    """Insert or update a step telemetry row in job_steps."""
    now = datetime.datetime.now()
    def _tx(tx):
        row = tx.query_one(
            "SELECT id FROM job_steps WHERE job_id = ? AND phase_code = ? AND step_code = ?",
            (job_id, phase_code, step_code),
        )
        if row:
            tx.execute(
                "UPDATE job_steps SET status = ?, items_total = ?, items_done = ?, "
                "throughput_rps = ?, error_message = ?, "
                "started_at = CASE WHEN status = 'running' THEN COALESCE(started_at, ?) ELSE started_at END, "
                "completed_at = CASE WHEN ? IN ('completed','failed','skipped') THEN ? ELSE completed_at END "
                "WHERE id = ?",
                (status, items_total, items_done, throughput_rps, error_message,
                 now, status, now, row["id"]),
            )
        else:
            tx.execute(
                "INSERT INTO job_steps (job_id, phase_code, step_code, step_name, status, "
                "items_total, items_done, throughput_rps, error_message, started_at) "
                "VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)",
                (job_id, phase_code, step_code, step_name, status,
                 items_total, items_done, throughput_rps, error_message,
                 now if status == "running" else None),
            )
    try:
        get_connector().run_transaction(_tx)
    except Exception:
        pass


def get_job_stage_images(job_id, phase_code, offset=0, limit=50):
    """Return work items (images + their phase status) for a specific job+stage."""
    try:
        phase_row = get_connector().query_one(
            "SELECT id FROM pipeline_phases WHERE code = ?", (phase_code,))
        if not phase_row:
            return {"items": [], "total": 0}
        phase_id = phase_row["id"]

        count_row = get_connector().query_one(
            "SELECT COUNT(*) AS cnt FROM image_phase_status ips "
            "JOIN images i ON ips.image_id = i.id "
            "WHERE ips.job_id = ? AND ips.phase_id = ?",
            (job_id, phase_id))
        total = (count_row["cnt"] if count_row else 0) or 0

        rows = get_connector().query(
            "SELECT i.id, i.file_path, i.file_name, ips.status, ips.started_at, ips.finished_at, ips.last_error "
            "FROM image_phase_status ips "
            "JOIN images i ON ips.image_id = i.id "
            "WHERE ips.job_id = ? AND ips.phase_id = ? "
            "ORDER BY i.id "
            "OFFSET ? ROWS FETCH NEXT ? ROWS ONLY",
            (job_id, phase_id, offset, limit))
        items = []
        for r in rows:
            started = r.get("started_at")
            finished = r.get("finished_at")
            duration_ms = None
            if started and finished:
                duration_ms = int((finished - started).total_seconds() * 1000)
            items.append({
                "image_id": r["id"],
                "image_path": r.get("file_path") or "",
                "filename": r.get("file_name") or (r.get("file_path") or "").rsplit("/", 1)[-1].rsplit("\\", 1)[-1],
                "status": r.get("status") or "pending",
                "duration_ms": duration_ms,
                "error": r.get("last_error"),
            })
        return {"items": items, "total": total}
    except Exception:
        return {"items": [], "total": 0}


def get_next_running_job_phase(job_id, tx=None):
    """Return current running phase for a job, if any."""
    conn = tx if tx else get_connector()
    row = conn.query_one(
        "SELECT phase_code FROM job_phases WHERE job_id = ? AND state = 'running' ORDER BY phase_order FETCH FIRST 1 ROWS ONLY",
        (job_id,),
    )
    return row["phase_code"] if row else None


def recover_running_jobs(mark_as="interrupted"):
    """Mark stale running jobs (and their in-flight job_phases) as interrupted."""
    rows = get_connector().query("SELECT id FROM jobs WHERE status = 'running'")
    recovered = [r["id"] for r in rows]
    if recovered:
        now = datetime.datetime.now()
        def _tx(tx):
            tx.execute(
                "UPDATE jobs SET status = ?, completed_at = ?, runner_state = ? WHERE status = 'running'",
                (mark_as, now, mark_as),
            )
            placeholders = ",".join("?" * len(recovered))
            tx.execute(
                f"UPDATE job_phases SET state = ?, completed_at = ? "
                f"WHERE job_id IN ({placeholders}) AND state = 'running'",
                [mark_as, now] + recovered,
            )
        get_connector().run_transaction(_tx)
    return recovered


def get_interrupted_jobs(job_type=None, limit=100):
    try:
        limit = int(limit)
    except (TypeError, ValueError):
        limit = 100

    if job_type:
        rows = get_connector().query(
            "SELECT * FROM jobs WHERE status = 'interrupted' AND job_type = ? ORDER BY created_at DESC FETCH FIRST ? ROWS ONLY",
            (job_type, limit),
        )
    else:
        rows = get_connector().query(
            "SELECT * FROM jobs WHERE status = 'interrupted' ORDER BY created_at DESC FETCH FIRST ? ROWS ONLY",
            (limit,),
        )
    return [dict(r) for r in rows]


def get_jobs(limit=50):
    try: limit = int(limit)
    except (ValueError, TypeError): limit = 50
    if limit < 0: limit = 50
    limit = min(limit, 1000)
    return [dict(r) for r in get_connector().query(
        "SELECT * FROM jobs ORDER BY created_at DESC FETCH FIRST ? ROWS ONLY", (limit,)
    )]


def get_all_images(sort_by="score", order="desc", limit=100):
    # Ensure limit is int
    try: limit = int(limit)
    except (ValueError, TypeError): limit = 100
    sort_by, order = _validate_sort(sort_by, order)
    if limit > 0:
        return list(get_connector().query(
            f"SELECT * FROM images ORDER BY {sort_by} {order} FETCH FIRST ? ROWS ONLY", (limit,)
        ))
    return list(get_connector().query(f"SELECT * FROM images ORDER BY {sort_by} {order}"))


def get_nef_paths_for_research(limit=500):
    """
    Fetch NEF file paths for research/assessment scripts.
    Returns a random sample of NEF images with id, file_path, score_general.
    Used by scripts/research_models.py for test set selection.
    """
    try:
        limit = int(limit)
    except (TypeError, ValueError):
        limit = 500
    limit = max(1, min(limit, 10000))
    fb_sql = (
        "SELECT id, file_path, score_general, score_technical, score_aesthetic "
        "FROM images WHERE LOWER(file_type) = 'nef' "
        "ORDER BY RAND() FETCH FIRST ? ROWS ONLY"
    )
    return list(get_connector().query(fb_sql, (limit,)))

def sync_folder_to_db(folder_path, job_id=None):
    """
    Scans a folder for .json files (generated by scoring) and upserts them to DB.
    """
    from modules import utils

    count = 0
    touched_folder_ids = set()
    folder = Path(folder_path)
    if not folder.exists():
        return 0

    # Collect rows first; execute in a single transaction
    upserts = []
    uuid_updates = []

    for json_file in folder.glob("*.json"):
        try:
            with open(json_file, 'r', encoding='utf-8') as f:
                data = json.load(f)

            if "summary" not in data and "average_normalized_score" not in data:
                continue

            # Extract score (General)
            score = 0
            if "summary" in data and "weighted_scores" in data["summary"]:
                score = data["summary"]["weighted_scores"].get("general", 0)
            elif "summary" in data:
                score = data["summary"].get("average_normalized_score", 0)

            image_path = data.get("image_path", "")
            if not image_path:
                stem = json_file.stem
                for ext in ['.jpg', '.nef', '.png']:
                    probe = folder / (stem + ext)
                    if probe.exists():
                        image_path = str(probe)
                        break

            if not image_path:
                image_path = str(json_file.with_suffix(''))

            file_name = Path(image_path).name

            upserts.append((job_id, str(image_path), file_name, score, json.dumps(data), utils.get_image_creation_time(str(image_path))))

            meta_dict = data.get("metadata") if isinstance(data.get("metadata"), dict) else None
            image_uuid = generate_image_uuid(meta_dict)
            uuid_updates.append((image_uuid, str(image_path)))

            folder_id = get_or_create_folder(os.path.dirname(str(image_path))) if image_path else None
            if folder_id:
                touched_folder_ids.add(folder_id)

            count += 1
        except Exception as e:
            logging.error(f"Failed to sync {json_file}: {e}")

    if upserts:
        def _batch(tx):
            for row in upserts:
                tx.execute(
                    '''UPDATE OR INSERT INTO images
                       (job_id, file_path, file_name, score, scores_json, created_at)
                       VALUES (?, ?, ?, ?, ?, ?)
                       MATCHING (file_path)''',
                    row,
                )
            for image_uuid, image_path in uuid_updates:
                tx.execute(
                    "UPDATE images SET image_uuid = ? WHERE file_path = ? AND (image_uuid IS NULL OR image_uuid = '')",
                    (image_uuid, image_path),
                )
        get_connector().run_transaction(_batch)

    if touched_folder_ids:
        for fid in touched_folder_ids:
            invalidate_folder_phase_aggregates(folder_id=fid)

    if count > 0:
        event_manager.broadcast_threadsafe("folder_scanned", {"folder_path": folder_path, "new_images": count})
    return count

def upsert_image(job_id, result, *, invalidate_agg=True, dirty_folder_ids=None):
    """
    Upsert a single image result from the streaming output.
    result is a generic dictionary (the JSON output from batch_process_images).

    Args:
        invalidate_agg: When False, skip per-row ``invalidate_folder_phase_aggregates``;
            caller should batch-invalidate (e.g. scoring ResultWorker + engine flush).
        dirty_folder_ids: If provided and ``invalidate_agg`` is False, folder_id values
            that would have been invalidated are added to this set for later flush.
    """
    from modules import utils

    def _record_phase_agg(fld_id, old_fld=None):
        if not fld_id and not old_fld:
            return
        if invalidate_agg:
            if fld_id:
                invalidate_folder_phase_aggregates(folder_id=fld_id)
            if old_fld and old_fld != fld_id:
                invalidate_folder_phase_aggregates(folder_id=old_fld)
        elif dirty_folder_ids is not None:
            if fld_id:
                dirty_folder_ids.add(fld_id)
            if old_fld and old_fld != fld_id:
                dirty_folder_ids.add(old_fld)

    # Extract fields
    image_path = result.get("image_path", "")
    file_name = result.get("image_name", Path(image_path).name)
    file_type = Path(image_path).suffix.lower().lstrip('.')
    
    # Handle score location variation (summary vs direct)
    # New Standard: "score" in root of engine result (which comes from weighted_scores['general'])
    # Or parsing directly from stored JSON if re-syncing
    
    score = 0
    normalized_score = 0
    
    if "score" in result:
        score = result["score"]
    elif "summary" in result and "weighted_scores" in result["summary"]:
         score = result["summary"]["weighted_scores"].get("general", 0)
    elif "summary" in result: # Legacy fallback
        score = result["summary"].get("average_normalized_score", 0)
    
    # normalized_score removal
    # normalized_score = score # General score is already 0-1
        
    # Individual Scores
    individual_scores = result.get("individual_scores", {})
    models_scores = result.get("models", {})
    
    def get_ind_score(name):
        # Try 'models' first (new format)
        if name in models_scores:
            m_data = models_scores[name]
            if isinstance(m_data, dict):
                 # Return None if status is not success
                 if m_data.get("status") != "success":
                     return None
                 return m_data.get("normalized_score", m_data.get("score"))
            return None
            
        # Try 'individual_scores' (legacy format)
        val = individual_scores.get(name)
        if isinstance(val, dict):
            if val.get("status") != "success" and "status" in val:
                return None
            return val.get("normalized_score", val.get("score"))
        if isinstance(val, (int, float)):
            return val
        return None

    score_spaq = get_ind_score("spaq")
    score_ava = get_ind_score("ava")
    score_koniq = get_ind_score("koniq")
    score_paq2piq = get_ind_score("paq2piq")
    score_liqe = get_ind_score("liqe")
        
    # Weighted Scores
    # Try to get from result (if passed from engine) or parse from summary
    score_technical = None
    score_aesthetic = None
    score_general = None
    
    if "score_technical" in result:
        score_technical = result["score_technical"]
        score_aesthetic = result["score_aesthetic"]
        score_general = result["score_general"]
    elif "summary" in result and "weighted_scores" in result["summary"]:
        ws = result["summary"]["weighted_scores"]
        score_technical = ws.get("technical")
        score_aesthetic = ws.get("aesthetic")
        score_general = ws.get("general")
    elif "full_results" in result: 
        # Engine passes full_results
        ws = result["full_results"].get("summary", {}).get("weighted_scores", {})
        score_technical = ws.get("technical")
        score_aesthetic = ws.get("aesthetic")
        score_general = ws.get("general")
        
    # Ensure main score matches general if not set
    if (score == 0 or score is None) and score_general is not None and score_general > 0:
        score = score_general


    
    thumbnail_path = result.get("thumbnail_path")
    from modules.thumbnails import thumb_path_to_win, thumb_path_to_wsl
    thumbnail_path_win = result.get("thumbnail_path_win") or thumb_path_to_win(thumbnail_path)

    # Extract Version
    model_version = "0.0.0"
    if "version" in result:
        model_version = result["version"]
    elif "full_results" in result:
        model_version = result["full_results"].get("version", "0.0.0")

    # Extract Metadata (Rating, Label)
    rating = 0
    label = ""
    
    # Try finding in nef_metadata
    nef_meta = None
    if "nef_metadata" in result: # Direct
         nef_meta = result["nef_metadata"]
    elif "full_results" in result and "summary" in result["full_results"]:
         nef_meta = result["full_results"]["summary"].get("nef_metadata")
    elif "summary" in result: # Legacy
         nef_meta = result["summary"].get("nef_metadata")
         
    if nef_meta:
        rating = nef_meta.get("rating", 0)
        label = nef_meta.get("label", "")
    


    # Keywords & Metadata (if present)
    keywords = result.get("keywords", [])
    if isinstance(keywords, list):
        keywords = ",".join(keywords)
        
    title = result.get("title", "")
    description = result.get("description", "")
        
    metadata = result.get("metadata", {})
    if isinstance(metadata, dict):
        metadata = json.dumps(metadata)


    image_hash = result.get("image_hash", None)

    # Resolve Folder ID
    folder_id = None
    if image_path:
        try:
             folder_id = get_or_create_folder(os.path.dirname(image_path))
        except Exception as e:
             logging.error(f"Error resolving folder for {image_path}: {e}")

    existing_image = None
    if image_path:
        try:
            existing_image = get_connector().query_one(
                "SELECT id, folder_id FROM images WHERE file_path = ?", (image_path,)
            )
        except Exception:
            existing_image = None

    # Prevent duplicates by image_uuid: if UUID exists in DB (different path), update that record instead of inserting
    meta_dict = result.get("metadata") if isinstance(result.get("metadata"), dict) else (
        json.loads(result.get("metadata")) if isinstance(result.get("metadata"), str) else None
    )
    image_uuid_val = generate_image_uuid(meta_dict)
    if image_uuid_val and image_uuid_val.strip():
        existing_id = find_image_id_by_uuid(image_uuid_val)
        if existing_id:
            # Same image (by UUID) at different path — update existing record, don't insert
            existing_path = None
            old_folder_id = None
            try:
                r = get_connector().query_one(
                    "SELECT file_path, folder_id FROM images WHERE id = ?", (existing_id,)
                )
                if r:
                    existing_path = r["file_path"]
                    old_folder_id = r["folder_id"]
            except Exception:
                pass
            if existing_path != image_path:
                logger.info("Duplicate by UUID %s: updating existing id=%s path %s -> %s",
                            image_uuid_val[:16], existing_id, existing_path, image_path)
                get_connector().execute(
                    '''UPDATE images SET
                       job_id=?, file_path=?, file_name=?, file_type=?,
                       score=?, score_spaq=?, score_ava=?, score_koniq=?, score_paq2piq=?, score_liqe=?,
                       score_technical=?, score_aesthetic=?, score_general=?, model_version=?,
                       rating=?, label=?, keywords=?, title=?, description=?, metadata=?, scores_json=?,
                       thumbnail_path=?, thumbnail_path_win=?, image_hash=?, folder_id=?
                       WHERE id=?''',
                    (job_id, image_path, file_name, file_type,
                     score, score_spaq, score_ava, score_koniq, score_paq2piq, score_liqe,
                     score_technical, score_aesthetic, score_general, model_version,
                     rating, label, keywords, title, description, metadata, json.dumps(result),
                     thumbnail_path, thumbnail_path_win, image_hash, folder_id, existing_id)
                )
                _sync_image_keywords(existing_id, keywords)
                register_image_path(existing_id, image_path)
                try:
                    resolve_windows_path(existing_id, image_path, verify=False)
                except Exception:
                    pass
                if image_path:
                    invalidate_folder_images_cache(os.path.dirname(image_path))
                _record_phase_agg(folder_id, old_folder_id)
                event_manager.broadcast_threadsafe("image_scored", {
                    "image_id": existing_id,
                    "file_path": image_path,
                    "score_general": score_general,
                    "score_technical": score_technical,
                    "score_aesthetic": score_aesthetic,
                    "rating": rating,
                    "label": label,
                    "image_hash": image_hash
                })
                return

    _upsert_params = (
        job_id, image_path, file_name, file_type,
        score,
        score_spaq, score_ava, score_koniq, score_paq2piq, score_liqe,
        score_technical, score_aesthetic, score_general, model_version,
        rating, label,
        keywords, title, description, metadata, json.dumps(result),
        thumbnail_path, thumbnail_path_win,
        image_hash, folder_id, utils.get_image_creation_time(image_path)
    )

    def _tx(tx):
        ret = tx.execute_returning(
            '''UPDATE OR INSERT INTO images
                  (job_id, file_path, file_name, file_type,
                   score,
                   score_spaq, score_ava, score_koniq, score_paq2piq, score_liqe,
                   score_technical, score_aesthetic, score_general, model_version,
                   rating, label,
                   keywords, title, description, metadata, scores_json,
                   thumbnail_path, thumbnail_path_win,
                   image_hash, folder_id, created_at)
                  VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                  MATCHING (file_path) RETURNING id''',
            _upsert_params,
        )
        img_id = ret[0]["id"] if ret else None
        if img_id:
            try:
                _meta = result.get("metadata") if isinstance(result.get("metadata"), dict) else (
                    json.loads(result.get("metadata")) if isinstance(result.get("metadata"), str) else None
                )
                new_uuid = generate_image_uuid(_meta)
                tx.execute(
                    "UPDATE images SET image_uuid = ? WHERE id = ? AND (image_uuid IS NULL OR image_uuid = '')",
                    (new_uuid, img_id),
                )
            except Exception as _uuid_err:
                logger.warning("Could not assign image_uuid for id %s: %s", img_id, _uuid_err)
        return img_id

    image_id = get_connector().run_transaction(_tx)
    old_folder_id = existing_image["folder_id"] if existing_image else None

    if image_id:
        _sync_image_keywords(image_id, keywords)

    if image_path:
        invalidate_folder_images_cache(os.path.dirname(image_path))

    _record_phase_agg(folder_id, old_folder_id)

    # Register path in file_paths
    if image_id:
        register_image_path(image_id, image_path)
        # Also resolve Windows path for native viewer
        resolve_windows_path(image_id, image_path, verify=False)

        event_manager.broadcast_threadsafe("image_scored", {
            "image_id": image_id,
            "file_path": image_path,
            "score_general": score_general,
            "score_technical": score_technical,
            "score_aesthetic": score_aesthetic,
            "rating": rating,
            "label": label,
            "image_hash": image_hash
        })



def get_image_details(file_path):
    row = get_connector().query_one(
        "SELECT * FROM images WHERE file_path = ?", (file_path,)
    )
    if not row:
        return {}
    data = dict(row)
    data['file_paths'] = get_all_paths(data['id'])
    data['resolved_path'] = get_resolved_path(data['id'], verified_only=False)
    return data


def upsert_image_exif(image_id: int, data: dict) -> bool:
    """
    Upsert EXIF metadata for an image into IMAGE_EXIF.
    data keys: make, model, lens_model, focal_length, focal_length_35mm,
    date_time_original, create_date, exposure_time, f_number, iso,
    exposure_compensation, image_width, image_height, orientation, flash,
    image_unique_id, shutter_count, sub_sec_time_original
    """
    if not image_id or not isinstance(data, dict):
        return False
    try:
        extracted_at = datetime.datetime.now()
        get_connector().execute(
            '''UPDATE OR INSERT INTO image_exif (
                image_id, make, model, lens_model, focal_length, focal_length_35mm,
                date_time_original, create_date, exposure_time, f_number, iso,
                exposure_compensation, image_width, image_height, orientation, flash,
                image_unique_id, shutter_count, sub_sec_time_original, extracted_at
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            MATCHING (image_id)''',
            (
                image_id,
                data.get('make'),
                data.get('model'),
                data.get('lens_model'),
                data.get('focal_length'),
                _safe_int(data.get('focal_length_35mm')),
                _parse_exif_timestamp(data.get('date_time_original')),
                _parse_exif_timestamp(data.get('create_date')),
                _str_or_none(data.get('exposure_time')),
                _str_or_none(data.get('f_number')),
                _safe_int(data.get('iso')),
                _str_or_none(data.get('exposure_compensation')),
                _safe_int(data.get('image_width')),
                _safe_int(data.get('image_height')),
                _safe_int(data.get('orientation')),
                _safe_int(data.get('flash')),
                _str_or_none(data.get('image_unique_id')),
                _safe_int(data.get('shutter_count')),
                _str_or_none(data.get('sub_sec_time_original')),
                extracted_at,
            ),
        )
        return True
    except Exception as e:
        logger.warning("upsert_image_exif failed for image_id %s: %s", image_id, e)
        return False


def upsert_image_xmp(image_id: int, data: dict) -> bool:
    """
    Upsert XMP sidecar metadata for an image into IMAGE_XMP.
    data keys: rating, label, pick_status, burst_uuid, stack_id, keywords,
    title, description, create_date, modify_date
    """
    if not image_id or not isinstance(data, dict):
        return False
    try:
        extracted_at = datetime.datetime.now()
        keywords_val = data.get('keywords')
        if isinstance(keywords_val, list):
            keywords_val = json.dumps(keywords_val) if keywords_val else None
        elif not isinstance(keywords_val, str):
            keywords_val = None
        get_connector().execute(
            '''UPDATE OR INSERT INTO image_xmp (
                image_id, rating, label, pick_status, burst_uuid, stack_id,
                keywords, title, description, create_date, modify_date, extracted_at
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            MATCHING (image_id)''',
            (
                image_id,
                _safe_int(data.get('rating')),
                _str_or_none(data.get('label')),
                _safe_int(data.get('pick_status')),
                _str_or_none(data.get('burst_uuid')),
                _str_or_none(data.get('stack_id')),
                keywords_val,
                _str_or_none(data.get('title')),
                _str_or_none(data.get('description')),
                _parse_exif_timestamp(data.get('create_date')),
                _parse_exif_timestamp(data.get('modify_date')),
                extracted_at,
            ),
        )
        return True
    except Exception as e:
        logger.warning("upsert_image_xmp failed for image_id %s: %s", image_id, e)
        return False


def get_image_exif(image_id: int) -> dict | None:
    """Get cached EXIF metadata for an image. Returns None if not found."""
    if not image_id:
        return None
    row = get_connector().query_one("SELECT * FROM image_exif WHERE image_id = ?", (image_id,))
    return dict(row) if row else None


def get_image_xmp(image_id: int) -> dict | None:
    """Get cached XMP metadata for an image. Returns None if not found."""
    if not image_id:
        return None
    row = get_connector().query_one("SELECT * FROM image_xmp WHERE image_id = ?", (image_id,))
    return dict(row) if row else None


def _safe_int(val):
    """Convert value to int, return None for invalid/empty."""
    if val is None:
        return None
    try:
        return int(float(val)) if val != '' else None
    except (ValueError, TypeError):
        return None


def _str_or_none(val, max_len=None):
    """Return string or None. Truncate if max_len given."""
    if val is None:
        return None
    s = str(val).strip()
    if not s:
        return None
    if max_len and len(s) > max_len:
        return s[:max_len]
    return s


def _parse_exif_timestamp(val):
    """Parse EXIF/XMP timestamp strings to datetime. Returns None on failure."""
    if val is None or (isinstance(val, str) and not val.strip()):
        return None
    if hasattr(val, 'year'):  # Already datetime
        return val
    s = str(val).strip()
    formats = [
        ("%Y:%m:%d %H:%M:%S", 19),
        ("%Y-%m-%dT%H:%M:%S", 19),
        ("%Y-%m-%d %H:%M:%S", 19),
        ("%Y:%m:%d", 10),
        ("%Y-%m-%d", 10),
    ]
    for fmt, min_len in formats:
        if len(s) >= min_len:
            try:
                return datetime.datetime.strptime(s[:len(fmt)], fmt)
            except ValueError:
                continue
    return None


def get_stack_ids_for_image_ids(image_ids):
    """
    Returns {image_id: stack_id} for images that have a stack_id.
    Use for batch lookup after clustering instead of N get_image_details calls.
    """
    if not image_ids:
        return {}
    placeholders = ",".join("?" * len(image_ids))
    rows = get_connector().query(
        f"SELECT id, stack_id FROM images WHERE id IN ({placeholders}) AND stack_id IS NOT NULL",
        tuple(image_ids))
    return {r["id"]: r["stack_id"] for r in rows}


def delete_image(file_path, delete_related: bool = True):
    """
    Remove an image record from the database.

    Note: This does NOT delete the image file on disk.
    It does clean up dependent rows (culling picks / resolved paths / file paths) and
    repairs stack metadata (best_image_id / empty stacks) when possible.

    Returns: (success: bool, message: str)
    """
    if not file_path:
        return False, "No file path provided"

    row = get_connector().query_one(
        "SELECT id, stack_id, thumbnail_path, thumbnail_path_win FROM images WHERE file_path = ?",
        (file_path,),
    )
    if not row:
        return False, "Image not found in DB"
    image_id = row["id"]
    stack_id = row["stack_id"]
    thumbnail_path = row["thumbnail_path"]
    thumbnail_path_win = row["thumbnail_path_win"]
    try:
        def _tx(tx):
            if delete_related:
                tx.execute("DELETE FROM culling_picks WHERE image_id = ?", (image_id,))
                tx.execute("DELETE FROM file_paths WHERE image_id = ?", (image_id,))
            tx.execute("DELETE FROM images WHERE id = ?", (image_id,))
            if stack_id is not None:
                cnt_row = tx.query_one("SELECT COUNT(*) AS cnt FROM images WHERE stack_id = ?", (stack_id,))
                remaining = cnt_row["cnt"] if cnt_row else 0
                if remaining == 0:
                    tx.execute("DELETE FROM stacks WHERE id = ?", (stack_id,))
                else:
                    best_row = tx.query_one("SELECT best_image_id FROM stacks WHERE id = ?", (stack_id,))
                    if best_row and best_row["best_image_id"] == image_id:
                        tx.execute(
                            """UPDATE stacks SET best_image_id = (
                                SELECT id FROM images WHERE stack_id = ?
                                ORDER BY score_general DESC NULLS LAST FETCH FIRST 1 ROWS ONLY
                            ) WHERE id = ?""",
                            (stack_id, stack_id),
                        )
        get_connector().run_transaction(_tx)
    except Exception as exc:
        return False, str(exc)
    invalidate_folder_images_cache(os.path.dirname(file_path))
    import platform as _plat
    _local_thumb = thumbnail_path_win if _plat.system() == "Windows" else thumbnail_path
    if _local_thumb:
        try:
            if os.path.exists(_local_thumb):
                os.remove(_local_thumb)
        except Exception:
            pass
    return True, f"Removed DB record for: {file_path}"

def backup_database(max_backups=5):
    """
    Creates a backup of the database file and rotates old backups.
    """
    if not os.path.exists(DB_PATH):
        return

    backup_dir = os.path.join(_PROJECT_ROOT, "backups")
    if not os.path.exists(backup_dir):
        os.makedirs(backup_dir)

    timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    backup_path = os.path.join(backup_dir, f"scoring_history_{timestamp}.fdb")

    try:
        # Copy file
        import shutil
        shutil.copy2(DB_PATH, backup_path)
        print(f"Database backup created: {backup_path}")

        # Rotate
        backups = sorted([os.path.join(backup_dir, f) for f in os.listdir(backup_dir) if f.startswith("scoring_history_") and f.endswith(".fdb")])
        
        while len(backups) > max_backups:
            oldest = backups.pop(0)
            try:
                os.remove(oldest)
                print(f"Removed old backup: {oldest}")
            except Exception as e:
                print(f"Failed to remove old backup {oldest}: {e}")

    except Exception as e:
        print(f"Backup failed: {e}")

def update_image_metadata(file_path, keywords, title, description, rating, label):
    """
    Updates the metadata fields for a given image path.
    """
    try:
        get_connector().execute(
            "UPDATE images SET keywords = ?, title = ?, description = ?, rating = ?, label = ? WHERE file_path = ?",
            (keywords, title, description, rating, label, file_path),
        )
        row = get_connector().query_one("SELECT id FROM images WHERE file_path = ?", (file_path,))
        if row:
            _sync_image_keywords(row["id"], keywords)

        # Broadcast image update
        try:
            from modules.events import event_manager
            event_manager.broadcast_threadsafe("image_updated", {
                "file_path": file_path,
                "updates": {
                    "keywords": keywords,
                    "title": title,
                    "description": description,
                    "rating": rating,
                    "label": label
                }
            })
        except Exception: pass

        return True
    except Exception as e:
        logging.error(f"Failed to update metadata for {file_path}: {e}")
        return False

def get_incomplete_records(limit: int | None = None):
    """
    Retrieves records that have missing scores or metadata.
    Criteria:
    - Composite scores (score_general / score_technical) missing or non-positive
    - Legacy score column missing or non-positive
    - Any model score <= 0 or NULL
    - Rating <= 0 or NULL
    - Label empty or NULL
    """
    score_checks = [
        "score_general IS NULL OR score_general <= 0",
        "score_technical IS NULL OR score_technical <= 0",
    ]
    models = ['spaq', 'ava', 'koniq', 'paq2piq', 'liqe']
    for m in models:
        score_checks.append(f"score_{m} IS NULL OR score_{m} <= 0")

    score_cond = " OR ".join(score_checks)

    query = f"""
        SELECT * FROM images
        WHERE
            (score IS NULL OR score <= 0) OR
            (rating IS NULL OR rating <= 0) OR
            (label IS NULL OR TRIM(label) = '') OR
            ({score_cond})
        ORDER BY created_at DESC NULLS LAST
    """
    if limit is not None and limit > 0:
        query = query.strip() + f"\n        FETCH FIRST {int(limit)} ROWS ONLY"

    return list(get_connector().query(query))

def export_db_to_json(output_path, folder_path=None, keyword_filter=None, rating_filter=None,
                      label_filter=None, min_score_general=0, min_score_aesthetic=0,
                      min_score_technical=0, date_range=None):
    """
    Exports the images table to a JSON file with optional filtering.
    
    Args:
        output_path: Path for the output JSON file
        folder_path: Optional folder path prefix to filter by
        keyword_filter: Optional keyword string to search for
        rating_filter: Optional list of ratings to filter by
        label_filter: Optional list of labels to filter by
        min_score_general, min_score_aesthetic, min_score_technical: Score thresholds
        date_range: Optional tuple (start_date, end_date) as strings "YYYY-MM-DD"
    
    Returns (success, message)
    """
    import json

    conditions, params = _build_export_where_clause(
        rating_filter, label_filter, keyword_filter,
        min_score_general, min_score_aesthetic, min_score_technical,
        date_range, folder_path
    )

    query = "SELECT * FROM images"
    if conditions:
        query += " WHERE " + " AND ".join(conditions)
    query += " ORDER BY id"
    rows = get_connector().query(query, tuple(params) if params else None)

    data = []
    for row in rows:
        item = dict(row)
        # Handle datetime serialization if necessary (e.g. created_at)
        # SQLite returns strings for timestamps usually, but just in case
        if 'created_at' in item and item['created_at']:
            item['created_at'] = str(item['created_at'])
            
        # Parse nested JSON strings for cleaner output?
        # scores_json is already a JSON string in DB. 
        # If we want the export to be a clean JSON object, we should probably parse it back to dict.
        # But for a raw backup, keeping it as string is safer. 
        # Let's try to parse it to make the export more usable.
        if 'scores_json' in item and isinstance(item['scores_json'], str):
            try:
                item['scores_json'] = json.loads(item['scores_json'])
            except (json.JSONDecodeError, ValueError):
                pass # Leave as string if fail

        if 'metadata' in item and isinstance(item['metadata'], str):
            try:
                item['metadata'] = json.loads(item['metadata'])
            except (json.JSONDecodeError, ValueError):
                pass
                
        data.append(item)
        
    try:
        with open(output_path, 'w', encoding='utf-8') as f:
            json.dump(data, f, indent=4, ensure_ascii=False)
        return True, f"Successfully exported {len(data)} records to {output_path}"
    except Exception as e:
        return False, f"Export failed: {e}"


def get_available_columns():
    """Returns list of all available columns in the images table."""
    if _get_db_engine() == "postgres":
        rows = get_connector().query(
            "SELECT column_name FROM information_schema.columns WHERE table_name = 'images' ORDER BY ordinal_position")
        return [row["column_name"] for row in rows]
    # Firebird: query system catalog
    rows = get_connector().query(
        "SELECT RDB$FIELD_NAME FROM RDB$RELATION_FIELDS WHERE RDB$RELATION_NAME = 'IMAGES'")
    return [row["rdb$field_name"].strip().lower() for row in rows]


def _build_export_where_clause(rating_filter=None, label_filter=None, keyword_filter=None,
                                min_score_general=0, min_score_aesthetic=0, min_score_technical=0,
                                date_range=None, folder_path=None):
    """
    Helper function to build WHERE clause for export queries.
    Returns (conditions_list, params_list)
    """
    conditions = []
    params = []

    if rating_filter:
        placeholders = ','.join(['?'] * len(rating_filter))
        conditions.append(f"rating IN ({placeholders})")
        params.extend(rating_filter)

    if label_filter:
        clean_labels = [l for l in label_filter if l != "None"]
        has_none = "None" in label_filter

        lbl_conds = []
        if clean_labels:
            placeholders = ','.join(['?'] * len(clean_labels))
            lbl_conds.append(f"label IN ({placeholders})")
            params.extend(clean_labels)

        if has_none:
            lbl_conds.append("(label IS NULL OR label = '')")

        if lbl_conds:
            conditions.append(f"({' OR '.join(lbl_conds)})")

    _add_keyword_filter(conditions, params, keyword_filter)

    # Score Filters
    if min_score_general > 0:
        conditions.append("score_general >= ?")
        params.append(min_score_general)
    
    if min_score_aesthetic > 0:
        conditions.append("score_aesthetic >= ?")
        params.append(min_score_aesthetic)

    if min_score_technical > 0:
        conditions.append("score_technical >= ?")
        params.append(min_score_technical)
        
    # Date Filter (Firebird-compatible: CAST instead of DATE())
    if date_range and len(date_range) == 2:
        start_date, end_date = date_range
        if start_date:
            conditions.append("CAST(created_at AS DATE) >= CAST(? AS DATE)")
            params.append(start_date)
        if end_date:
            conditions.append("CAST(created_at AS DATE) <= CAST(? AS DATE)")
            params.append(end_date)
    
    # Folder Filter
    if folder_path:
        conditions.append("file_path LIKE ?")
        params.append(f"{folder_path}%")
    
    return conditions, params


def export_db_to_csv(output_path, columns=None, rating_filter=None, label_filter=None, 
                     keyword_filter=None, min_score_general=0, min_score_aesthetic=0, 
                     min_score_technical=0, date_range=None, folder_path=None):
    """
    Exports the images table to a CSV file with optional filtering.
    
    Args:
        output_path: Path for the output CSV file
        columns: Optional list of column names to export. If None, exports common columns.
        rating_filter: Optional list of ratings to filter by
        label_filter: Optional list of labels to filter by
        keyword_filter: Optional keyword string to search for
        min_score_general: Minimum general score threshold
        min_score_aesthetic: Minimum aesthetic score threshold
        min_score_technical: Minimum technical score threshold
        date_range: Optional tuple (start_date, end_date) as strings "YYYY-MM-DD"
        folder_path: Optional folder path prefix to filter by
    
    Returns (success, message)
    """
    import csv
    
    # Default columns for export (most useful ones)
    default_columns = [
        'id', 'file_path', 'file_name', 'file_type',
        'score_general', 'score_technical', 'score_aesthetic',
        'score_spaq', 'score_ava', 'score_koniq', 'score_paq2piq', 'score_liqe',
        'rating', 'label', 'keywords', 'title', 'description',
        'stack_id', 'created_at'
    ]
    
    columns_to_export = columns if columns else default_columns

    try:
        existing_cols = set(get_available_columns())
        valid_columns = [col for col in columns_to_export if col in existing_cols]

        if not valid_columns:
            return False, "No valid columns to export"

        conditions, params = _build_export_where_clause(
            rating_filter, label_filter, keyword_filter,
            min_score_general, min_score_aesthetic, min_score_technical,
            date_range, folder_path
        )

        query = f"SELECT {', '.join(valid_columns)} FROM images"
        if conditions:
            query += " WHERE " + " AND ".join(conditions)
        query += " ORDER BY id"

        rows = get_connector().query(query, tuple(params) if params else None)

        with open(output_path, 'w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)
            writer.writerow(valid_columns)
            for row in rows:
                writer.writerow([row.get(col) for col in valid_columns])

        return True, f"Successfully exported {len(rows)} records to {output_path}"

    except Exception as e:
        return False, f"CSV export failed: {e}"


def export_db_to_excel(output_path, columns=None, rating_filter=None, label_filter=None,
                       keyword_filter=None, min_score_general=0, min_score_aesthetic=0,
                       min_score_technical=0, date_range=None, folder_path=None):
    """
    Exports the images table to an Excel file with optional filtering.
    Requires openpyxl to be installed.
    
    Args:
        output_path: Path for the output Excel file (.xlsx)
        columns: Optional list of column names to export. If None, exports common columns.
        rating_filter: Optional list of ratings to filter by
        label_filter: Optional list of labels to filter by
        keyword_filter: Optional keyword string to search for
        min_score_general: Minimum general score threshold
        min_score_aesthetic: Minimum aesthetic score threshold
        min_score_technical: Minimum technical score threshold
        date_range: Optional tuple (start_date, end_date) as strings "YYYY-MM-DD"
        folder_path: Optional folder path prefix to filter by
    
    Returns (success, message)
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        return False, "openpyxl is required for Excel export. Install with: pip install openpyxl"
    
    # Default columns for export
    default_columns = [
        'id', 'file_path', 'file_name', 'file_type',
        'score_general', 'score_technical', 'score_aesthetic',
        'score_spaq', 'score_ava', 'score_koniq', 'score_paq2piq', 'score_liqe',
        'rating', 'label', 'keywords', 'title', 'description',
        'stack_id', 'created_at'
    ]
    
    columns_to_export = columns if columns else default_columns

    try:
        existing_cols = set(get_available_columns())
        valid_columns = [col for col in columns_to_export if col in existing_cols]

        if not valid_columns:
            return False, "No valid columns to export"

        conditions, params = _build_export_where_clause(
            rating_filter, label_filter, keyword_filter,
            min_score_general, min_score_aesthetic, min_score_technical,
            date_range, folder_path
        )

        query = f"SELECT {', '.join(valid_columns)} FROM images"
        if conditions:
            query += " WHERE " + " AND ".join(conditions)
        query += " ORDER BY id"

        rows = get_connector().query(query, tuple(params) if params else None)

        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Image Scores"

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")

        for col_idx, col_name in enumerate(valid_columns, 1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')

        for row_idx, row in enumerate(rows, 2):
            for col_idx, col_name in enumerate(valid_columns, 1):
                ws.cell(row=row_idx, column=col_idx, value=row.get(col_name))

        for col_idx, col_name in enumerate(valid_columns, 1):
            max_length = len(col_name)
            for row in rows[:100]:
                val = row.get(col_name)
                if val:
                    max_length = max(max_length, min(len(str(val)), 50))
            ws.column_dimensions[chr(64 + col_idx) if col_idx <= 26 else 'A'].width = max_length + 2

        ws.freeze_panes = 'A2'
        wb.save(output_path)

        return True, f"Successfully exported {len(rows)} records to {output_path}"

    except Exception as e:
        return False, f"Excel export failed: {e}"


# --- Stack Management ---

def clear_stacks():
    """
    Clears all stacks and resets stack_id in images.
    """
    try:
        def _tx(tx):
            tx.execute("DELETE FROM stacks")
            tx.execute("UPDATE images SET stack_id = NULL")
        get_connector().run_transaction(_tx)
        event_manager.broadcast_threadsafe("stacks_cleared", {})
    except Exception as e:
        logging.error(f"Failed to clear stacks: {e}")

def create_stack(name, best_image_id=None):
    """
    Creates a new stack.
    """
    stack_id = None
    try:
        rows = get_connector().execute_returning(
            "INSERT INTO stacks (name, best_image_id, created_at) VALUES (?, ?, ?) RETURNING id",
            (name, best_image_id, datetime.datetime.now()),
        )
        stack_id = rows[0]["id"] if rows else None
        if stack_id:
            event_manager.broadcast_threadsafe("stack_created", {"stack_id": stack_id})
    except Exception as e:
        logging.error(f"Failed to create stack: {e}")
    return stack_id

def update_image_fields_batch(updates):
    """
    Batch updates multiple fields for multiple images.
    updates: list of (image_id, dict) where dict has field_name -> value.
    Valid fields: keywords, title, description, rating, label, etc. (see valid_fields in update_image_field)
    """
    valid_fields = {
        'burst_uuid', 'rating', 'label', 'score_general', 'score_aesthetic',
        'score_technical', 'keywords', 'title', 'description', 'stack_id',
        'thumbnail_path', 'thumbnail_path_win', 'metadata', 'image_hash',
        'cull_decision', 'cull_policy_version', 'image_uuid'
    }
    if not updates:
        return
    try:
        def _tx(tx):
            for image_id, fields in updates:
                if not isinstance(fields, dict):
                    continue
                for fname, val in fields.items():
                    if fname not in valid_fields:
                        continue
                    tx.execute(f"UPDATE images SET {fname} = ? WHERE id = ?", (val, image_id))
        get_connector().run_transaction(_tx)

        # Keyword Sync (Dual Writing)
        for image_id, fields in updates:
            if isinstance(fields, dict) and 'keywords' in fields:
                try:
                    _sync_image_keywords(image_id, fields['keywords'])
                except Exception as e:
                    logging.warning(f"Batch keyword sync failed for image {image_id}: {e}")

        invalidate_folder_images_cache()
        for image_id, fields in updates:
            if isinstance(fields, dict):
                try:
                    event_manager.broadcast_threadsafe("image_updated", {
                        "image_id": image_id,
                        "updates": fields
                    })
                except Exception:
                    pass
    except Exception as e:
        logging.error(f"Failed batch update_image_fields: {e}")


def update_image_stack_batch(updates):
    """
    Batch updates image stack_ids.
    updates: list of (stack_id, image_id) tuples
    """
    try:
        get_connector().execute_many("UPDATE images SET stack_id = ? WHERE id = ?", list(updates))

        invalidate_folder_images_cache()
        for stack_id, image_id in updates:
            event_manager.broadcast_threadsafe("image_updated", {
                "image_id": image_id,
                "updates": {"stack_id": stack_id}
            })
    except Exception as e:
        logging.error(f"Failed to batch update image stacks: {e}")


def batch_update_cull_decisions(updates: list, policy_version: str = "1.0", batch_size: int = 1000):
    """
    Batch updates cull_decision and cull_policy_version for images.
    updates: list of (image_id, cull_decision, file_path) tuples.
    cull_decision: 'pick' | 'reject' | 'neutral'
    """
    if not updates:
        return
    try:
        all_params = [(decision, policy_version, img_id) for img_id, decision, _ in updates]
        get_connector().execute_many(
            "UPDATE images SET cull_decision = ?, cull_policy_version = ? WHERE id = ?",
            all_params,
        )

        invalidate_folder_images_cache()
        for img_id, decision, file_path in updates:
            event_manager.broadcast_threadsafe("image_updated", {
                "image_id": img_id,
                "file_path": file_path,
                "updates": {"cull_decision": decision}
            })
    except Exception as e:
        logging.error("Failed to batch update cull decisions: %s", e)


def get_stacks():
    """
    Returns all stacks joined with their best image info.
    """
    sql = """
        SELECT s.*, i.file_path as best_image_path, i.score_general as best_image_score,
        (SELECT COUNT(*) FROM images WHERE stack_id = s.id) as image_count
        FROM stacks s
        LEFT JOIN images i ON s.best_image_id = i.id
        ORDER BY s.id ASC
    """
    return list(get_connector().query(sql))


def get_stack_context(image_id):
    """
    Get stack context for a single image.
    Returns dict with stack_id, stack_size, is_best, stack_name or None if not in stack.
    """
    row = get_connector().query_one("SELECT stack_id FROM images WHERE id = ?", (image_id,))
    if not row or not row['stack_id']:
        return None

    stack_id = row['stack_id']

    # Get stack info
    stack_row = get_connector().query_one("""
        SELECT
            s.name,
            s.best_image_id,
            (SELECT COUNT(*) FROM images WHERE stack_id = s.id) as stack_size
        FROM stacks s
        WHERE s.id = ?
    """, (stack_id,))

    if not stack_row:
        return None

    return {
        'stack_id': stack_id,
        'stack_name': stack_row['name'],
        'stack_size': stack_row['stack_size'],
        'is_best': stack_row['best_image_id'] == image_id
    }


def get_stack_contexts_batch(image_ids):
    """
    Get stack context for multiple images in a single query.
    Returns dict mapping image_id to stack context.
    Efficient for gallery display with many images.
    """
    if not image_ids:
        return {}

    # Query all images and their stack info in one go
    placeholders = ','.join(['?'] * len(image_ids))
    query = f"""
        SELECT
            i.id as image_id,
            i.stack_id,
            s.name as stack_name,
            s.best_image_id,
            (SELECT COUNT(*) FROM images i2 WHERE i2.stack_id = s.id) as stack_size
        FROM images i
        LEFT JOIN stacks s ON i.stack_id = s.id
        WHERE i.id IN ({placeholders})
    """
    rows = list(get_connector().query(query, list(image_ids)))

    result = {}
    for row in rows:
        if row['stack_id']:
            result[row['image_id']] = {
                'stack_id': row['stack_id'],
                'stack_name': row['stack_name'],
                'stack_size': row['stack_size'],
                'is_best': row['best_image_id'] == row['image_id']
            }

    return result


def get_stacks_for_display(folder_path=None, sort_by="score_general", order="desc"):
    """
    Returns stacks with dynamic cover image based on sort criteria.
    Uses CTE with ROW_NUMBER() instead of correlated subquery for better performance.
    """
    # Resolve folder_id if path provided
    folder_id = None
    if folder_path:
        folder_id = get_or_create_folder(folder_path)
        if not folder_id:
            return []

    # Map sort_by to column
    # If sort_by is invalid, default to score_general
    valid_cols = ["created_at", "id", "score_general", "score_technical", "score_aesthetic",
                  "score_spaq", "score_ava", "score_koniq", "score_paq2piq", "score_liqe"]
    if sort_by not in valid_cols:
        sort_by = "score_general"

    agg_func = "MAX" if order.lower() == "desc" else "MIN"
    order_dir = order.upper()

    # Use CTE with ROW_NUMBER() to compute cover images in a single pass
    # This avoids the N+1 correlated subquery problem

    where_clause = ""
    params = []
    cte_where = ""
    if folder_id:
        where_clause = "WHERE i.folder_id = ?"
        cte_where = "AND i2.folder_id = ?"
        # Need 2 params: one for main WHERE clause, one for subquery
        params.append(folder_id)  # For main SELECT WHERE clause
        params.append(folder_id)  # For subquery WHERE clause

    # db.py is called from both WSL and Windows; pick the SQL column
    # that matches the caller's environment so the returned cover_path is usable directly.
    import platform as _plat
    _thumb_col = "thumbnail_path_win" if _plat.system() == "Windows" else "thumbnail_path"

    query = f'''
        SELECT
            s.id,
            s.name,
            COUNT(i.id) as image_count,
            {agg_func}(i.{sort_by}) as sort_val,
            (SELECT FIRST 1 COALESCE(NULLIF(i2.{_thumb_col}, ''), NULLIF(i2.thumbnail_path, ''), i2.file_path)
             FROM images i2
             WHERE i2.stack_id = s.id {cte_where}
             ORDER BY i2.{sort_by} {order_dir}) as cover_path
        FROM stacks s
        JOIN images i ON s.id = i.stack_id
        {where_clause}
        GROUP BY s.id, s.name
        ORDER BY sort_val {order_dir}
    '''

    rows = list(get_connector().query(query, tuple(params)))

    # Path-prefix fallback: if folder_id filter returns 0 rows, retry with file_path LIKE (handles
    # folder_id mismatch, NULL folder_id, or WSL vs Windows path differences).
    if folder_path and len(rows) == 0:
        try:
            from modules import utils
            fp_norm = os.path.normpath(folder_path).rstrip("/\\")
            p_win = fp_norm + os.sep
            p_win_fwd = fp_norm.replace("\\", "/").rstrip("/") + "/"
            p_wsl = utils.convert_path_to_wsl(fp_norm).rstrip("/") + "/"
            where_clause = "WHERE (i.file_path LIKE ? OR i.file_path LIKE ? OR i.file_path LIKE ?)"
            params_fb = [p_win + "%", p_win_fwd + "%", p_wsl + "%"]
            query_fb = f'''
                WITH ranked_covers AS (
                    SELECT stack_id,
                        COALESCE(NULLIF({_thumb_col}, ''), NULLIF(thumbnail_path, ''), file_path) as cover_path,
                        ROW_NUMBER() OVER (PARTITION BY stack_id ORDER BY {sort_by} {order_dir}) as rn
                    FROM images WHERE stack_id IS NOT NULL
                )
                SELECT s.id, s.name, COUNT(i.id) as image_count,
                    {agg_func}(i.{sort_by}) as sort_val, rc.cover_path
                FROM stacks s
                JOIN images i ON s.id = i.stack_id
                LEFT JOIN ranked_covers rc ON s.id = rc.stack_id AND rc.rn = 1
                {where_clause}
                GROUP BY s.id, s.name, rc.cover_path
                ORDER BY sort_val {order_dir}
            '''
            rows = list(get_connector().query(query_fb, tuple(params_fb)))
        except Exception:
            pass

    return rows

def get_images_in_stack(stack_id):
    """
    Returns all images in a stack.
    """
    return list(get_connector().query("SELECT * FROM images WHERE stack_id = ? ORDER BY score_general DESC", (stack_id,)))

def get_stack_count():
    row = get_connector().query_one("SELECT COUNT(*) AS cnt FROM stacks")
    return row["cnt"] if row else 0

def get_clustered_folders():
    """
    Returns a set of folders that have been clustered.
    """
    try:
        rows = list(get_connector().query("SELECT folder_path FROM cluster_progress"))
        return {row["folder_path"] for row in rows}
    except Exception as e:
        logging.error(f"Error reading cluster progress: {e}")
        return set()

def mark_folder_clustered(folder_path):
    """
    Marks a folder as successfully clustered.
    """
    try:
        get_connector().execute(
            "UPDATE OR INSERT INTO cluster_progress (folder_path, last_run) VALUES (?, ?) MATCHING (folder_path)",
            (folder_path, datetime.datetime.now())
        )
        event_manager.broadcast_threadsafe("folder_updated", {"folder_path": folder_path})
    except Exception as e:
        logging.error(f"Failed to mark folder as clustered: {e}")

def clear_cluster_progress():
    """
    Clears cluster progress and stacks.
    Also resets culling phase status from running to done to allow force re-run.
    """
    try:
        # Get all images with running culling phase, then reset them
        rows = list(get_connector().query(
            """SELECT DISTINCT ips.image_id
               FROM IMAGE_PHASE_STATUS ips
               JOIN pipeline_phases pp ON ips.phase_id = pp.id
               WHERE pp.code = ? AND ips.status = ?""",
            ("culling", "running")
        ))
        running_image_ids = [row["image_id"] for row in rows]
        logging.info(f"[Force Rescan - All Folders] Found {len(running_image_ids)} images with running culling phase, resetting to done")

        # Reset each image's culling phase to done
        for image_id in running_image_ids:
            logging.debug(f"[Force Rescan - All Folders] Resetting culling phase for image {image_id} from running to done")
            set_image_phase_status(image_id, "culling", "done")

        # Now clear stacks
        def _tx(tx):
            tx.execute("DELETE FROM cluster_progress")
            tx.execute("DELETE FROM stacks")
            tx.execute("UPDATE images SET stack_id = NULL")
        get_connector().run_transaction(_tx)
    except Exception as e:
        logging.error(f"Failed to clear cluster progress: {e}")


def clear_stacks_in_folder(folder_path):
    """
    Clears stacks for images in a specific folder only.
    Used for targeted re-clustering of a single folder.
    
    Steps:
    1. Get all image IDs in the folder
    2. Get stack_ids for those images
    3. Set stack_id = NULL for those images
    4. Delete stacks that are now empty
    5. Remove folder from cluster_progress
    
    Returns (success, message)
    """
    import os
    folder_path = os.path.normpath(folder_path)

    try:
        def _tx(tx):
            folder_row = tx.query_one("SELECT id FROM folders WHERE path = ?", (folder_path,))

            if not folder_row:
                stack_rows = tx.query("""
                    SELECT DISTINCT stack_id FROM images
                    WHERE file_path LIKE ? AND stack_id IS NOT NULL
                """, (folder_path + '%',))
            else:
                f_id = folder_row["id"]
                stack_rows = tx.query("""
                    SELECT DISTINCT stack_id FROM images
                    WHERE folder_id = ? AND stack_id IS NOT NULL
                """, (f_id,))

            affected_stacks = [r["stack_id"] for r in stack_rows]

            if not affected_stacks:
                tx.execute("DELETE FROM cluster_progress WHERE folder_path = ?", (folder_path,))
                return 0, 0, []

            if folder_row:
                updated_count = tx.execute("UPDATE images SET stack_id = NULL WHERE folder_id = ?", (folder_row["id"],))
            else:
                updated_count = tx.execute("UPDATE images SET stack_id = NULL WHERE file_path LIKE ?", (folder_path + '%',))

            deleted_stacks = 0
            for sid in affected_stacks:
                cnt_row = tx.query_one("SELECT COUNT(*) AS cnt FROM images WHERE stack_id = ?", (sid,))
                remaining = cnt_row["cnt"] if cnt_row else 0
                if remaining == 0:
                    tx.execute("DELETE FROM stacks WHERE id = ?", (sid,))
                    deleted_stacks += 1
                else:
                    tx.execute("""
                        UPDATE stacks SET best_image_id = (
                            SELECT id FROM images
                            WHERE stack_id = ?
                            ORDER BY score_general DESC NULLS LAST
                            FETCH FIRST 1 ROWS ONLY
                        ) WHERE id = ?
                    """, (sid, sid))

            tx.execute("DELETE FROM cluster_progress WHERE folder_path = ?", (folder_path,))

            if folder_row:
                run_rows = tx.query(
                    """SELECT i.id
                       FROM images i
                       JOIN IMAGE_PHASE_STATUS ips ON i.id = ips.image_id
                       JOIN pipeline_phases pp ON ips.phase_id = pp.id
                       WHERE i.folder_id = ? AND pp.code = ? AND ips.status = ?""",
                    (folder_row["id"], "culling", "running"))
            else:
                run_rows = tx.query(
                    """SELECT DISTINCT i.id
                       FROM images i
                       JOIN IMAGE_PHASE_STATUS ips ON i.id = ips.image_id
                       JOIN pipeline_phases pp ON ips.phase_id = pp.id
                       WHERE i.file_path LIKE ? AND pp.code = ? AND ips.status = ?""",
                    (folder_path + '%', "culling", "running"))
            running_image_ids = [r["id"] for r in run_rows]

            if running_image_ids:
                logging.info(f"[Force Rescan] Found {len(running_image_ids)} images with running culling phase, resetting to done")

            return updated_count, deleted_stacks, running_image_ids

        updated_count, deleted_stacks, running_image_ids = get_connector().run_transaction(_tx)

        if not updated_count and not deleted_stacks and not running_image_ids:
            return True, f"No stacks found in folder: {folder_path}"

        for image_id in running_image_ids:
            logging.debug(f"[Force Rescan] Resetting culling phase for image {image_id} from running to done")
            set_image_phase_status(image_id, "culling", "done")

        invalidate_folder_images_cache(folder_path)
        event_manager.broadcast_threadsafe("folder_updated", {"folder_path": folder_path})
        event_manager.broadcast_threadsafe("stacks_cleared", {"folder_path": folder_path})

        msg = f"Cleared {deleted_stacks} stacks, updated {updated_count} images in folder: {folder_path}"
        logging.info(msg)
        return True, msg

    except Exception as e:
        logging.error(f"Failed to clear stacks in folder {folder_path}: {e}")
        return False, str(e)


def create_stacks_batch(stacks_data):
    """
    Creates multiple stacks and updates associations in a single transaction.
    stacks_data: list of dicts { 'name': str, 'best_image_id': int, 'image_ids': [int] }
    """
    created_ids = []
    timestamp = datetime.datetime.now()

    def _tx(tx):
        for data in stacks_data:
            # Create Stack
            ret = tx.execute_returning(
                "INSERT INTO stacks (name, best_image_id, created_at) VALUES (?, ?, ?) RETURNING id",
                (data['name'], data['best_image_id'], timestamp)
            )
            stack_id = ret[0]["id"] if ret else None
            created_ids.append(stack_id)
            # Update Images
            if stack_id and data.get('image_ids'):
                for img_id in data['image_ids']:
                    tx.execute("UPDATE images SET stack_id = ? WHERE id = ?", (stack_id, img_id))

    try:
        get_connector().run_transaction(_tx)
        created_count = len(stacks_data)
        # Broadcast outside transaction
        for stack_id in created_ids:
            if stack_id:
                event_manager.broadcast_threadsafe("stack_created", {"stack_id": stack_id})
        return True, f"Created {created_count} stacks."
    except Exception as e:
        logging.error(f"Failed to batch create stacks: {e}")
        return False, str(e)

# --- Manual Stack Operations ---

def create_stack_from_images(image_ids, name=None):
    """
    Creates a new stack from a list of image IDs (manual grouping).
    Returns (success, stack_id or error message).
    """
    if not image_ids or len(image_ids) < 2:
        return False, "Need at least 2 images to create a stack"

    try:
        def _tx(tx):
            nonlocal name
            placeholders = ','.join(['?'] * len(image_ids))
            rows = tx.query(f"SELECT id, score_general FROM images WHERE id IN ({placeholders})", tuple(image_ids))
            if len(rows) != len(image_ids):
                raise ValueError(f"Some images not found. Expected {len(image_ids)}, found {len(rows)}")

            best_id = None
            best_score = -1
            for r in rows:
                score = r["score_general"] if r.get("score_general") else 0
                if score > best_score:
                    best_score = score
                    best_id = r["id"]

            if not name:
                cnt_row = tx.query_one("SELECT COUNT(*) AS cnt FROM stacks")
                stack_count = cnt_row["cnt"] if cnt_row else 0
                timestamp = datetime.datetime.now().strftime("%Y-%m-%d")
                name = f"Stack {timestamp} #{stack_count + 1:03d}"

            ins_rows = tx.execute_returning(
                "INSERT INTO stacks (name, best_image_id, created_at) VALUES (?, ?, ?) RETURNING id",
                (name, best_id, datetime.datetime.now()))
            sid = ins_rows[0]["id"] if ins_rows else None

            updates = [(sid, img_id) for img_id in image_ids]
            tx.execute_many("UPDATE images SET stack_id = ? WHERE id = ?", updates)
            return sid

        stack_id = get_connector().run_transaction(_tx)
        if stack_id:
            event_manager.broadcast_threadsafe("stack_created", {"stack_id": stack_id})
        return True, stack_id
    except ValueError as e:
        return False, str(e)
    except Exception as e:
        logging.error(f"Failed to create stack from images: {e}")
        return False, str(e)


def remove_images_from_stack(image_ids):
    """
    Removes images from their current stacks (sets stack_id to NULL).
    Also cleans up empty stacks after removal.
    Returns (success, message).
    """
    if not image_ids:
        return False, "No images specified"

    try:
        def _tx(tx):
            placeholders = ','.join(['?'] * len(image_ids))
            aff_rows = tx.query(
                f"SELECT DISTINCT stack_id FROM images WHERE id IN ({placeholders}) AND stack_id IS NOT NULL",
                tuple(image_ids))
            affected_stacks = [r["stack_id"] for r in aff_rows]

            removed_count = tx.execute(
                f"UPDATE images SET stack_id = NULL WHERE id IN ({placeholders})", tuple(image_ids))

            deleted_stacks = 0
            for sid in affected_stacks:
                cnt_row = tx.query_one("SELECT COUNT(*) AS cnt FROM images WHERE stack_id = ?", (sid,))
                remaining = cnt_row["cnt"] if cnt_row else 0
                if remaining == 0:
                    tx.execute("DELETE FROM stacks WHERE id = ?", (sid,))
                    deleted_stacks += 1
                else:
                    tx.execute("""
                        UPDATE stacks SET best_image_id = (
                            SELECT id FROM images
                            WHERE stack_id = ?
                            ORDER BY score_general DESC NULLS LAST
                            FETCH FIRST 1 ROWS ONLY
                        ) WHERE id = ?
                    """, (sid, sid))
                    event_manager.broadcast_threadsafe("stack_updated", {"stack_id": sid})
            return removed_count, deleted_stacks

        removed_count, deleted_stacks = get_connector().run_transaction(_tx)

        for img_id in image_ids:
            event_manager.broadcast_threadsafe("image_updated", {
                "image_id": img_id,
                "updates": {"stack_id": None}
            })

        msg = f"Removed {removed_count} images from stacks"
        if deleted_stacks > 0:
            msg += f", deleted {deleted_stacks} empty stack(s)"
        return True, msg
    except Exception as e:
        logging.error(f"Failed to remove images from stack: {e}")
        return False, str(e)


def dissolve_stack(stack_id):
    """
    Completely dissolves a stack - removes all images and deletes the stack.
    Returns (success, message).
    """
    if not stack_id:
        return False, "No stack specified"

    try:
        def _tx(tx):
            cnt_row = tx.query_one("SELECT COUNT(*) AS cnt FROM images WHERE stack_id = ?", (stack_id,))
            image_count = cnt_row["cnt"] if cnt_row else 0

            name_row = tx.query_one("SELECT name FROM stacks WHERE id = ?", (stack_id,))
            stack_name = name_row["name"] if name_row else f"Stack #{stack_id}"

            tx.execute("UPDATE images SET stack_id = NULL WHERE stack_id = ?", (stack_id,))
            tx.execute("DELETE FROM stacks WHERE id = ?", (stack_id,))
            return stack_name, image_count

        stack_name, image_count = get_connector().run_transaction(_tx)
        event_manager.broadcast_threadsafe("stack_deleted", {"stack_id": stack_id})
        return True, f"Dissolved '{stack_name}' ({image_count} images ungrouped)"
    except Exception as e:
        logging.error(f"Failed to dissolve stack: {e}")
        return False, str(e)


def set_stack_cover_image(stack_id, image_id):
    """
    Sets a specific image as the cover (best_image_id) for a stack.
    Allows manual override of the auto-selected best image.
    
    Args:
        stack_id: ID of the stack to update
        image_id: ID of the image to set as cover
    
    Returns (success, message)
    """
    if not stack_id or not image_id:
        return False, "Stack ID and Image ID are required"

    try:
        stack_row = get_connector().query_one("SELECT name FROM stacks WHERE id = ?", (stack_id,))
        if not stack_row:
            return False, f"Stack {stack_id} not found"
        img_row = get_connector().query_one("SELECT file_name, stack_id FROM images WHERE id = ?", (image_id,))
        if not img_row:
            return False, f"Image {image_id} not found"
        if img_row["stack_id"] != stack_id:
            return False, f"Image {image_id} does not belong to stack {stack_id}"
        get_connector().execute("UPDATE stacks SET best_image_id = ? WHERE id = ?", (image_id, stack_id))
        event_manager.broadcast_threadsafe("stack_updated", {"stack_id": stack_id})
        return True, f"Set '{img_row['file_name']}' as cover for '{stack_row['name']}'"
    except Exception as e:
        logging.error(f"Failed to set stack cover image: {e}")
        return False, str(e)


def get_image_ids_by_paths(file_paths):
    """
    Returns image IDs for given file paths.
    Useful for converting gallery selection (paths) to DB IDs.
    """
    if not file_paths:
        return []

    ids = []
    try:
        for path in file_paths:
            row = get_connector().query_one("SELECT id FROM images WHERE file_path = ?", (path,))
            if row:
                ids.append(row["id"])
            else:
                basename = os.path.basename(path)
                row = get_connector().query_one("SELECT id FROM images WHERE file_name = ?", (basename,))
                if row:
                    logging.warning(f"Path lookup fallback used: {path} -> id {row['id']} (matched by filename)")
                    ids.append(row["id"])
    except Exception as e:
        logging.error(f"Failed to get image IDs by paths: {e}")
    return ids


# --- Culling Session Management ---

def create_culling_session(folder_path, mode='automated'):
    """
    Creates a new culling session for a folder.
    Returns session_id.
    """
    try:
        rows = get_connector().execute_returning(
            "INSERT INTO culling_sessions (folder_path, mode, status, created_at) VALUES (?, ?, 'active', ?) RETURNING id",
            (folder_path, mode, datetime.datetime.now()),
        )
        session_id = rows[0]["id"] if rows else None
        return session_id
    except Exception as e:
        logging.error(f"Failed to create culling session: {e}")
        return None


def get_culling_session(session_id):
    """Returns culling session details."""
    row = get_connector().query_one("SELECT * FROM culling_sessions WHERE id = ?", (session_id,))
    return dict(row) if row else None


def get_active_culling_sessions():
    """Returns all active (incomplete) culling sessions."""
    return [dict(r) for r in get_connector().query("SELECT * FROM culling_sessions WHERE status = 'active' ORDER BY created_at DESC")]


def update_culling_session(session_id, **kwargs):
    """
    Updates culling session fields.
    Accepts: status, total_images, total_groups, reviewed_groups, picked_count, rejected_count
    """
    allowed = ['status', 'total_images', 'total_groups', 'reviewed_groups', 
               'picked_count', 'rejected_count', 'completed_at']
    updates = {k: v for k, v in kwargs.items() if k in allowed}
    
    if not updates:
        return False

    try:
        set_clause = ', '.join([f"{k} = ?" for k in updates.keys()])
        params = list(updates.values()) + [session_id]
        get_connector().execute(f"UPDATE culling_sessions SET {set_clause} WHERE id = ?", params)
        return True
    except Exception as e:
        logging.error(f"Failed to update culling session: {e}")
        return False


def complete_culling_session(session_id):
    """Marks a culling session as completed."""
    return update_culling_session(
        session_id, 
        status='completed', 
        completed_at=datetime.datetime.now()
    )


# --- Culling Picks Management ---

def add_images_to_culling_session(session_id, image_ids, group_assignments=None):
    """
    Adds images to a culling session.
    group_assignments: dict of {image_id: group_id} if groups are pre-computed.
    """
    if not image_ids:
        logging.warning(f"No image_ids provided for session {session_id}")
        return False

    now = datetime.datetime.now()
    try:
        added_count = 0

        def _tx(tx):
            nonlocal added_count
            for img_id in image_ids:
                group_id = group_assignments.get(img_id) if group_assignments else None
                try:
                    tx.execute(
                        "UPDATE OR INSERT INTO culling_picks"
                        " (session_id, image_id, group_id, decision, auto_suggested, is_best_in_group, created_at)"
                        " VALUES (?, ?, ?, ?, ?, ?, ?)"
                        " MATCHING (session_id, image_id)",
                        (session_id, img_id, group_id, None, 0, 0, now),
                    )
                    added_count += 1
                except Exception as e:
                    logging.error(f"Failed to add image {img_id} to session {session_id}: {e}")

        get_connector().run_transaction(_tx)
        logging.info(f"Added {added_count}/{len(image_ids)} images to culling session {session_id}")
        return added_count > 0
    except Exception as e:
        logging.error(f"Failed to add images to culling session: {e}")
        logging.error(traceback.format_exc())
        return False


def set_pick_decision(session_id, image_id, decision, auto_suggested=False):
    """
    Sets pick/reject decision for an image in a culling session.
    decision: 'pick', 'reject', 'maybe', or None (to clear)
    """
    try:
        rowcount = get_connector().execute(
            "UPDATE culling_picks SET decision = ?, auto_suggested = ? WHERE session_id = ? AND image_id = ?",
            (decision, auto_suggested, session_id, image_id),
        )

        if rowcount > 0:
            event_manager.broadcast_threadsafe("image_updated", {
                "image_id": image_id,
                "updates": {"cull_decision": decision}
            })
        return rowcount > 0
    except Exception as e:
        logging.error(f"Failed to set pick decision: {e}")
        return False


def set_best_in_group(session_id, image_id, group_id):
    """Marks an image as the best in its group."""
    try:
        def _tx(tx):
            tx.execute(
                "UPDATE culling_picks SET is_best_in_group = ? WHERE session_id = ? AND group_id = ?",
                (0, session_id, group_id),
            )
            tx.execute(
                "UPDATE culling_picks SET is_best_in_group = ?, decision = 'pick', auto_suggested = ? WHERE session_id = ? AND image_id = ?",
                (1, 1, session_id, image_id),
            )
        get_connector().run_transaction(_tx)
        return True
    except Exception as e:
        logging.error(f"Failed to set best in group: {e}")
        return False


def get_session_picks(session_id, decision_filter=None):
    """
    Returns all picks for a session with image details.
    decision_filter: None for all, or 'pick', 'reject', 'maybe'
    """
    query = """
        SELECT cp.*, i.file_path, i.file_name, i.thumbnail_path, i.thumbnail_path_win,
               i.score_general, i.score_technical, i.score_aesthetic,
               i.rating, i.label
        FROM culling_picks cp
        JOIN images i ON cp.image_id = i.id
        WHERE cp.session_id = ?
    """
    params = [session_id]

    if decision_filter:
        query += " AND cp.decision = ?"
        params.append(decision_filter)

    query += " ORDER BY cp.group_id, i.score_general DESC"

    return [dict(row) for row in get_connector().query(query, tuple(params))]


def get_session_groups(session_id):
    """
    Returns images grouped by group_id for a session.
    Returns list of groups, each group is a dict with group info and list of images.
    """
    query = """
        SELECT cp.group_id, cp.image_id, cp.decision, cp.auto_suggested, cp.is_best_in_group,
               i.file_path, i.file_name, i.thumbnail_path, i.thumbnail_path_win,
               i.score_general, i.score_technical, i.score_aesthetic
        FROM culling_picks cp
        JOIN images i ON cp.image_id = i.id
        WHERE cp.session_id = ?
        ORDER BY cp.group_id, i.score_general DESC
    """

    rows = list(get_connector().query(query, (session_id,)))

    # Group by group_id
    groups = {}
    for row in rows:
        gid = row['group_id'] if row['group_id'] else 0  # Singles in group 0
        if gid not in groups:
            groups[gid] = {
                'group_id': gid,
                'images': [],
                'has_pick': False,
                'best_image_id': None
            }
        
        img = dict(row)
        groups[gid]['images'].append(img)
        
        if img['decision'] == 'pick':
            groups[gid]['has_pick'] = True
        if img['is_best_in_group']:
            groups[gid]['best_image_id'] = img['image_id']
    
    return list(groups.values())


def get_session_stats(session_id):
    """Returns statistics for a culling session."""
    con = get_connector()

    total    = (con.query_one("SELECT COUNT(*) AS n FROM culling_picks WHERE session_id = ?", (session_id,)) or {}).get("n", 0) or 0
    picked   = (con.query_one("SELECT COUNT(*) AS n FROM culling_picks WHERE session_id = ? AND decision = 'pick'", (session_id,)) or {}).get("n", 0) or 0
    rejected = (con.query_one("SELECT COUNT(*) AS n FROM culling_picks WHERE session_id = ? AND decision = 'reject'", (session_id,)) or {}).get("n", 0) or 0
    groups   = (con.query_one("SELECT COUNT(DISTINCT group_id) AS n FROM culling_picks WHERE session_id = ? AND group_id IS NOT NULL", (session_id,)) or {}).get("n", 0) or 0
    # Groups with at least one decision
    reviewed = (con.query_one("SELECT COUNT(DISTINCT group_id) AS n FROM culling_picks WHERE session_id = ? AND group_id IS NOT NULL AND decision IS NOT NULL", (session_id,)) or {}).get("n", 0) or 0

    return {
        'total_images': int(total),
        'total_groups': int(groups),
        'reviewed_groups': int(reviewed),
        'picked_count': int(picked),
        'rejected_count': int(rejected),
        'unreviewed': int(total) - int(picked) - int(rejected),
    }


def clear_culling_picks(session_id):
    """
    Removes all picks from a culling session.
    Used before re-importing groups from updated stacks.
    """
    try:
        deleted = get_connector().execute("DELETE FROM culling_picks WHERE session_id = ?", (session_id,))
        logging.info(f"Cleared {deleted} picks from session {session_id}")
        return True
    except Exception as e:
        logging.error(f"Failed to clear picks for session {session_id}: {e}")
        return False


def reset_culling_decisions(session_id):
    """
    Resets all decisions (pick/reject) in a session without removing the picks.
    Used before re-running auto-pick.
    """
    try:
        updated = get_connector().execute(
            "UPDATE culling_picks SET decision = NULL, auto_suggested = 0, is_best_in_group = 0 WHERE session_id = ?",
            (session_id,),
        )
        logging.info(f"Reset {updated} decisions in session {session_id}")
        return True
    except Exception as e:
        logging.error(f"Failed to reset decisions for session {session_id}: {e}")
        return False


def get_image_culling_status(file_path):
    """
    Returns the most recent culling decision for an image.
    Returns dict with 'decision' ('pick', 'reject', 'maybe', or None) and 'session_id'.
    'pick' = Accepted, 'reject' = Rejected
    """
    try:
        con = get_connector()
        # First get image_id from file_path
        row = con.query_one("SELECT id FROM images WHERE file_path = ?", (file_path,))
        if not row:
            return None

        image_id = row["id"]

        # Get most recent culling decision for this image
        pick_row = con.query_one(
            """SELECT cp.decision, cp.session_id, cp.is_best_in_group, cs.folder_path
               FROM culling_picks cp
               JOIN culling_sessions cs ON cp.session_id = cs.id
               WHERE cp.image_id = ?
               ORDER BY cs.created_at DESC
               FETCH FIRST 1 ROWS ONLY""",
            (image_id,),
        )
        if not pick_row:
            return None

        return {
            'decision': pick_row['decision'],
            'session_id': pick_row['session_id'],
            'is_best_in_group': pick_row['is_best_in_group'],
            'folder_path': pick_row['folder_path'],
        }
    except Exception as e:
        logging.error(f"Failed to get culling status for {file_path}: {e}")
        return None


def is_folder_keywords_processed(folder_path):
    """
    Checks if a folder is marked as fully processed for keywords.
    """
    folder_path = os.path.normpath(folder_path)
    try:
        row = get_connector().query_one("SELECT is_keywords_processed FROM folders WHERE path = ?", (folder_path,))
        return bool(row and row["is_keywords_processed"])
    except Exception as e:
        logging.error(f"Error checking folder keyword status: {e}")
        return False

def check_and_update_folder_keywords_status(folder_path):
    """
    Checks if all images in a folder have keywords and updates the folder status.
    """
    try:
        folder_path = os.path.normpath(folder_path)

        # 1. Get Folder ID
        row = get_connector().query_one("SELECT id FROM folders WHERE path = ?", (folder_path,))
        if not row:
            return  # Folder not tracked yet? or just insert?
            # If we are strictly checking, maybe we should insert?
            # But usually get_or_create_folder handles insertion.
            # If it's missing, let's assume we can't mark it processed.

        folder_id = row["id"]

        # 2. Check for any images in this folder that have NO keywords
        # We check for NULL or Empty string
        # And we only care about images that are actually registered (e.g. have an ID)

        # We need to be careful: if a folder has NO images, is it processed?
        # Yes, effectively.

        # Check count of *unprocessed* images
        cnt_row = get_connector().query_one("""
            SELECT COUNT(*) AS cnt FROM images
            WHERE folder_id = ?
            AND (keywords IS NULL OR keywords = '')
        """, (folder_id,))

        pending_count = cnt_row["cnt"] if cnt_row else 0

        is_processed = 1 if pending_count == 0 else 0

        # Update status
        get_connector().execute(
            "UPDATE folders SET is_keywords_processed = ? WHERE id = ?",
            (is_processed, folder_id)
        )
        return is_processed == 1

    except Exception as e:
        logging.error(f"Error updating folder keyword status: {e}")
        return False

def get_stack_count_for_folder(folder_path):
    """
    Returns the number of stacks associated with images in a specific folder.
    Used to check if we can reuse existing stacks for culling.
    """
    try:
        # Normalize path
        norm_path = os.path.normpath(folder_path)

        # Get folder_id
        row = get_connector().query_one("SELECT id FROM folders WHERE path = ?", (norm_path,))
        if not row:
            return 0
        folder_id = row["id"]

        # Count stacks for images in this folder
        cnt_row = get_connector().query_one("""
            SELECT COUNT(DISTINCT stack_id) AS cnt
            FROM images
            WHERE folder_id = ? AND stack_id IS NOT NULL
        """, (folder_id,))
        return cnt_row["cnt"] if cnt_row else 0
    except Exception as e:
        print(f"Error counting stacks for folder: {e}")
        return 0

def _embedding_bytes_to_pg(embedding_bytes):
    """Convert raw float32 bytes to a numpy array suitable for pgvector."""
    import numpy as np
    return np.frombuffer(embedding_bytes, dtype=np.float32)


def update_image_embedding(image_id, embedding_bytes):
    """Store a raw float32 embedding blob for an image."""
    try:
        conn = get_connector()
        if conn.type == 'postgres':
            vec = _embedding_bytes_to_pg(embedding_bytes) if embedding_bytes is not None else None
            db_postgres.execute_write(
                "UPDATE images SET image_embedding = %s WHERE id = %s",
                (vec, image_id),
            )
        else:
            conn.execute("UPDATE images SET image_embedding = ? WHERE id = ?", (embedding_bytes, image_id))
    except Exception as e:
        print(f"Error updating embedding for image {image_id}: {e}")


def update_image_embeddings_batch(pairs):
    """
    Batch-update embeddings.  pairs: list of (image_id, embedding_bytes).
    """
    if not pairs:
        return
    try:
        conn = get_connector()
        if conn.type == 'postgres':
            with db_postgres.PGConnectionManager(commit=True) as pg_conn:
                with pg_conn.cursor() as cur:
                    for image_id, embedding_bytes in pairs:
                        vec = _embedding_bytes_to_pg(embedding_bytes) if embedding_bytes is not None else None
                        cur.execute(
                            "UPDATE images SET image_embedding = %s WHERE id = %s",
                            (vec, image_id),
                        )
        else:
            def _tx(tx):
                for image_id, embedding_bytes in pairs:
                    tx.execute("UPDATE images SET image_embedding = ? WHERE id = ?", (embedding_bytes, image_id))
            conn.run_transaction(_tx)
    except Exception as e:
        print(f"Error batch-updating embeddings: {e}")


def _pg_vec_to_bytes(vec) -> "bytes | None":
    """Convert a pgvector result (numpy array) back to raw float32 bytes."""
    if vec is None:
        return None
    import numpy as np
    return np.array(vec, dtype=np.float32).tobytes()


def get_image_embedding(image_id):
    """Return the raw embedding bytes for an image, or None."""
    try:
        conn = get_connector()
        if conn.type == 'postgres':
            row = db_postgres.execute_select_one(
                "SELECT image_embedding FROM images WHERE id = %s", (image_id,)
            )
            return _pg_vec_to_bytes(row["image_embedding"]) if row else None
        row = conn.query_one("SELECT image_embedding FROM images WHERE id = ?", (image_id,))
        if row and row.get("image_embedding"):
            return bytes(row["image_embedding"])
        return None
    except Exception as e:
        print(f"Error getting embedding for image {image_id}: {e}")
        return None

def get_image_embeddings_batch(image_ids: list[int]) -> dict[int, bytes]:
    """Return a dictionary mapping image_id to raw embedding bytes for the given sequence of image IDs."""
    if not image_ids:
        return {}

    try:
        conn = get_connector()
        if conn.type == 'postgres':
            placeholders = ','.join(['%s'] * len(image_ids))
            rows = db_postgres.execute_select(
                f"SELECT id, image_embedding FROM images WHERE id IN ({placeholders})",
                tuple(image_ids),
            )
            return {r["id"]: _pg_vec_to_bytes(r["image_embedding"]) for r in rows if r["image_embedding"] is not None}

        placeholders = ','.join(['?'] * len(image_ids))
        rows = conn.query(
            f"SELECT id, image_embedding FROM images WHERE id IN ({placeholders})",
            tuple(image_ids),
        )
        return {r["id"]: bytes(r["image_embedding"]) for r in rows if r.get("image_embedding") is not None}
    except Exception as e:
        logger.error(f"Error getting batch embeddings: {e}")
        return {}


def get_embeddings_for_search(folder_path=None, limit=None):
    """
    Return (image_id, file_path, embedding_bytes) for images with stored embeddings.
    Optionally filter by folder_path and cap results with limit.
    """
    if _get_db_engine() == "postgres":
        if folder_path:
            norm = os.path.normpath(folder_path)
            sql = ("SELECT i.id, i.file_path, i.image_embedding FROM images i "
                   "JOIN folders f ON f.id = i.folder_id "
                   "WHERE i.image_embedding IS NOT NULL AND f.path = ?")
            params: list = [norm]
        else:
            sql = "SELECT i.id, i.file_path, i.image_embedding FROM images i WHERE i.image_embedding IS NOT NULL"
            params = []
        if limit:
            sql += " ROWS ?"
            params.append(limit)
        conn = get_connector()
        if conn.type == 'postgres':
            # pgvector returns numpy arrays; convert via db_postgres to get proper bytes
            import psycopg2.extras
            pg_sql = _translate_fb_to_pg(sql)
            with db_postgres.PGConnectionManager() as pg_conn:
                with pg_conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
                    cur.execute(pg_sql, tuple(params) if params else None)
                    return [(r["id"], r["file_path"], bytes(r["image_embedding"])) for r in cur.fetchall()]
        rows = conn.query(sql, tuple(params) if params else None)
        return [(r["id"], r["file_path"], bytes(r["image_embedding"])) for r in rows]


def get_embeddings_with_metadata(folder_path=None, limit=None):
    """
    Return embedding vectors together with display metadata for each image.

    Each returned dict has keys:
        image_id, file_path, embedding (bytes), thumbnail_path,
        label, rating, score_general
    Optionally filter by folder_path and cap results with limit.
    """
    if _get_db_engine() == "postgres":
        if folder_path:
            norm = os.path.normpath(folder_path)
            sql = ("SELECT i.id, i.file_path, i.image_embedding, i.thumbnail_path, "
                   "       i.label, i.rating, i.score_general FROM images i "
                   "JOIN folders f ON f.id = i.folder_id "
                   "WHERE i.image_embedding IS NOT NULL AND f.path = ?")
            params: list = [norm]
        else:
            sql = ("SELECT i.id, i.file_path, i.image_embedding, i.thumbnail_path, "
                   "       i.label, i.rating, i.score_general FROM images i "
                   "WHERE i.image_embedding IS NOT NULL")
            params = []
        if limit:
            sql += " ROWS ?"
            params.append(limit)
        conn = get_connector()
        if conn.type == 'postgres':
            import psycopg2.extras
            pg_sql = _translate_fb_to_pg(sql)
            with db_postgres.PGConnectionManager() as pg_conn:
                with pg_conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
                    cur.execute(pg_sql, tuple(params) if params else None)
                    return [{
                        "image_id": r["id"],
                        "file_path": r["file_path"],
                        "embedding": bytes(r["image_embedding"]),
                        "thumbnail_path": r["thumbnail_path"],
                        "label": r["label"],
                        "rating": r["rating"],
                        "score_general": float(r["score_general"]) if r["score_general"] is not None else None,
                    } for r in cur.fetchall()]
        rows = conn.query(sql, tuple(params) if params else None)
        return [{
            "image_id": r["id"],
            "file_path": r["file_path"],
            "embedding": bytes(r["image_embedding"]),
            "thumbnail_path": r["thumbnail_path"],
            "label": r["label"],
            "rating": r["rating"],
            "score_general": float(r["score_general"]) if r["score_general"] is not None else None,
        } for r in rows]


def get_images_missing_embeddings(folder_path=None, limit=None, min_id_exclusive=None):
    """
    Return image rows with image_embedding IS NULL.
    Columns: id, file_path, thumbnail_path, thumbnail_path_win (for path resolution in WSL).
    Optionally filter by folder_path, cap with limit, and resume with min_id_exclusive (id > value).
    Rows are ordered by id ascending for stable checkpointing.
    """
    try:
        id_resume = ""
        resume_params: list = []
        if min_id_exclusive is not None:
            id_resume = " AND id > ?"
            resume_params.append(int(min_id_exclusive))

        if folder_path:
            norm = os.path.normpath(folder_path)
            frow = get_connector().query_one("SELECT id FROM folders WHERE path = ?", (norm,))
            if not frow:
                return []
            folder_id = frow["id"]
            query = f"""
                SELECT id, file_path, thumbnail_path, thumbnail_path_win
                FROM images
                WHERE image_embedding IS NULL AND folder_id = ?{id_resume}
                ORDER BY id
            """
            params = [folder_id] + resume_params
        else:
            query = f"""
                SELECT id, file_path, thumbnail_path, thumbnail_path_win
                FROM images
                WHERE image_embedding IS NULL{id_resume}
                ORDER BY id
            """
            params = list(resume_params)

        if limit:
            query += " FETCH FIRST ? ROWS ONLY"
            params.append(limit)

        return list(get_connector().query(query, tuple(params)))
    except Exception as e:
        logger.error("Error loading images missing embeddings: %s", e)
        return []


def get_images_for_tag_propagation(folder_path=None):
    """
    Return two lists for tag propagation:
      untagged: [(image_id, file_path, embedding_bytes)] — images with embeddings but no keywords
      tagged:   [(image_id, file_path, embedding_bytes, keywords_str)] — images with embeddings AND keywords

    Optional folder_path narrows scope to a single folder.
    """
    try:
        conn = get_connector()
        folder_filter = ""
        params: list = []
        if folder_path:
            norm = os.path.normpath(folder_path)
            frow = conn.query_one("SELECT id FROM folders WHERE path = ?", (norm,))
            if not frow:
                return [], []
            folder_filter = " AND i.folder_id = ?"
            params.append(frow["id"])

        # Untagged images with embeddings.
        # Phase 2/3 migration: prefer normalized IMAGE_KEYWORDS; keep legacy
        # IMAGES.KEYWORDS fallback for rows not yet dual-written.
        q_untagged = (
            "SELECT i.id, i.file_path, i.image_embedding FROM images i "
            "WHERE i.image_embedding IS NOT NULL "
            "AND NOT EXISTS (SELECT 1 FROM image_keywords ik WHERE ik.image_id = i.id) "
            "AND (i.keywords IS NULL OR i.keywords = '')" + folder_filter
        )

        # Tagged images with embeddings. Build keyword CSV from normalized
        # tables when possible, otherwise fall back to legacy IMAGES.KEYWORDS.
        q_tagged = (
            "SELECT i.id, i.file_path, i.image_embedding, "
            "COALESCE((SELECT LIST(COALESCE(kd.keyword_display, kd.keyword_norm), ', ') "
            "FROM image_keywords ik JOIN keywords_dim kd ON kd.keyword_id = ik.keyword_id "
            "WHERE ik.image_id = i.id), i.keywords) AS keywords_csv "
            "FROM images i "
            "WHERE i.image_embedding IS NOT NULL "
            "AND (EXISTS (SELECT 1 FROM image_keywords ik WHERE ik.image_id = i.id) "
            "OR (i.keywords IS NOT NULL AND i.keywords != ''))" + folder_filter
        )

        if conn.type == 'postgres':
            # pgvector requires direct cursor for proper binary embedding access
            import psycopg2.extras
            pg_untagged = _translate_fb_to_pg(q_untagged)
            pg_tagged = _translate_fb_to_pg(q_tagged)
            pg_params = tuple(params) if params else None
            with db_postgres.PGConnectionManager() as pg_conn:
                with pg_conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
                    cur.execute(pg_untagged, pg_params)
                    untagged = [(r["id"], r["file_path"], bytes(r["image_embedding"])) for r in cur.fetchall()]
                    cur.execute(pg_tagged, pg_params)
                    tagged = [(r["id"], r["file_path"], bytes(r["image_embedding"]), r["keywords_csv"]) for r in cur.fetchall()]
            return untagged, tagged

        # Firebird / other connector path
        p = tuple(params) if params else None
        untagged_rows = conn.query(q_untagged, p)
        untagged = [(r["id"], r["file_path"], bytes(r["image_embedding"])) for r in untagged_rows]
        tagged_rows = conn.query(q_tagged, p)
        tagged = [(r["id"], r["file_path"], bytes(r["image_embedding"]), r["keywords_csv"]) for r in tagged_rows]
        return untagged, tagged
    except Exception as e:
        logging.error("Error loading images for tag propagation: %s", e)
        return [], []



def get_image_tag_propagation_focus(image_id: int):
    """
    Embedding, paths, and keyword CSV for one image (for dry-run preview on a
    specific row, including already-tagged images).

    Returns:
        (embedding_bytes, file_path, folder_path_or_none, keywords_csv) or None
        if the image has no embedding.
    """
    conn = get_db()
    c = conn.cursor()
    try:
        c.execute(
            "SELECT i.image_embedding, i.file_path, f.path, "
            "COALESCE((SELECT LIST(COALESCE(kd.keyword_display, kd.keyword_norm), ', ') "
            "FROM image_keywords ik JOIN keywords_dim kd ON kd.keyword_id = ik.keyword_id "
            "WHERE ik.image_id = i.id), COALESCE(i.keywords, '')) AS kw "
            "FROM images i LEFT JOIN folders f ON f.id = i.folder_id WHERE i.id = ?",
            (image_id,),
        )
        row = c.fetchone()
        if not row or not row[0]:
            return None
        emb = bytes(row[0])
        fp = row[1]
        folder_path = row[2]
        kw = row[3] if row[3] is not None else ""
        return (emb, fp, folder_path, kw)
    except Exception as e:
        logging.error("get_image_tag_propagation_focus(%s): %s", image_id, e)
        return None
    finally:
        conn.close()


def list_folder_paths_with_missing_keywords(require_embedding: bool = False):
    """
    Folders (direct image parent) with at least one image lacking keywords.

    Matches the untagged predicate in get_images_for_tag_propagation: no row in
    image_keywords and legacy images.keywords is null or empty. When
    require_embedding is True, also requires image_embedding IS NOT NULL
    (same as propagation's untagged set).

    Returns:
        List of (folder_path, untagged_image_count) sorted by count desc, then path.
    """
    embed_clause = ""
    if require_embedding:
        embed_clause = " AND i.image_embedding IS NOT NULL "
    query = (
        "SELECT f.path, COUNT(*) AS cnt "
        "FROM images i "
        "JOIN folders f ON f.id = i.folder_id "
        "WHERE NOT EXISTS (SELECT 1 FROM image_keywords ik WHERE ik.image_id = i.id) "
        "AND (i.keywords IS NULL OR i.keywords = '') "
        f"{embed_clause}"
        "GROUP BY f.path "
        "HAVING COUNT(*) > 0 "
        "ORDER BY COUNT(*) DESC, f.path"
    )
    try:
        result = []
        for row in get_connector().query(query):
            path = row.get("path")
            cnt = int(row.get("cnt") or 0)
            if path:
                result.append((str(path), cnt))
        return result
    except Exception as e:
        logging.error("list_folder_paths_with_missing_keywords: %s", e)
        return []


# ===========================================================================
# Pipeline Phases — helper functions
# ===========================================================================

# Module-level cache: phase_code -> phase_id
_phase_id_cache = {}


def seed_pipeline_phases():
    """
    Insert default pipeline phases into PIPELINE_PHASES table.
    Idempotent — skips existing codes.
    """
    from modules.phases import SEED_PHASES

    try:
        def _tx(tx):
            # Fix any broken rows accidentally inserted as "PhaseCode.INDEXING"
            tx.execute("UPDATE pipeline_phases SET code = REPLACE(code, 'PhaseCode.', '') WHERE code LIKE 'PhaseCode.%'")

            for phase in SEED_PHASES:
                code = phase["code"].value if hasattr(phase["code"], "value") else str(phase["code"])
                existing = tx.query_one("SELECT id FROM pipeline_phases WHERE code = ?", (code,))
                if existing is None:
                    tx.execute(
                        "INSERT INTO pipeline_phases (code, name, description, sort_order, enabled, optional, default_skip) "
                        "VALUES (?, ?, ?, ?, 1, ?, ?)",
                        (code, phase["name"], phase.get("description", ""), phase["sort_order"],
                         1 if phase.get("optional") else 0, 1 if phase.get("default_skip") else 0))
                else:
                    tx.execute(
                        "UPDATE pipeline_phases SET optional = ?, default_skip = ? WHERE code = ?",
                        (1 if phase.get("optional") else 0, 1 if phase.get("default_skip") else 0, code))

        get_connector().run_transaction(_tx)
        logger.info("Pipeline phases seeded successfully.")
    except Exception as e:
        logger.error("Failed to seed pipeline phases: %s", e)
    # Clear cache so it's rebuilt on next access
    _phase_id_cache.clear()


def get_phase_id(phase_code):
    """
    Look up pipeline_phases.id by code.  Result is cached per-process.

    Args:
        phase_code: str or PhaseCode enum value.

    Returns:
        int or None if not found.
    """
    code = phase_code.value if hasattr(phase_code, "value") else str(phase_code)
    
    if code in _phase_id_cache:
        return _phase_id_cache[code]

    row = get_connector().query_one("SELECT id FROM pipeline_phases WHERE code = ?", (code,))
    if row:
        _phase_id_cache[code] = row["id"]
        return row["id"]
    return None


def set_image_phase_status(image_id, phase_code, status,
                            app_version=None, executor_version=None,
                            job_id=None, error=None, skip_reason=None, skipped_by=None):
    """
    Upsert a row in IMAGE_PHASE_STATUS for (image_id, phase).

    Increments attempt_count on reruns (done/failed/skipped → running).
    Sets started_at on 'running', finished_at on terminal states.
    """
    from modules.phases import PhaseStatus

    phase_id = get_phase_id(phase_code)
    if phase_id is None:
        logger.warning("set_image_phase_status: unknown phase '%s'", phase_code)
        return

    now = datetime.datetime.now()

    def _tx(tx):
        # Check existing row
        existing = tx.query_one(
            "SELECT id, status, attempt_count FROM image_phase_status "
            "WHERE image_id = ? AND phase_id = ?",
            (image_id, phase_id)
        )

        if existing:
            row_id = existing["id"]
            old_status = (existing["status"] or "not_started").strip()
            attempt_count = existing["attempt_count"] or 0

            # Guard: running → running is not allowed (duplicate job protection)
            if old_status == PhaseStatus.RUNNING and status == PhaseStatus.RUNNING:
                logger.warning(
                    "set_image_phase_status: running→running guard triggered "
                    "(img=%s, phase=%s) — skipping duplicate update", image_id, phase_code
                )
                return None

            # Increment attempt on rerun transitions
            if status == PhaseStatus.RUNNING and old_status in (
                PhaseStatus.DONE, PhaseStatus.FAILED, PhaseStatus.SKIPPED
            ):
                attempt_count += 1

            # Build UPDATE
            fields = ["status = ?", "updated_at = ?", "attempt_count = ?"]
            params = [status, now, attempt_count]

            if status == PhaseStatus.RUNNING:
                fields.append("started_at = ?")
                params.append(now)
            elif status in (PhaseStatus.DONE, PhaseStatus.FAILED, PhaseStatus.SKIPPED):
                fields.append("finished_at = ?")
                params.append(now)

            if app_version is not None:
                fields.append("app_version = ?")
                params.append(app_version)
            if executor_version is not None:
                fields.append("executor_version = ?")
                params.append(executor_version)
            if job_id is not None:
                fields.append("job_id = ?")
                params.append(job_id)
            if error is not None:
                fields.append("last_error = ?")
                params.append(error)
            elif status == PhaseStatus.DONE:
                fields.append("last_error = NULL")

            if status == PhaseStatus.SKIPPED:
                fields.append("skip_reason = ?")
                params.append(skip_reason)
                fields.append("skipped_by = ?")
                params.append(skipped_by)
            elif status == PhaseStatus.RUNNING:
                fields.append("skip_reason = NULL")
                fields.append("skipped_by = NULL")

            params.append(row_id)
            tx.execute(f"UPDATE image_phase_status SET {', '.join(fields)} WHERE id = ?", params)
        else:
            # INSERT new row
            started = now if status == PhaseStatus.RUNNING else None
            finished = now if status in (PhaseStatus.DONE, PhaseStatus.FAILED, PhaseStatus.SKIPPED) else None

            tx.execute(
                "INSERT INTO image_phase_status "
                "(image_id, phase_id, status, app_version, executor_version, "
                " job_id, attempt_count, last_error, started_at, finished_at, updated_at, skip_reason, skipped_by) "
                "VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)",
                (image_id, phase_id, status, app_version, executor_version,
                 job_id, 0, error, started, finished, now,
                 skip_reason if status == PhaseStatus.SKIPPED else None,
                 skipped_by if status == PhaseStatus.SKIPPED else None)
            )

        frow = tx.query_one("SELECT folder_id FROM images WHERE id = ?", (image_id,))
        return frow["folder_id"] if frow else None

    try:
        folder_id = get_connector().run_transaction(_tx)
    except Exception as e:
        logger.error("set_image_phase_status failed (img=%s, phase=%s): %s", image_id, phase_code, e)
        folder_id = None

    phase_text = phase_code.value if hasattr(phase_code, "value") else str(phase_code)
    event_type = "progress" if status == PhaseStatus.RUNNING else "state-change"
    severity = "error" if status == PhaseStatus.FAILED else ("warning" if status == PhaseStatus.SKIPPED else "info")
    record_pipeline_event(
        "error" if status == PhaseStatus.FAILED else event_type,
        f"Image #{image_id} phase {phase_text}: {status}",
        workflow_run=job_id,
        stage_run=phase_text,
        step_run=f"image:{image_id}",
        category="phase",
        severity=severity,
        metadata={"image_id": image_id, "phase": phase_text, "status": status, "error": error, "skip_reason": skip_reason},
        critical=status == PhaseStatus.FAILED,
        noisy=True,
        source="db.set_image_phase_status",
    )

    if folder_id:
        invalidate_folder_phase_aggregates(folder_id=folder_id)


def get_image_phase_statuses(image_id):
    """
    Return phase statuses for one image.

    Returns:
        dict: {phase_code: {status, executor_version, app_version, updated_at, attempt_count, last_error}}
    """
    rows = get_connector().query(
        "SELECT pp.code, ips.status, ips.executor_version, ips.app_version, "
        "       ips.updated_at, ips.attempt_count, ips.last_error, ips.skip_reason, ips.skipped_by "
        "FROM image_phase_status ips "
        "JOIN pipeline_phases pp ON pp.id = ips.phase_id "
        "WHERE ips.image_id = ? ORDER BY pp.sort_order",
        (image_id,)
    )
    result = {}
    for r in rows:
        code = (r["code"] or "").strip()
        result[code] = {
            "status": (r["status"] or "not_started").strip(),
            "executor_version": r["executor_version"],
            "app_version": r["app_version"],
            "updated_at": r["updated_at"],
            "attempt_count": r["attempt_count"],
            "last_error": r["last_error"],
            "skip_reason": r["skip_reason"],
            "skipped_by": r["skipped_by"],
        }
    return result


def get_all_phases(enabled_only=True):
    """
    Return all phase definitions from PIPELINE_PHASES table.

    Returns:
        list[dict]: [{id, code, name, description, sort_order, enabled}, ...]
    """
    query = "SELECT id, code, name, description, sort_order, enabled, optional, default_skip FROM pipeline_phases"
    if enabled_only:
        query += " WHERE enabled = 1"
    query += " ORDER BY sort_order"
    return [{
        "id": r["id"],
        "code": (r["code"] or "").strip(),
        "name": (r["name"] or "").strip(),
        "description": r["description"],
        "sort_order": r["sort_order"],
        "enabled": r["enabled"],
        "optional": bool(r["optional"]),
        "default_skip": bool(r["default_skip"]),
    } for r in get_connector().query(query)]


def _derive_folder_phase_status(total, done, running, failed, skipped):
    """Derive normalized folder status from per-image phase counts."""
    if total <= 0:
        return "not_started"
    if done == total:
        return "done"
    if running > 0:
        return "running"
    if failed > 0 and done == 0 and skipped == 0:
        return "failed"
    if skipped == total:
        return "skipped"
    if done > 0 or failed > 0 or skipped > 0:
        return "partial"
    return "not_started"


def get_folder_phase_summary(folder_path, force_refresh=False):
    """
    Return phase status summary for a folder and descendants.

    Uses folder-level cache (`folders.phase_agg_json`) and recomputes live data
    when `phase_agg_dirty = 1`. Pass force_refresh=True to bypass cache and
    always recompute (e.g. when user selects a folder or clicks Refresh).
    """
    from modules import utils

    wsl_path = utils.convert_path_to_wsl(folder_path) if hasattr(utils, 'convert_path_to_wsl') else folder_path
    target_path = wsl_path if wsl_path else folder_path
    base_folder_id = get_or_create_folder(target_path)
    if not base_folder_id:
        return []

    try:
        if not force_refresh:
            cache_row = get_connector().query_one(
                "SELECT phase_agg_dirty, phase_agg_json FROM folders WHERE id = ?",
                (base_folder_id,))
            if cache_row and (cache_row.get("phase_agg_dirty") or 0) == 0 and cache_row.get("phase_agg_json"):
                try:
                    return json.loads(cache_row["phase_agg_json"])
                except Exception:
                    pass

        path_like_unix = target_path + "/%"
        path_like_win = target_path + "\\%"

        rows = get_connector().query(
            """
            SELECT
                pp.code,
                pp.name,
                pp.sort_order,
                COUNT(i.id) as total_images,
                COALESCE(SUM(CASE WHEN ips.status = 'done' THEN 1 ELSE 0 END), 0) as done_count,
                COALESCE(SUM(CASE WHEN ips.status = 'failed' THEN 1 ELSE 0 END), 0) as failed_count,
                COALESCE(SUM(CASE WHEN ips.status = 'running' THEN 1 ELSE 0 END), 0) as running_count,
                COALESCE(SUM(CASE WHEN ips.status = 'queued' THEN 1 ELSE 0 END), 0) as queued_count,
                COALESCE(SUM(CASE WHEN ips.status = 'paused' THEN 1 ELSE 0 END), 0) as paused_count,
                COALESCE(SUM(CASE WHEN ips.status = 'cancel_requested' THEN 1 ELSE 0 END), 0) as cancel_requested_count,
                COALESCE(SUM(CASE WHEN ips.status = 'restarting' THEN 1 ELSE 0 END), 0) as restarting_count,
                COALESCE(SUM(CASE WHEN ips.status = 'skipped' THEN 1 ELSE 0 END), 0) as skipped_count,
                pp.optional
            FROM pipeline_phases pp
            CROSS JOIN (
                SELECT id FROM images
                WHERE folder_id IN (
                    SELECT id FROM folders
                    WHERE path = ? OR path LIKE ? OR path LIKE ?
                )
            ) i
            LEFT JOIN image_phase_status ips
                ON ips.image_id = i.id AND ips.phase_id = pp.id
            WHERE pp.enabled = 1
            GROUP BY pp.code, pp.name, pp.sort_order, pp.optional
            ORDER BY pp.sort_order
            """,
            (target_path, path_like_unix, path_like_win))

        result = []
        scoring_done = False
        for row in rows:
            code = row["code"].strip() if isinstance(row["code"], str) else row["code"]
            name = row["name"].strip() if isinstance(row["name"], str) else row["name"]
            sort_order = row["sort_order"]
            total = row["total_images"] or 0
            done = row["done_count"] or 0
            failed = row["failed_count"] or 0
            running = row["running_count"] or 0
            queued = row["queued_count"] or 0
            paused = row["paused_count"] or 0
            cancel_requested = row["cancel_requested_count"] or 0
            restarting = row["restarting_count"] or 0
            skipped = row["skipped_count"] or 0
            is_optional = bool(row["optional"])

            advance_ready = done + skipped if is_optional else done
            if total == 0:
                status = "not_started"
            elif done == total:
                status = "done"
            elif skipped == total and is_optional:
                status = "skipped"
            elif advance_ready == total and is_optional:
                status = "done"
            elif running > 0:
                status = "running"
            elif paused > 0:
                status = "paused"
            elif queued > 0:
                status = "queued"
            elif restarting > 0:
                status = "restarting"
            elif cancel_requested > 0:
                status = "cancel_requested"
            elif done > 0 or skipped > 0:
                status = "partial"
            elif failed > 0:
                status = "failed"
            else:
                status = "not_started"

            if code == 'scoring' and status == 'done':
                scoring_done = True

            result.append({
                "code": code,
                "name": name,
                "sort_order": sort_order,
                "status": status,
                "done_count": done,
                "failed_count": failed,
                "running_count": running,
                "queued_count": queued,
                "paused_count": paused,
                "cancel_requested_count": cancel_requested,
                "restarting_count": restarting,
                "skipped_count": skipped,
                "total_count": total,
                "optional": is_optional,
                "advance_ready": advance_ready == total if total > 0 else False,
            })

        # Cache the computed result (non-fatal if this fails)
        try:
            get_connector().execute(
                "UPDATE folders SET phase_agg_dirty = 0, phase_agg_updated_at = ?, phase_agg_json = ?, is_fully_scored = ? WHERE id = ?",
                (datetime.datetime.now(), json.dumps(result), 1 if scoring_done else 0, base_folder_id))
        except Exception as cache_err:
            logger.debug("get_folder_phase_summary cache write failed for '%s' (non-fatal): %s", folder_path, cache_err)
        return result
    except Exception as e:
        logger.error("get_folder_phase_summary failed for '%s': %s", folder_path, e)
        return []


def get_all_folder_phase_summaries_bulk():
    """Single-query bulk load of phase summary cache for all folders.

    Returns dict mapping normalized local path -> list of phase summary dicts.
    Includes all folders that have a cached phase_agg_json (clean or dirty).
    Dirty/stale data is acceptable here — callers use this for sidebar display
    where slightly stale is far better than N individual expensive queries.
    """
    from modules import utils
    rows = get_connector().query("SELECT path, phase_agg_json FROM folders WHERE phase_agg_json IS NOT NULL")
    result = {}
    for row in rows:
        wsl_path = row["path"]
        local_path = utils.convert_path_to_local(wsl_path) if hasattr(utils, 'convert_path_to_local') else wsl_path
        if not local_path:
            continue
        try:
            result[os.path.normpath(local_path)] = json.loads(row["phase_agg_json"])
        except Exception:
            pass
    return result




def set_folder_phase_status(folder_path, phase_code, status, reason=None, actor=None, app_version=None, executor_version=None, job_id=None):
    """
    Bulk update image_phase_status for all images in a folder/subfolders.

    Args:
        folder_path: Target folder path.
        phase_code:  Phase code or enum.
        status:      PhaseStatus value to apply to each image.
        reason:      Optional skip reason (for skipped status).
        actor:       Optional user/actor label (for skipped status).
        app_version: Optional app version stamp.
        executor_version: Optional executor version stamp.
        job_id: Optional job id associated with this transition.

    Returns:
        int: number of image rows updated.
    """
    from modules import utils

    if not folder_path:
        return 0

    wsl_path = utils.convert_path_to_wsl(folder_path) if hasattr(utils, 'convert_path_to_wsl') else folder_path
    target_path = wsl_path if wsl_path else folder_path
    path_like_unix = target_path + "/%"
    path_like_win = target_path + "\\%"

    rows = get_connector().query(
        """
        SELECT i.id
        FROM images i
        JOIN folders f ON f.id = i.folder_id
        WHERE f.path = ? OR f.path LIKE ? OR f.path LIKE ?
        """,
        (target_path, path_like_unix, path_like_win))
    image_ids = [row["id"] for row in rows]

    for image_id in image_ids:
        set_image_phase_status(
            image_id=image_id,
            phase_code=phase_code,
            status=status,
            app_version=app_version,
            executor_version=executor_version,
            job_id=job_id,
            skip_reason=reason,
            skipped_by=actor,
        )

    return len(image_ids)


def backfill_index_meta_for_folder(folder_path):
    """
    Set INDEXING=DONE and METADATA=DONE for images that have SCORING=DONE
    but lack INDEXING or METADATA status (backfill gap from legacy runs or new-image flow).

    Args:
        folder_path: Target folder path (includes subfolders).

    Returns:
        int: number of images updated.
    """
    from modules import utils

    if not folder_path:
        return 0

    wsl_path = utils.convert_path_to_wsl(folder_path) if hasattr(utils, 'convert_path_to_wsl') else folder_path
    target_path = wsl_path if wsl_path else folder_path
    path_like_unix = target_path + "/%"
    path_like_win = target_path + "\\%"

    rows = get_connector().query(
        """
        SELECT i.id
        FROM images i
        JOIN folders f ON f.id = i.folder_id
        WHERE f.path = ? OR f.path LIKE ? OR f.path LIKE ?
        AND EXISTS (
            SELECT 1 FROM image_phase_status ips
            JOIN pipeline_phases pp ON pp.id = ips.phase_id
            WHERE ips.image_id = i.id AND pp.code = 'scoring' AND ips.status = 'done'
        )
        AND (
            NOT EXISTS (
                SELECT 1 FROM image_phase_status ips2
                JOIN pipeline_phases pp2 ON pp2.id = ips2.phase_id
                WHERE ips2.image_id = i.id AND pp2.code = 'indexing' AND ips2.status = 'done'
            )
            OR NOT EXISTS (
                SELECT 1 FROM image_phase_status ips3
                JOIN pipeline_phases pp3 ON pp3.id = ips3.phase_id
                WHERE ips3.image_id = i.id AND pp3.code = 'metadata' AND ips3.status = 'done'
            )
        )
        """,
        (target_path, path_like_unix, path_like_win))
    image_ids = [row["id"] for row in rows]

    for image_id in image_ids:
        set_image_phase_status(image_id, "indexing", "done")
        set_image_phase_status(image_id, "metadata", "done")

    if image_ids:
        invalidate_folder_phase_aggregates(folder_path=target_path)

    return len(image_ids)


def backfill_index_meta_global(limit=None, dry_run=False):
    """
    Same semantics as backfill_index_meta_for_folder, but for all images in the DB.

    Sets INDEXING=DONE and METADATA=DONE when SCORING is DONE but either earlier
    phase is missing or not done (repairs GAP-D / WORKFLOW_STAGES_ANALYSIS gaps).

    Args:
        limit: Optional max number of images to update (for staged rollouts).
        dry_run: If True, only return the count of images that would be updated.

    Returns:
        int: number of images that would be or were updated.
    """
    rows = get_connector().query(
        """
        SELECT i.id
        FROM images i
        WHERE EXISTS (
            SELECT 1 FROM image_phase_status ips
            JOIN pipeline_phases pp ON pp.id = ips.phase_id
            WHERE ips.image_id = i.id
              AND LOWER(TRIM(pp.code)) = 'scoring'
              AND LOWER(TRIM(ips.status)) = 'done'
        )
        AND (
            NOT EXISTS (
                SELECT 1 FROM image_phase_status ips2
                JOIN pipeline_phases pp2 ON pp2.id = ips2.phase_id
                WHERE ips2.image_id = i.id
                  AND LOWER(TRIM(pp2.code)) = 'indexing'
                  AND LOWER(TRIM(ips2.status)) = 'done'
            )
            OR NOT EXISTS (
                SELECT 1 FROM image_phase_status ips3
                JOIN pipeline_phases pp3 ON pp3.id = ips3.phase_id
                WHERE ips3.image_id = i.id
                  AND LOWER(TRIM(pp3.code)) = 'metadata'
                  AND LOWER(TRIM(ips3.status)) = 'done'
            )
        )
        """)
    image_ids = [row["id"] for row in rows]

    if isinstance(limit, int) and limit > 0:
        image_ids = image_ids[:limit]

    if dry_run:
        return len(image_ids)

    for image_id in image_ids:
        set_image_phase_status(image_id, "indexing", "done")
        set_image_phase_status(image_id, "metadata", "done")

    if image_ids:
        get_connector().execute("UPDATE folders SET phase_agg_dirty = 1")

    return len(image_ids)


def repair_stuck_running_ips(hours=2, phase_code=None, dry_run=False):
    """
    Mark IMAGE_PHASE_STATUS rows stuck in running/queued (older than ``hours``)
    as failed so the UI and policy layer can retry.

    Uses a direct UPDATE to avoid invalid FSM transitions (e.g. queued→failed).

    Args:
        hours: Minimum age of updated_at or started_at before repair.
        phase_code: If set, only rows for this pipeline_phases.code (e.g. 'culling').
        dry_run: If True, only count matching rows.

    Returns:
        dict with keys matched, updated.
    """
    import datetime as _dt

    threshold = _dt.datetime.now() - _dt.timedelta(hours=hours)
    msg = (
        "repair_stuck_running_ips: stuck running/queued beyond threshold; "
        "likely lost job_id in worker queue. Re-run this phase if needed."
    )

    sql = """
        SELECT ips.id
        FROM image_phase_status ips
        JOIN pipeline_phases pp ON pp.id = ips.phase_id
        WHERE LOWER(TRIM(ips.status)) IN ('running', 'queued')
          AND (
              (ips.updated_at IS NOT NULL AND ips.updated_at < ?)
              OR (ips.updated_at IS NULL AND ips.started_at IS NOT NULL AND ips.started_at < ?)
          )
    """
    params = [threshold, threshold]
    if phase_code:
        sql += " AND LOWER(TRIM(pp.code)) = ?"
        params.append(str(phase_code).strip().lower())

    rows = get_connector().query(sql, tuple(params))
    ids = [row["id"] for row in rows]

    if dry_run:
        return {"matched": len(ids), "updated": 0}

    def _tx(tx):
        now = _dt.datetime.now()
        for ips_id in ids:
            tx.execute(
                """
                UPDATE image_phase_status
                SET status = 'failed',
                    last_error = ?,
                    finished_at = ?,
                    updated_at = ?,
                    job_id = NULL
                WHERE id = ?
                """,
                (msg, now, now, ips_id))
        if ids:
            tx.execute("UPDATE folders SET phase_agg_dirty = 1")

    get_connector().run_transaction(_tx)
    return {"matched": len(ids), "updated": len(ids)}


def repair_legacy_keywords_junction(limit=None, dry_run=False):
    """
    Dual-write repair: images.keywords is non-empty but image_keywords has no rows.

    Calls _sync_image_keywords for each matching image (GAP-E / bird keyword queries).

    Args:
        limit: Optional max images to process.
        dry_run: If True, return counts only.

    Returns:
        dict: matched, synced
    """
    rows = get_connector().query(
        """
        SELECT i.id, i.keywords
        FROM images i
        WHERE i.keywords IS NOT NULL
        AND NOT EXISTS (
            SELECT 1 FROM image_keywords ik WHERE ik.image_id = i.id
        )
        """)
    pairs = [(row["id"], str(row["keywords"]) if row.get("keywords") is not None else None) for row in rows]

    if isinstance(limit, int) and limit > 0:
        pairs = pairs[:limit]

    if dry_run:
        return {"matched": len(pairs), "synced": 0}

    synced = 0
    nulled = 0
    for image_id, keywords_str in pairs:
        if not keywords_str or not keywords_str.strip():
            try:
                get_connector().execute("UPDATE images SET keywords = NULL WHERE id = ?", (image_id,))
                nulled += 1
            except Exception as e:
                logger.warning("repair_legacy_keywords_junction: null-cleanup image %s: %s", image_id, e)
        else:
            try:
                _sync_image_keywords(image_id, keywords_str, source="repair_legacy_keywords_junction")
                synced += 1
            except Exception as e:
                logger.warning("repair_legacy_keywords_junction: image %s: %s", image_id, e)

    return {"matched": len(pairs), "synced": synced, "nulled": nulled}


def repair_culling_ips_failed_has_data(dry_run=False):
    """
    Repair GAP-C: culling IPS is 'failed' but the image has a stack_id / embedding.
    This happens when a job was force-failed by the repair script but the work
    actually completed. Since data is present, the accurate state is 'done'.

    Returns:
        dict: matched, repaired
    """
    rows = get_connector().query(
        """
        SELECT ips.id
        FROM image_phase_status ips
        JOIN pipeline_phases pp ON pp.id = ips.phase_id
        JOIN images i ON i.id = ips.image_id
        WHERE LOWER(TRIM(pp.code)) = 'culling'
          AND LOWER(TRIM(ips.status)) = 'failed'
          AND (i.stack_id IS NOT NULL OR i.image_embedding IS NOT NULL)
        """)
    ids = [row["id"] for row in rows]

    if dry_run:
        return {"matched": len(ids), "repaired": 0}

    def _tx(tx):
        now = datetime.datetime.now()
        for ips_id in ids:
            tx.execute(
                """
                UPDATE image_phase_status
                SET status = 'done',
                    last_error = NULL,
                    finished_at = COALESCE(finished_at, ?),
                    updated_at = ?
                WHERE id = ?
                """,
                (now, now, ips_id))
        if ids:
            tx.execute("UPDATE folders SET phase_agg_dirty = 1")

    get_connector().run_transaction(_tx)
    return {"matched": len(ids), "repaired": len(ids)}


def backfill_keywords_ips_done(dry_run=False):
    """
    Repair GAP-B: image has image_keywords rows but no 'keywords' IPS record.
    Often happens when images are tagged via legacy batch-update paths.

    Returns:
        dict: matched, created
    """
    rows = get_connector().query(
        """
        SELECT DISTINCT ik.image_id
        FROM image_keywords ik
        WHERE NOT EXISTS (
            SELECT 1 FROM image_phase_status ips
            JOIN pipeline_phases pp ON pp.id = ips.phase_id
            WHERE ips.image_id = ik.image_id
              AND LOWER(TRIM(pp.code)) = 'keywords'
        )
        """)
    ids = [row["image_id"] for row in rows]

    if dry_run:
        return {"matched": len(ids), "created": 0}

    count = 0
    for image_id in ids:
        set_image_phase_status(image_id, "keywords", "done")
        count += 1

    if ids:
        get_connector().execute("UPDATE folders SET phase_agg_dirty = 1")

    return {"matched": len(ids), "created": count}


def _is_firebird_running(host_ip, port=3050):
    import socket
    try:
        with socket.socket(socket.AF_INET, socket.SOCK_STREAM) as s:
            s.settimeout(1)
            result = s.connect_ex((host_ip, port))
            return result == 0
    except (OSError, Exception):
        return False

def _launch_firebird_server_wsl(fb_exe_path):
    import subprocess
    try:
        # Use cmd.exe to launch via Windows
        # start /B runs in background (same window), but -a might pop a window.
        # We try to minimize intrusion.
        cmd = f'cmd.exe /c start /B "" "{fb_exe_path}" -a'
        subprocess.Popen(cmd, shell=True, stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
    except Exception as e:
        print(f"Failed to launch Firebird Server: {e}")


def _sync_image_keywords(image_id, keywords_str, source="auto", confidence=1.0):
    """
    Dual-write sync: Parses the legacy keywords CSV string and updates the normalized
    IMAGE_KEYWORDS and KEYWORDS_DIM tables.
    """
    if not image_id:
        return

    try:
        def _tx(tx):
            tx.execute("DELETE FROM image_keywords WHERE image_id = ?", (image_id,))

            if not keywords_str:
                return

            kws = [k.strip() for k in keywords_str.split(',') if k.strip()]
            if not kws:
                return

            for kw in kws:
                kw_norm = kw.lower()
                row = tx.query_one("SELECT keyword_id FROM keywords_dim WHERE keyword_norm = ?", (kw_norm,))
                if row:
                    kw_id = row["keyword_id"]
                else:
                    ins = tx.execute_returning(
                        "INSERT INTO keywords_dim (keyword_norm, keyword_display) VALUES (?, ?) RETURNING keyword_id",
                        (kw_norm, kw))
                    kw_id = ins[0]["keyword_id"]

                tx.execute(
                    "UPDATE OR INSERT INTO image_keywords (image_id, keyword_id, source, confidence) VALUES (?, ?, ?, ?) MATCHING (image_id, keyword_id)",
                    (image_id, kw_id, source, confidence))

        get_connector().run_transaction(_tx)
    except Exception as e:
        logging.warning(f"_sync_image_keywords failed for image {image_id}: {e}")

def _backfill_keywords():
    """One-time migration to move BLOB keywords to the normalized tables."""
    print("  [2.1c] Backfilling keywords from images...")
    conn = get_db()
    c = conn.cursor()
    try:
        # Check if already backfilled to avoid redundant work
        c.execute("SELECT FIRST 1 1 FROM image_keywords")
        if c.fetchone():
            print("  [2.1c] Keywords already backfilled.")
            return

        c.execute("SELECT id, keywords FROM images WHERE keywords IS NOT NULL AND keywords <> ''")
        rows = c.fetchall()
        for row in rows:
            _sync_image_keywords(row[0], row[1], source="legacy_backfill")
        print(f"  [2.1c] Successfully backfilled keywords for {len(rows)} images.")
    except Exception as e:
        import logging
        logging.error(f"Error backfilling keywords: {e}")
    finally:
        conn.close()

def _backfill_image_xmp():
    """Backfill IMAGE_XMP rows for images that have metadata but no XMP record."""
    print("  [2.6] Backfilling IMAGE_XMP from images...")
    conn = get_db()
    c = conn.cursor()
    try:
        c.execute("""
            SELECT i.id, i.rating, i.label, i.keywords, i.title, i.description
            FROM images i
            LEFT JOIN image_xmp x ON i.id = x.image_id
            WHERE x.image_id IS NULL
              AND (i.rating IS NOT NULL OR i.label IS NOT NULL
                   OR i.keywords IS NOT NULL OR i.title IS NOT NULL
                   OR i.description IS NOT NULL)
        """)
        rows = c.fetchall()

        if not rows:
            print("  [2.6] No images need IMAGE_XMP backfill.")
            return

        count = 0
        for row in rows:
            # RowWrapper.__iter__ yields (key, value) pairs — do not tuple-unpack as columns.
            image_id, rating, label, keywords, title, description = (
                row[0], row[1], row[2], row[3], row[4], row[5]
            )
            c.execute("""
                UPDATE OR INSERT INTO image_xmp
                    (image_id, rating, label, keywords, title, description, extracted_at)
                VALUES (?, ?, ?, ?, ?, ?, CURRENT_TIMESTAMP)
                MATCHING (image_id)
            """, (image_id, rating, label, keywords, title, description))
            count += 1

        conn.commit()
        print(f"  [2.6] Backfilled IMAGE_XMP for {count} images.")
    except Exception as e:
        logging.error(f"Error backfilling IMAGE_XMP: {e}")
        try:
            conn.rollback()
        except Exception:
            pass
    finally:
        conn.close()
